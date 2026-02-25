"""
Build a standalone, formula-driven FTE Baseload Excel Model.
Run this script to generate FTE_Baseload_Model_Live.xlsx.

Single-scenario model: user enters one value per parameter.
The yearly range comes from within-year variation (min/max monthly FTE).
"""

import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

# ── Palette ──────────────────────────────────────────────────────────────
NAVY = "051C2C"
BLUE = "2251FF"
TEAL = "00A9F4"
GREY = "7F8C8D"
LIGHT = "F5F6F7"
WHITE = "FFFFFF"
DARK = "1A1A2E"
YELLOW_INPUT = "FFF9E6"
GREEN_OK = "E8F5E9"

# ── Styles ───────────────────────────────────────────────────────────────
navy_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
light_fill = PatternFill(start_color=LIGHT, end_color=LIGHT, fill_type="solid")
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
input_fill = PatternFill(start_color=YELLOW_INPUT, end_color=YELLOW_INPUT, fill_type="solid")
green_fill = PatternFill(start_color=GREEN_OK, end_color=GREEN_OK, fill_type="solid")

title_font = Font(name="Calibri", size=16, bold=True, color=NAVY)
section_font = Font(name="Calibri", size=12, bold=True, color=NAVY)
label_font = Font(name="Calibri", size=10, color=DARK)
input_font = Font(name="Calibri", size=10, color=BLUE, bold=True)
formula_font = Font(name="Calibri", size=10, color=GREY, italic=True)
hdr_font = Font(name="Calibri", size=10, bold=True, color=WHITE)
body_font = Font(name="Calibri", size=10, color=DARK)
note_font = Font(name="Calibri", size=9, color=GREY, italic=True)
bold_font = Font(name="Calibri", size=10, bold=True, color=DARK)

thin_border = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="thin", color="D0D5DD"),
)

align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Archetype definitions (defaults — midpoints of the original ranges) ──
ARCHETYPES = [
    {
        "name": "Chemistry",
        "share": 0.15,
        "stages": {
            "TRL 1-4": {"dur": 7, "cost": 6.5, "res": 3.5, "dev": 1.5},
            "TRL 5-7": {"dur": 12, "cost": 12.5, "res": 1.5, "dev": 3.5},
        },
    },
    {
        "name": "Process (Hardware)",
        "share": 0.70,
        "stages": {
            "TRL 1-4": {"dur": 9, "cost": 12.5, "res": 6.5, "dev": 1.5},
            "TRL 5-7": {"dur": 15, "cost": 15.0, "res": 1.5, "dev": 6.5},
        },
    },
    {
        "name": "Algorithm (Software)",
        "share": 0.15,
        "stages": {
            "TRL 1-4": {"dur": 6, "cost": 4.25, "res": 0.5, "dev": 0.5},
            "TRL 5-7": {"dur": 6, "cost": 4.25, "res": 0.5, "dev": 0.5},
        },
    },
]


def _style_header_row(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = navy_fill
        cell.font = hdr_font
        cell.alignment = align_center
        cell.border = thin_border


def _style_data_cell(ws, row, col, is_input=False, is_formula=False):
    cell = ws.cell(row=row, column=col)
    cell.border = thin_border
    cell.alignment = Alignment(vertical="center")
    if is_input:
        cell.fill = input_fill
        cell.font = input_font
    elif is_formula:
        cell.fill = green_fill
        cell.font = formula_font
    else:
        cell.font = body_font


# ═════════════════════════════════════════════════════════════════════════
# SHEET 1: INPUTS
# ═════════════════════════════════════════════════════════════════════════
def build_inputs_sheet(wb):
    ws = wb.active
    ws.title = "Inputs"
    ws.sheet_properties.tabColor = BLUE

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 5
    ws.column_dimensions["E"].width = 38
    ws.column_dimensions["F"].width = 16

    ws.merge_cells("B2:F2")
    ws["B2"] = "FTE Baseload Model — Inputs"
    ws["B2"].font = title_font

    ws["B3"] = "Yellow cells = you can change these.  Green cells = calculated by formulas.  All other sheets update automatically."
    ws["B3"].font = note_font

    row = 5

    # ── BUDGET & TIMELINE ──
    ws.cell(row=row, column=2, value="BUDGET & TIMELINE").font = section_font
    row += 1

    labels_budget = [
        ("Total R&D budget (USD millions)", 400, "Budget", True, None),
        ("Overhead deduction (%)", 0.30, "Overhead", True, "0%"),
        ("Net project budget (M)", None, "NetBudget", False, "#,##0"),
        None,
        ("First year of new projects", 2026, "StartYear", False, "0"),
        ("Last year of new projects", 2029, "EndYear", False, "0"),
        ("Intake window (months per year)", 6, "IntakeMonths", True, "0"),
        ("Utilization rate", 1.0, "Utilization", True, "0%"),
    ]

    for item in labels_budget:
        if item is None:
            row += 1
            continue
        lbl, val, named, is_input, fmt = item
        ws.cell(row=row, column=2, value=lbl).font = label_font
        cell = ws.cell(row=row, column=3)

        if named == "NetBudget":
            cell.value = "=Budget*(1-Overhead)"
            _style_data_cell(ws, row, 3, is_formula=True)
        else:
            cell.value = val
            _style_data_cell(ws, row, 3, is_input=True)

        if fmt:
            cell.number_format = fmt

        if named:
            ref = f"Inputs!$C${row}"
            wb.defined_names.add(DefinedName(named, attr_text=ref))

        row += 1

    row += 1

    # ── PIPELINE STAGES ──
    ws.cell(row=row, column=2, value="PIPELINE STAGES").font = section_font
    row += 1

    pipe_items = [
        ("TRL 1-4: % of new projects that start here", 0.20, "Alloc_Early", "0%"),
        ("TRL 1-4: % of completers that advance to TRL 5-7", 0.50, "Conv_Early", "0%"),
        ("TRL 5-7: % of new projects that start here directly", 0.80, "Alloc_Late", "0%"),
    ]
    for lbl, val, named, fmt in pipe_items:
        ws.cell(row=row, column=2, value=lbl).font = label_font
        cell = ws.cell(row=row, column=3, value=val)
        cell.number_format = fmt
        _style_data_cell(ws, row, 3, is_input=True)
        wb.defined_names.add(DefinedName(named, attr_text=f"Inputs!$C${row}"))
        row += 1

    ws.cell(row=row, column=2, value="Total direct allocation (should = 100%)").font = label_font
    ws.cell(row=row, column=3, value="=Alloc_Early+Alloc_Late")
    ws.cell(row=row, column=3).number_format = "0%"
    _style_data_cell(ws, row, 3, is_formula=True)
    row += 2

    # ── PORTFOLIO MIX ──
    ws.cell(row=row, column=2, value="PORTFOLIO MIX").font = section_font
    row += 1
    ws.cell(row=row, column=2, value="What share of your projects fall into each type? Must add to 100%.").font = note_font
    row += 1

    arch_share_names = ["Share_Chem", "Share_HW", "Share_SW"]
    for ai, arch in enumerate(ARCHETYPES):
        ws.cell(row=row, column=2, value=f"{arch['name']} (%)").font = label_font
        cell = ws.cell(row=row, column=3, value=arch["share"])
        cell.number_format = "0%"
        _style_data_cell(ws, row, 3, is_input=True)
        wb.defined_names.add(DefinedName(arch_share_names[ai], attr_text=f"Inputs!$C${row}"))
        row += 1

    ws.cell(row=row, column=2, value="Total portfolio share").font = label_font
    ws.cell(row=row, column=3, value="=Share_Chem+Share_HW+Share_SW")
    ws.cell(row=row, column=3).number_format = "0%"
    _style_data_cell(ws, row, 3, is_formula=True)
    row += 2

    # ── ARCHETYPE PARAMETERS ──
    ws.cell(row=row, column=2, value="PROJECT TYPE PARAMETERS").font = section_font
    row += 1
    ws.cell(row=row, column=2, value="For each project type and stage, set your best estimate for duration, cost, and team size.").font = note_font
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 1

    arch_short = ["Chem", "HW", "SW"]
    stage_short = ["E", "L"]
    stage_names = ["TRL 1-4", "TRL 5-7"]

    for ai, arch in enumerate(ARCHETYPES):
        ws.cell(row=row, column=2, value=arch["name"]).font = Font(name="Calibri", size=11, bold=True, color=NAVY)
        row += 1

        for si, sname in enumerate(stage_names):
            prefix = f"{arch_short[ai]}_{stage_short[si]}"
            sp = arch["stages"][sname]

            ws.cell(row=row, column=2, value=f"  {sname}").font = Font(name="Calibri", size=10, bold=True, color=GREY)
            row += 1

            params = [
                ("Duration (months)", "Dur", sp["dur"], "0"),
                ("Cost per project (M)", "Cost", sp["cost"], "#,##0.0"),
                ("Research FTE per project", "Res", sp["res"], "0.0"),
                ("Developer FTE per project", "Dev", sp["dev"], "0.0"),
            ]

            _style_header_row(ws, row, 2, 3)
            ws.cell(row=row, column=2, value="Parameter")
            ws.cell(row=row, column=3, value="Value")
            row += 1

            for plbl, pshort, pval, pfmt in params:
                ws.cell(row=row, column=2, value=f"    {plbl}").font = label_font
                ws.cell(row=row, column=2).border = thin_border

                name = f"{prefix}_{pshort}"

                cell = ws.cell(row=row, column=3, value=pval)
                cell.number_format = pfmt
                _style_data_cell(ws, row, 3, is_input=True)
                wb.defined_names.add(DefinedName(name, attr_text=f"Inputs!$C${row}"))

                row += 1

            row += 1

    row += 1

    # ── DERIVED VALUES ──
    ws.cell(row=row, column=2, value="DERIVED VALUES (formulas — do not edit)").font = section_font
    row += 1
    ws.cell(row=row, column=2, value="These are calculated from the inputs above. They show how the model converts budget into project count.").font = note_font
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    row += 1

    _style_header_row(ws, row, 2, 4)
    ws.cell(row=row, column=2, value="Metric")
    ws.cell(row=row, column=3, value="Value")
    ws.cell(row=row, column=4, value="How calculated")
    ws.column_dimensions["D"].width = 55
    row += 1

    # Weighted cost per project
    # Expected cost entering early = cost_early + conv * cost_late
    # Expected cost entering late = cost_late
    # Archetype cost = alloc_early * expected_early + alloc_late * expected_late
    # Portfolio cost = sum(share_i * arch_cost_i)
    wc_formula = (
        "=Share_Chem*(Alloc_Early*(Chem_E_Cost + Conv_Early*Chem_L_Cost) + Alloc_Late*Chem_L_Cost)"
        "+Share_HW*(Alloc_Early*(HW_E_Cost + Conv_Early*HW_L_Cost) + Alloc_Late*HW_L_Cost)"
        "+Share_SW*(Alloc_Early*(SW_E_Cost + Conv_Early*SW_L_Cost) + Alloc_Late*SW_L_Cost)"
    )

    ws.cell(row=row, column=2, value="Weighted cost per project (M)").font = label_font
    ws.cell(row=row, column=2).border = thin_border
    c_wc = ws.cell(row=row, column=3, value=wc_formula)
    c_wc.number_format = "#,##0.0"
    _style_data_cell(ws, row, 3, is_formula=True)
    wb.defined_names.add(DefinedName("WtdCost", attr_text=f"Inputs!$C${row}"))

    ws.cell(row=row, column=4, value="Portfolio-weighted average cost, accounting for stage mix and conversion").font = note_font
    row += 1

    ws.cell(row=row, column=2, value="Projects per year").font = label_font
    ws.cell(row=row, column=2).border = thin_border
    c_pp = ws.cell(row=row, column=3, value="=IF(WtdCost>0, NetBudget/WtdCost, 0)")
    c_pp.number_format = "#,##0.0"
    _style_data_cell(ws, row, 3, is_formula=True)
    wb.defined_names.add(DefinedName("ProjPerYr", attr_text=f"Inputs!$C${row}"))

    ws.cell(row=row, column=4, value="Net budget ÷ weighted cost per project").font = note_font
    row += 1

    return ws


# ═════════════════════════════════════════════════════════════════════════
# SHEET 2: HOW THIS MODEL WORKS
# ═════════════════════════════════════════════════════════════════════════
def build_how_it_works_sheet(wb):
    ws = wb.create_sheet("How This Model Works", 0)
    ws.sheet_properties.tabColor = NAVY

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 100

    ws["B2"] = "FTE Baseload Model — How It Works"
    ws["B2"].font = title_font

    content = [
        ("WHAT THIS MODEL DOES", section_font),
        ("It answers: 'Given our R&D budget and project portfolio, how many researchers and developers do we need?'", label_font),
        ("You provide your best estimates for cost, duration, and team size. The model calculates a single headcount figure.", label_font),
        ("The yearly range (min–max) reflects natural within-year variation as projects start, overlap, and complete.", label_font),
        ("", None),
        ("HOW IT CALCULATES HEADCOUNT — 5 STEPS", section_font),
        ("", None),
        ("Step 1: Start with the money", Font(name="Calibri", size=10, bold=True, color=NAVY)),
        ("Total R&D budget minus overhead = net project budget.", label_font),
        ("", None),
        ("Step 2: Figure out how many projects you can afford", Font(name="Calibri", size=10, bold=True, color=NAVY)),
        ("Net budget ÷ weighted average cost per project = projects per year.", label_font),
        ("The cost is weighted by your portfolio mix and which stages projects go through.", label_font),
        ("", None),
        ("Step 3: Distribute projects across types and stages", Font(name="Calibri", size=10, bold=True, color=NAVY)),
        ("Projects are split across archetypes (Chemistry, Hardware, Software) by portfolio share.", label_font),
        ("Within each type, some start at TRL 1-4 (early) and some start directly at TRL 5-7 (late).", label_font),
        ("When early-stage projects finish, a percentage advance to TRL 5-7 as additional projects.", label_font),
        ("", None),
        ("Step 4: Simulate the pipeline month by month", Font(name="Calibri", size=10, bold=True, color=NAVY)),
        ("Each year's new projects are spread across the first N months (the intake window).", label_font),
        ("A project stays 'active' from its start month through start + duration months.", label_font),
        ("The Engine sheet tracks how many projects are running in each stage, every month.", label_font),
        ("", None),
        ("Step 5: Convert active projects into headcount", Font(name="Calibri", size=10, bold=True, color=NAVY)),
        ("Active projects × staff per project = FTE needed that month.", label_font),
        ("If utilization < 100%, the model inflates by 1 ÷ utilization to get gross headcount.", label_font),
        ("", None),
        ("UNDERSTANDING STEADY STATE AND THE YEARLY RANGE", section_font),
        ("", None),
        ("Why does headcount grow at first?", Font(name="Calibri", size=10, bold=True, color=NAVY)),
        ("In Year 1, projects start but none have finished yet — the pipeline only fills up.", label_font),
        ("In Year 2, new projects start while Year 1 projects are still running. Headcount keeps climbing", label_font),
        ("until the rate of new starts roughly equals the rate of completions.", label_font),
        ("Once that happens, headcount stabilizes — this is the STEADY STATE.", label_font),
        ("The steady-state headcount is the long-run staffing level your hiring plan should target.", label_font),
        ("", None),
        ("Why is there a range within each year?", Font(name="Calibri", size=10, bold=True, color=NAVY)),
        ("Because new projects start during an intake window (not all at once), FTE demand varies month to month.", label_font),
        ("The Summary sheet shows:", label_font),
        ("  • Avg monthly FTE — the steady-state staffing level for that year", label_font),
        ("  • Min monthly FTE — the quietest month (e.g. just before a new cohort starts)", label_font),
        ("  • Max monthly FTE — the busiest month (e.g. when old and new cohorts overlap most)", label_font),
        ("Early years have a wide range (pipeline is still filling). Later years converge (pipeline has stabilized).", label_font),
        ("", None),
        ("KEY TERMS", section_font),
        ("FTE = Full-Time Equivalent. 1 FTE = one person working full time.", label_font),
        ("TRL = Technology Readiness Level. A 1-to-9 scale of technology maturity.", label_font),
        ("Pipeline = The sequence of stages a project goes through.", label_font),
        ("Archetype = A type of R&D project (e.g. Chemistry, Hardware, Software).", label_font),
        ("Conversion rate = % of projects finishing one stage that advance to the next.", label_font),
        ("", None),
        ("THINGS THIS EXCEL ASSUMES THAT CANNOT BE CHANGED", section_font),
        ("(Changing these would require rebuilding the sheet structure)", note_font),
        ("", None),
        ("• 2 pipeline stages (TRL 1-4 and TRL 5-7)", label_font),
        ("• 3 project types (Chemistry, Process Hardware, Algorithm Software)", label_font),
        ("• 2 FTE roles (Research and Developer)", label_font),
        ("• Monthly granularity — projects tracked month by month", label_font),
        ("• No ramp-up — projects start at full staffing immediately", label_font),
        ("• No mid-stage cancellation — projects run to completion", label_font),
        ("• Same budget every year across the planning horizon", label_font),
        ("• Projects spread evenly across the intake window", label_font),
        ("• Linear pipeline — no branching or looping between stages", label_font),
        ("• No economies of scale — cost per project is constant regardless of volume", label_font),
        ("", None),
        ("SHEET GUIDE", section_font),
        ("Inputs — the only sheet you need to edit. All yellow cells are changeable.", label_font),
        ("Engine — monthly calculations for all archetypes and stages. All formulas.", label_font),
        ("Summary — annual averages and within-year range from the engine. All formulas.", label_font),
    ]

    row = 4
    for text, font in content:
        if text or font:
            cell = ws.cell(row=row, column=2, value=text)
            if font:
                cell.font = font
        row += 1

    return ws


# ═════════════════════════════════════════════════════════════════════════
# SHEET 3: ENGINE (single scenario)
# ═════════════════════════════════════════════════════════════════════════
def build_engine_sheet(wb):
    """Build a single engine sheet using user's direct input values."""
    ws = wb.create_sheet("Engine")
    ws.sheet_properties.tabColor = BLUE

    arch_short = ["Chem", "HW", "SW"]
    share_names = ["Share_Chem", "Share_HW", "Share_SW"]

    n_months = 60  # 5 years (2026–2030) — enough for project tails past EndYear
    data_start_row = 3

    cols_per_arch = 9
    arch_start_col = 4

    # Headers row 1: group headers
    for ai, arch in enumerate(ARCHETYPES):
        start_c = arch_start_col + ai * cols_per_arch
        ws.merge_cells(start_row=1, start_column=start_c, end_row=1, end_column=start_c + cols_per_arch - 1)
        ws.cell(row=1, column=start_c, value=arch["name"]).font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        ws.cell(row=1, column=start_c).fill = navy_fill
        ws.cell(row=1, column=start_c).alignment = align_center
        for c in range(start_c, start_c + cols_per_arch):
            ws.cell(row=1, column=c).fill = navy_fill

    totals_col = arch_start_col + 3 * cols_per_arch
    ws.merge_cells(start_row=1, start_column=totals_col, end_row=1, end_column=totals_col + 2)
    ws.cell(row=1, column=totals_col, value="GRAND TOTALS").font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    ws.cell(row=1, column=totals_col).fill = navy_fill
    ws.cell(row=1, column=totals_col).alignment = align_center
    for c in range(totals_col, totals_col + 3):
        ws.cell(row=1, column=c).fill = navy_fill

    # Headers row 2
    _style_header_row(ws, 2, 1, totals_col + 2)
    ws.cell(row=2, column=1, value="Date")
    ws.cell(row=2, column=2, value="Year")
    ws.cell(row=2, column=3, value="Month")

    sub_headers = [
        "Early Starts/mo", "Early Active",
        "Late Conv Starts", "Late Direct Starts", "Late Total Starts", "Late Active",
        "Research FTE", "Developer FTE", "Total FTE",
    ]
    for ai in range(3):
        sc = arch_start_col + ai * cols_per_arch
        for si, sh in enumerate(sub_headers):
            ws.cell(row=2, column=sc + si, value=sh)

    ws.cell(row=2, column=totals_col, value="Total Research")
    ws.cell(row=2, column=totals_col + 1, value="Total Developer")
    ws.cell(row=2, column=totals_col + 2, value="Total FTE")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 5
    for c in range(4, totals_col + 3):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # ── Data rows ──
    ds = data_start_row
    for r_offset in range(n_months):
        r = ds + r_offset

        ws.cell(row=r, column=1, value=f"=DATE(StartYear + INT({r_offset}/12), MOD({r_offset},12)+1, 1)")
        ws.cell(row=r, column=1).number_format = "YYYY-MM"
        ws.cell(row=r, column=2, value=f"=YEAR(A{r})")
        ws.cell(row=r, column=3, value=f"=MONTH(A{r})")

        for ai in range(3):
            sc = arch_start_col + ai * cols_per_arch
            ashort = arch_short[ai]
            share = share_names[ai]

            early_starts_col = get_column_letter(sc)
            early_active_col = get_column_letter(sc + 1)
            late_conv_col = get_column_letter(sc + 2)
            late_direct_col = get_column_letter(sc + 3)
            late_total_col = get_column_letter(sc + 4)
            late_active_col = get_column_letter(sc + 5)
            res_col = get_column_letter(sc + 6)
            dev_col = get_column_letter(sc + 7)
            tot_col = get_column_letter(sc + 8)

            dur_early = f"{ashort}_E_Dur"
            dur_late = f"{ashort}_L_Dur"
            res_early = f"{ashort}_E_Res"
            dev_early = f"{ashort}_E_Dev"
            res_late = f"{ashort}_L_Res"
            dev_late = f"{ashort}_L_Dev"

            # Early Direct Starts per month
            ws.cell(row=r, column=sc).value = (
                f"=IF(AND(B{r}>=StartYear, B{r}<=EndYear, C{r}<=IntakeMonths),"
                f" ProjPerYr*{share}*Alloc_Early/IntakeMonths, 0)"
            )
            ws.cell(row=r, column=sc).number_format = "0.00"

            # Early Active Stock (sliding window)
            if r_offset == 0:
                ws.cell(row=r, column=sc + 1).value = f"={early_starts_col}{r}"
            else:
                ws.cell(row=r, column=sc + 1).value = (
                    f"=IF({r}-{ds} < {dur_early},"
                    f" SUM({early_starts_col}${ds}:{early_starts_col}{r}),"
                    f" SUM(OFFSET({early_starts_col}{r},-{dur_early}+1,0,{dur_early},1)))"
                )
            ws.cell(row=r, column=sc + 1).number_format = "0.00"

            # Late Conversion Starts (early completions × conv rate)
            ws.cell(row=r, column=sc + 2).value = (
                f"=IF({r}-{ds} >= {dur_early},"
                f" OFFSET({early_starts_col}{r}, -{dur_early}, 0) * Conv_Early, 0)"
            )
            ws.cell(row=r, column=sc + 2).number_format = "0.00"

            # Late Direct Starts
            ws.cell(row=r, column=sc + 3).value = (
                f"=IF(AND(B{r}>=StartYear, B{r}<=EndYear, C{r}<=IntakeMonths),"
                f" ProjPerYr*{share}*Alloc_Late/IntakeMonths, 0)"
            )
            ws.cell(row=r, column=sc + 3).number_format = "0.00"

            # Late Total Starts
            ws.cell(row=r, column=sc + 4).value = f"={late_conv_col}{r}+{late_direct_col}{r}"
            ws.cell(row=r, column=sc + 4).number_format = "0.00"

            # Late Active Stock (sliding window)
            if r_offset == 0:
                ws.cell(row=r, column=sc + 5).value = f"={late_total_col}{r}"
            else:
                ws.cell(row=r, column=sc + 5).value = (
                    f"=IF({r}-{ds} < {dur_late},"
                    f" SUM({late_total_col}${ds}:{late_total_col}{r}),"
                    f" SUM(OFFSET({late_total_col}{r},-{dur_late}+1,0,{dur_late},1)))"
                )
            ws.cell(row=r, column=sc + 5).number_format = "0.00"

            # FTE
            util = "Utilization"
            ws.cell(row=r, column=sc + 6).value = (
                f"=({early_active_col}{r}*{res_early} + {late_active_col}{r}*{res_late}) / {util}"
            )
            ws.cell(row=r, column=sc + 6).number_format = "0.0"

            ws.cell(row=r, column=sc + 7).value = (
                f"=({early_active_col}{r}*{dev_early} + {late_active_col}{r}*{dev_late}) / {util}"
            )
            ws.cell(row=r, column=sc + 7).number_format = "0.0"

            ws.cell(row=r, column=sc + 8).value = f"={res_col}{r}+{dev_col}{r}"
            ws.cell(row=r, column=sc + 8).number_format = "0.0"

        # Grand Totals
        res_cols = [get_column_letter(arch_start_col + ai * cols_per_arch + 6) for ai in range(3)]
        dev_cols = [get_column_letter(arch_start_col + ai * cols_per_arch + 7) for ai in range(3)]

        ws.cell(row=r, column=totals_col).value = f"={'+'.join(f'{c}{r}' for c in res_cols)}"
        ws.cell(row=r, column=totals_col).number_format = "0.0"
        ws.cell(row=r, column=totals_col + 1).value = f"={'+'.join(f'{c}{r}' for c in dev_cols)}"
        ws.cell(row=r, column=totals_col + 1).number_format = "0.0"
        tc = get_column_letter(totals_col)
        tc1 = get_column_letter(totals_col + 1)
        ws.cell(row=r, column=totals_col + 2).value = f"={tc}{r}+{tc1}{r}"
        ws.cell(row=r, column=totals_col + 2).number_format = "0.0"

    # Hide intermediate columns (Early Starts, Active, Conv, Direct, Total Starts, Late Active)
    # Keep only Research FTE, Developer FTE, Total FTE visible per archetype
    for ai in range(3):
        sc = arch_start_col + ai * cols_per_arch
        for offset in range(6):  # first 6 sub-columns are intermediate
            col_letter = get_column_letter(sc + offset)
            ws.column_dimensions[col_letter].hidden = True

    return ws, totals_col


# ═════════════════════════════════════════════════════════════════════════
# SHEET 4: SUMMARY
# ═════════════════════════════════════════════════════════════════════════
def build_summary_sheet(wb, totals_col):
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = NAVY

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 10
    for c_letter in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[c_letter].width = 18

    ws["B2"] = "Annual FTE Summary"
    ws["B2"].font = title_font
    ws["B3"] = "Avg = average across all months. Min/Max = lowest and highest single-month FTE that year."
    ws["B3"].font = note_font

    _style_header_row(ws, 5, 2, 7)
    headers = ["Year", "Avg monthly FTE", "Min monthly FTE", "Max monthly FTE", "Avg Research FTE", "Avg Developer FTE"]
    for ci, h in enumerate(headers):
        ws.cell(row=5, column=2 + ci, value=h)

    tc_total = get_column_letter(totals_col + 2)
    tc_res = get_column_letter(totals_col)
    tc_dev = get_column_letter(totals_col + 1)

    ds = 3
    de = ds + 60 - 1

    n_years = 2029 - 2026 + 1  # EndYear - StartYear + 1
    for yi in range(n_years):
        r = 6 + yi
        ws.cell(row=r, column=2, value=f"=StartYear+{yi}")
        ws.cell(row=r, column=2).number_format = "0"
        ws.cell(row=r, column=2).font = bold_font
        ws.cell(row=r, column=2).border = thin_border

        year_ref = f"B{r}"

        # Avg monthly FTE (exclude months with zero FTE)
        ws.cell(row=r, column=3).value = (
            f"=IFERROR(AVERAGEIFS(Engine!{tc_total}${ds}:{tc_total}${de},"
            f"Engine!$B${ds}:$B${de},{year_ref},"
            f"Engine!{tc_total}${ds}:{tc_total}${de},\">0\"), 0)"
        )
        ws.cell(row=r, column=3).number_format = "0.0"
        _style_data_cell(ws, r, 3, is_formula=True)

        # Min monthly FTE (exclude months with zero FTE)
        ws.cell(row=r, column=4).value = (
            f"=IFERROR(_xlfn.MINIFS(Engine!{tc_total}${ds}:{tc_total}${de},"
            f"Engine!$B${ds}:$B${de},{year_ref},"
            f"Engine!{tc_total}${ds}:{tc_total}${de},\">0\"), 0)"
        )
        ws.cell(row=r, column=4).number_format = "0.0"
        _style_data_cell(ws, r, 4, is_formula=True)

        # Max monthly FTE (exclude months with zero FTE)
        ws.cell(row=r, column=5).value = (
            f"=IFERROR(_xlfn.MAXIFS(Engine!{tc_total}${ds}:{tc_total}${de},"
            f"Engine!$B${ds}:$B${de},{year_ref},"
            f"Engine!{tc_total}${ds}:{tc_total}${de},\">0\"), 0)"
        )
        ws.cell(row=r, column=5).number_format = "0.0"
        _style_data_cell(ws, r, 5, is_formula=True)

        # Avg Research FTE (exclude months with zero total FTE)
        ws.cell(row=r, column=6).value = (
            f"=IFERROR(AVERAGEIFS(Engine!{tc_res}${ds}:{tc_res}${de},"
            f"Engine!$B${ds}:$B${de},{year_ref},"
            f"Engine!{tc_total}${ds}:{tc_total}${de},\">0\"), 0)"
        )
        ws.cell(row=r, column=6).number_format = "0.0"
        _style_data_cell(ws, r, 6, is_formula=True)

        # Avg Developer FTE (exclude months with zero total FTE)
        ws.cell(row=r, column=7).value = (
            f"=IFERROR(AVERAGEIFS(Engine!{tc_dev}${ds}:{tc_dev}${de},"
            f"Engine!$B${ds}:$B${de},{year_ref},"
            f"Engine!{tc_total}${ds}:{tc_total}${de},\">0\"), 0)"
        )
        ws.cell(row=r, column=7).number_format = "0.0"
        _style_data_cell(ws, r, 7, is_formula=True)

    return ws


# ═════════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════════
def main():
    wb = Workbook()

    print("Building Inputs sheet...")
    build_inputs_sheet(wb)

    print("Building How This Model Works sheet...")
    build_how_it_works_sheet(wb)

    print("Building Engine sheet...")
    _, totals_col = build_engine_sheet(wb)

    print("Building Summary sheet...")
    build_summary_sheet(wb, totals_col)

    desired_order = [
        "How This Model Works",
        "Inputs",
        "Engine",
        "Summary",
    ]
    sheet_indices = {ws.title: i for i, ws in enumerate(wb.worksheets)}
    new_order = [sheet_indices[name] for name in desired_order]
    wb._sheets = [wb.worksheets[i] for i in new_order]

    out_path = r"c:\Users\Debdoot Ray\genAI training\fte_model\FTE_Baseload_Model_Live.xlsx"
    wb.save(out_path)
    print(f"\nSaved to: {out_path}")
    print("Open in Excel and try changing a yellow cell on the Inputs sheet!")


if __name__ == "__main__":
    main()

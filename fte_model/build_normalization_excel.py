"""Build v5 normalization Excel workbook — Staffing Norms methodology.

Core idea: extract peer PY/MYR norms from Shell+Chevron, apply to
PETRONAS portfolio, compute implied vs actual person-years per bucket.
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from decimal import Decimal, ROUND_HALF_UP
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path


def round_half_up(val, places):
    return float(Decimal(str(val)).quantize(Decimal(10) ** -places, rounding=ROUND_HALF_UP))

# ══════════════════════════════════════════════════════════════════
# Colors and styles
# ══════════════════════════════════════════════════════════════════
TEAL = "003A6C"
DARK_BLUE = "2E86C1"
DARK_GREEN = "27AE60"
ORANGE = "E67E22"
WHITE = "FFFFFF"
LIGHT_YELLOW = "FFF9E6"
LIGHT_GREY = "F2F2F2"
LIGHT_BLUE = "E8F0FE"
LIGHT_GREEN = "E6F5E6"
LIGHT_RED = "FDE8E8"
LIGHT_ORANGE = "FEF3E2"

title_font = Font(name="Calibri", size=14, bold=True, color=TEAL)
header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
section_font = Font(name="Calibri", size=11, bold=True, color=TEAL)
normal_font = Font(name="Calibri", size=11)
bold_font = Font(name="Calibri", size=11, bold=True)
small_font = Font(name="Calibri", size=10, italic=True, color="666666")
input_font = Font(name="Calibri", size=11, bold=True, color="003A6C")

header_fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
input_fill = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")
calc_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
result_fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
caveat_fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
grey_fill = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
orange_fill = PatternFill(start_color=LIGHT_ORANGE, end_color=LIGHT_ORANGE, fill_type="solid")

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def cell(ws, row, col, value, font=None, fill=None, align=None, fmt=None, border=True):
    c = ws.cell(row=row, column=col, value=value)
    if font:
        c.font = font
    if fill:
        c.fill = fill
    c.alignment = align if align else center
    if fmt:
        c.number_format = fmt
    if border:
        c.border = thin_border
    return c


def header_row(ws, row, col_start, labels):
    for i, label in enumerate(labels):
        cell(ws, row, col_start + i, label, font=header_font, fill=header_fill)


def merge_title(ws, row, col_start, col_end, text, font=title_font):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.font = font
    c.alignment = Alignment(horizontal="left", vertical="center")


def main():
    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════
    # Data
    # ══════════════════════════════════════════════════════════════════
    ref_projects = {
        ("Chemistry", "TRL 1-4"): [
            ("Fgo", 36, 5, 2.4),
            ("Garcinia", 36, 4, 9.8),
        ],
        ("Chemistry", "TRL 5-7"): [
            ("Fgo", 60, 8, 2.7),
            ("WIF", 24, 8, 5.1),
            ("PMB", 24, 3, 17.9),
        ],
        ("HW Process", "TRL 1-4"): [
            ("Methane Pyrolisis", 24, 1, 3.8),
            ("Nanobubble", 8, 2, 1.3),
            ("Bio Meg", 84, 4, 19.2),
        ],
        ("HW Process", "TRL 5-7"): [
            ("Methane Pyrolisis", 48, 11, 31.4),
            ("Nanobubble", 60, 5, 15.6),
            ("Garcinia", 72, 5, 42.1),
            ("Bio Meg", 96, 7, 109.0),
        ],
        ("HW Mechanical", "TRL 1-4"): [
            ("Elektra", 24, 1, 5.7),
            ("ASAT", 36, 3, 15.0),
        ],
        ("HW Mechanical", "TRL 5-7"): [
            ("Elektra", 12, 1, 4.9),
            ("ASAT", 24, 2, 1.8),
        ],
        ("Algorithm", "TRL 1-4"): [
            ("Zod", 12, 1, 3.0),
            ("Reseis", 48, 4, 4.0),
        ],
        ("Algorithm", "TRL 5-7"): [
            ("Zod", 12, 1, 3.9),
            ("Reseis", 24, 2, 2.0),
        ],
    }

    petronas_projects = {}
    for (arch, stage), projects in ref_projects.items():
        durs = [p[1] for p in projects]
        ftes = [p[2] for p in projects]
        pys = [p[2] * (p[1] / 12.0) for p in projects]
        costs = [p[3] for p in projects]
        n = len(projects)
        petronas_projects[(arch, stage)] = {
            "dur": round(sum(durs) / n),
            "fte": round_half_up(sum(ftes) / n, 1),
            "py": round_half_up(sum(pys) / n, 1),
            "cost_myr": round_half_up(sum(costs) / n, 1),
        }

    shell_projects = {
        ("Chemistry", "TRL 1-4"):     {"dur": 48, "fte": 6.5,  "py": 26.0,   "cost_myr": 40},
        ("Chemistry", "TRL 5-7"):     {"dur": 72, "fte": 10,   "py": 60.0,   "cost_myr": 200},
        ("HW Mechanical", "TRL 1-4"): {"dur": 48, "fte": 12,   "py": 48.0,   "cost_myr": 40},
        ("HW Mechanical", "TRL 5-7"): {"dur": 72, "fte": 17.5, "py": 105.0,  "cost_myr": 100},
        ("HW Process", "TRL 1-4"):    {"dur": 60, "fte": 6.5,  "py": 32.5,   "cost_myr": 80},
        ("HW Process", "TRL 5-7"):    {"dur": 78, "fte": 15,   "py": 97.5,   "cost_myr": 200},
        ("Algorithm", "TRL 1-4"):     {"dur": 18, "fte": 4.5,  "py": 6.8,    "cost_myr": 8},
        ("Algorithm", "TRL 5-7"):     {"dur": 18, "fte": 10,   "py": 15.0,   "cost_myr": 24},
    }

    chevron_projects = {
        ("Chemistry", "TRL 1-4"):     {"dur": 48, "fte": 2.5,  "py": 10.0,   "cost_myr": 6},
        ("Chemistry", "TRL 5-7"):     {"dur": 72, "fte": 15.5, "py": 93.0,   "cost_myr": 140},
        ("HW Mechanical", "TRL 1-4"): {"dur": 48, "fte": 8,    "py": 32.0,   "cost_myr": 6},
        ("HW Mechanical", "TRL 5-7"): {"dur": 72, "fte": 27.5, "py": 165.0,  "cost_myr": 140},
        ("HW Process", "TRL 1-4"):    {"dur": 60, "fte": 8.5,  "py": 42.5,   "cost_myr": 6},
        ("HW Process", "TRL 5-7"):    {"dur": 90, "fte": 25,   "py": 187.5,  "cost_myr": 300},
        ("Algorithm", "TRL 1-4"):     {"dur": 18, "fte": 3.5,  "py": 5.25,   "cost_myr": 6},
        ("Algorithm", "TRL 5-7"):     {"dur": 18, "fte": 7.5,  "py": 11.25,  "cost_myr": 30},
    }

    basf_projects = {
        ("Chemistry", "TRL 1-4"):     {"dur": 48, "fte": 5,    "py": 20.0,   "cost_myr": 6},
        ("Chemistry", "TRL 5-7"):     {"dur": 36, "fte": 27.5, "py": 82.5,   "cost_myr": 100},
        ("HW Mechanical", "TRL 1-4"): {"dur": 48, "fte": 5,    "py": 20.0,   "cost_myr": 6},
        ("HW Mechanical", "TRL 5-7"): {"dur": 36, "fte": 27.5, "py": 82.5,   "cost_myr": 40},
        ("HW Process", "TRL 1-4"):    {"dur": 48, "fte": 10,   "py": 40.0,   "cost_myr": 30},
        ("HW Process", "TRL 5-7"):    {"dur": 54, "fte": 60,   "py": 270.0,  "cost_myr": 300},
        ("Algorithm", "TRL 1-4"):     {"dur": 18, "fte": 4.5,  "py": 6.8,    "cost_myr": 8},
        ("Algorithm", "TRL 5-7"):     {"dur": 18, "fte": 10,   "py": 15.0,   "cost_myr": 24},
    }

    arch_order = ["Chemistry", "HW Mechanical", "HW Process", "Algorithm"]
    stages = ["TRL 1-4", "TRL 5-7"]
    all_companies = [
        ("PETRONAS", petronas_projects),
        ("Shell",    shell_projects),
        ("Chevron",  chevron_projects),
        ("BASF",     basf_projects),
    ]

    # ══════════════════════════════════════════════════════════════════
    # Sheet 1: Inputs
    # ══════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Inputs"
    ws1.sheet_properties.tabColor = TEAL

    for col in range(1, 10):
        ws1.column_dimensions[get_column_letter(col)].width = 22

    r = 1
    merge_title(ws1, r, 1, 8, "R&D Staffing Norms \u2014 Inputs")
    r += 1
    cell(ws1, r, 1, "Yellow cells are editable. Change them to run your own scenarios.",
         font=small_font, fill=None, align=left_wrap, border=False)
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

    # --- Global assumptions ---
    r += 2
    cell(ws1, r, 1, "GLOBAL ASSUMPTIONS", font=section_font, align=left_wrap, border=False)

    r += 1
    cell(ws1, r, 1, "Parameter", font=bold_font, fill=grey_fill, align=left_wrap)
    cell(ws1, r, 2, "Value", font=bold_font, fill=grey_fill)
    cell(ws1, r, 3, "Unit", font=bold_font, fill=grey_fill)
    cell(ws1, r, 4, "Notes", font=bold_font, fill=grey_fill)
    ws1.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)

    params = [
        ("FX Rate", 4.0, "MYR per USD", "Reference \u2014 not used in staffing norms calculation"),
        ("PETRONAS FTE Cost", 80000, "USD / year", "Reference \u2014 annual cost per researcher"),
        ("Shell FTE Cost", 200000, "USD / year", "Reference \u2014 annual cost per researcher"),
        ("Chevron FTE Cost", 200000, "USD / year", "Reference \u2014 annual cost per researcher"),
        ("BASF FTE Cost", 120000, "USD / year", "Reference \u2014 includes technicians at lower cost"),
        ("Total Active Projects", 160, "Count", "Total active R&D projects in PETRONAS portfolio"),
        ("TRL 1-4 Share", 0.50, "Fraction", "Share of projects in TRL 1-4 stage (rest = TRL 5-7)"),
    ]

    for label, val, unit, note in params:
        r += 1
        cell(ws1, r, 1, label, font=normal_font, align=left_wrap)
        cell(ws1, r, 2, val, font=input_font, fill=input_fill,
             fmt="#,##0.00" if isinstance(val, float) and val < 1 else "#,##0")
        cell(ws1, r, 3, unit, font=normal_font, align=left_wrap)
        cell(ws1, r, 4, note, font=small_font, align=left_wrap)
        ws1.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)

    fx_row = 6
    projects_row = 11
    trl_split_row = 12

    # --- Archetype shares ---
    r += 2
    cell(ws1, r, 1, "ARCHETYPE SHARES", font=section_font, align=left_wrap, border=False)

    r += 1
    cell(ws1, r, 1, "Archetype", font=bold_font, fill=grey_fill, align=left_wrap)
    cell(ws1, r, 2, "Share", font=bold_font, fill=grey_fill)

    archetypes = [
        ("Chemistry", 0.40),
        ("HW Mechanical", 0.18),
        ("HW Process", 0.25),
        ("Algorithm", 0.17),
    ]

    share_start_row = r + 1
    for aname, share in archetypes:
        r += 1
        cell(ws1, r, 1, aname, font=normal_font, align=left_wrap)
        cell(ws1, r, 2, share, font=input_font, fill=input_fill, fmt="0%")
    share_end_row = r

    r += 1
    cell(ws1, r, 1, "Total", font=bold_font, align=left_wrap)
    cell(ws1, r, 2, None, font=bold_font, fill=calc_fill, fmt="0%")
    ws1[f"B{r}"] = f"=SUM(B{share_start_row}:B{share_end_row})"

    # --- Portfolio breakdown (derived) ---
    r += 2
    cell(ws1, r, 1, "PORTFOLIO BREAKDOWN (auto-calculated)", font=section_font, align=left_wrap, border=False)
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

    r += 1
    cell(ws1, r, 1, "Archetype", font=bold_font, fill=grey_fill, align=left_wrap)
    cell(ws1, r, 2, "TRL Stage", font=bold_font, fill=grey_fill)
    cell(ws1, r, 3, "# Projects", font=bold_font, fill=grey_fill)

    breakdown_rows = {}
    for arch_idx, (aname, _) in enumerate(archetypes):
        for stage in stages:
            r += 1
            cell(ws1, r, 1, aname, font=normal_font, align=left_wrap)
            cell(ws1, r, 2, stage, font=normal_font)
            cell(ws1, r, 3, None, font=bold_font, fill=calc_fill, fmt="#,##0.0")
            share_ref = f"B{share_start_row + arch_idx}"
            if stage == "TRL 1-4":
                ws1[f"C{r}"] = f"=B{projects_row}*{share_ref}*B{trl_split_row}"
            else:
                ws1[f"C{r}"] = f"=B{projects_row}*{share_ref}*(1-B{trl_split_row})"
            breakdown_rows[(aname, stage)] = r

    bd_start = min(breakdown_rows.values())
    bd_end = max(breakdown_rows.values())
    r += 1
    cell(ws1, r, 1, "Total", font=bold_font, align=left_wrap)
    cell(ws1, r, 2, "", font=normal_font)
    cell(ws1, r, 3, None, font=bold_font, fill=calc_fill, fmt="#,##0.0")
    ws1[f"C{r}"] = f"=SUM(C{bd_start}:C{bd_end})"

    # ══════════════════════════════════════════════════════════════════
    # Sheet 2: Staffing Norms
    # ══════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Staffing Norms")
    ws2.sheet_properties.tabColor = DARK_BLUE

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 10
    for col in range(3, 16):
        ws2.column_dimensions[get_column_letter(col)].width = 14
    MCOLS2 = 15

    r = 1
    merge_title(ws2, r, 1, MCOLS2, "Staffing Norms \u2014 Peer-Implied Demand Sizing")
    r += 1
    cell(ws2, r, 1,
         "Apply Shell+Chevron staffing norms (PY/MYR M) to PETRONAS portfolio to compute implied demand.",
         font=small_font, fill=None, align=left_wrap, border=False)
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MCOLS2)

    # ------------------------------------------------------------------
    # Section A: Peer Staffing Norms
    # ------------------------------------------------------------------
    r += 2
    cell(ws2, r, 1, "A. PEER STAFFING NORMS (PY per MYR M of project cost)",
         font=section_font, align=left_wrap, border=False)
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MCOLS2)
    r += 1
    cell(ws2, r, 1,
         "PETRONAS: averaged from actual projects (see Reference Data). Peers: outside-in benchmark estimates. PY/MYR = PY / Cost. Higher = more effort per MYR M.",
         font=small_font, fill=None, align=left_wrap, border=False)
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MCOLS2)

    r += 1
    header_row(ws2, r, 1, [
        "Archetype", "TRL Stage",
        "PET PY", "PET Cost (MYR M)", "PET PY/MYR",
        "Shell PY", "Shell Cost (MYR M)", "Shell PY/MYR",
        "Chev PY", "Chev Cost (MYR M)", "Chev PY/MYR",
        "BASF PY (ref)", "BASF Cost (ref)", "BASF PY/MYR (ref)",
        "Shell+Chev Avg",
    ])

    norm_rows = {}
    for arch in arch_order:
        for stage in stages:
            r += 1
            key = (arch, stage)
            p = petronas_projects[key]
            s = shell_projects[key]
            c = chevron_projects[key]
            b = basf_projects[key]

            cell(ws2, r, 1, arch, font=normal_font, align=left_wrap)
            cell(ws2, r, 2, stage, font=normal_font)
            # PETRONAS: PY | Cost | PY/MYR
            cell(ws2, r, 3, p["py"], font=normal_font, fmt="#,##0.0")
            cell(ws2, r, 4, p["cost_myr"], font=normal_font, fmt="#,##0.0")
            cell(ws2, r, 5, None, font=normal_font, fill=calc_fill, fmt="0.00")
            ws2[f"E{r}"] = f"=C{r}/D{r}"
            # Shell: PY | Cost | PY/MYR
            cell(ws2, r, 6, s["py"], font=normal_font, fmt="#,##0.0")
            cell(ws2, r, 7, s["cost_myr"], font=normal_font, fmt="#,##0.0")
            cell(ws2, r, 8, None, font=normal_font, fill=calc_fill, fmt="0.00")
            ws2[f"H{r}"] = f"=F{r}/G{r}"
            # Chevron: PY | Cost | PY/MYR
            cell(ws2, r, 9, c["py"], font=normal_font, fmt="#,##0.0")
            cell(ws2, r, 10, c["cost_myr"], font=normal_font, fmt="#,##0.0")
            cell(ws2, r, 11, None, font=normal_font, fill=calc_fill, fmt="0.00")
            ws2[f"K{r}"] = f"=I{r}/J{r}"
            # BASF (ref): PY | Cost | PY/MYR
            cell(ws2, r, 12, b["py"], font=normal_font, fmt="#,##0.0")
            cell(ws2, r, 13, b["cost_myr"], font=normal_font, fmt="#,##0.0")
            cell(ws2, r, 14, None, font=normal_font, fill=calc_fill, fmt="0.00")
            ws2[f"N{r}"] = f"=L{r}/M{r}"
            # Shell+Chevron Avg PY/MYR
            cell(ws2, r, 15, None, font=bold_font, fill=result_fill, fmt="0.000")
            ws2[f"O{r}"] = f"=(H{r}+K{r})/2"

            norm_rows[(arch, stage)] = r

    r += 1
    cell(ws2, r, 1,
         "Chevron TRL 1-4 costs are identical (6 MYR M) across all archetypes \u2014 likely blanket estimate. Inflates Chevron PY/MYR for TRL 1-4.",
         font=small_font, fill=caveat_fill, align=left_wrap, border=False)
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MCOLS2)
    r += 1
    cell(ws2, r, 1,
         "BASF includes plant/lab technicians in FTE \u2014 shown for reference only, not in average.",
         font=small_font, fill=caveat_fill, align=left_wrap, border=False)
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MCOLS2)

    # ------------------------------------------------------------------
    # Section B: PETRONAS Portfolio
    # ------------------------------------------------------------------
    r += 2
    cell(ws2, r, 1, "B. PETRONAS PORTFOLIO",
         font=section_font, align=left_wrap, border=False)
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MCOLS2)
    r += 1
    cell(ws2, r, 1,
         "Project counts from Inputs sheet. Avg Cost/Project and Avg PY/Project are PETRONAS per-project averages (see Reference Data sheet).",
         font=small_font, fill=None, align=left_wrap, border=False)
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MCOLS2)

    r += 1
    header_row(ws2, r, 1, [
        "Archetype", "TRL Stage", "# Projects",
        "Avg Cost/Project (MYR M)", "Total Cost (MYR M)",
        "Avg PY/Project", "Total Actual PY",
    ])

    portfolio_rows = {}
    for arch in arch_order:
        for stage in stages:
            r += 1
            key = (arch, stage)
            p = petronas_projects[key]

            cell(ws2, r, 1, arch, font=normal_font, align=left_wrap)
            cell(ws2, r, 2, stage, font=normal_font)
            cell(ws2, r, 3, None, font=normal_font, fill=calc_fill, fmt="#,##0.0")
            ws2[f"C{r}"] = f"=Inputs!C{breakdown_rows[(arch, stage)]}"
            cell(ws2, r, 4, p["cost_myr"], font=normal_font, fmt="#,##0.0")
            cell(ws2, r, 5, None, font=bold_font, fill=calc_fill, fmt="#,##0.0")
            ws2[f"E{r}"] = f"=C{r}*D{r}"
            cell(ws2, r, 6, p["py"], font=normal_font, fmt="#,##0.0")
            cell(ws2, r, 7, None, font=bold_font, fill=calc_fill, fmt="#,##0.0")
            ws2[f"G{r}"] = f"=C{r}*F{r}"

            portfolio_rows[(arch, stage)] = r

    pf_start = min(portfolio_rows.values())
    pf_end = max(portfolio_rows.values())
    r += 1
    cell(ws2, r, 1, "TOTAL", font=bold_font, align=left_wrap, fill=grey_fill)
    cell(ws2, r, 2, "", font=normal_font, fill=grey_fill)
    for ci in [3, 5, 7]:
        cl = get_column_letter(ci)
        cell(ws2, r, ci, None, font=bold_font, fill=grey_fill, fmt="#,##0.0")
        ws2[f"{cl}{r}"] = f"=SUM({cl}{pf_start}:{cl}{pf_end})"
    cell(ws2, r, 4, "", font=normal_font, fill=grey_fill)
    cell(ws2, r, 6, "", font=normal_font, fill=grey_fill)
    portfolio_total_row = r

    # ------------------------------------------------------------------
    # Section C: Implied Staffing
    # ------------------------------------------------------------------
    r += 2
    cell(ws2, r, 1, "C. IMPLIED STAFFING (Peer Norms Applied to PETRONAS Portfolio)",
         font=section_font, align=left_wrap, border=False)
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MCOLS2)
    r += 1
    cell(ws2, r, 1,
         "Implied PY = Total Cost x Peer PY/MYR Norm. Gap = Implied \u2212 Actual.",
         font=small_font, fill=None, align=left_wrap, border=False)
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MCOLS2)

    r += 1
    header_row(ws2, r, 1, [
        "Archetype", "TRL Stage", "Peer Norm (PY/MYR)",
        "Total Cost (MYR M)", "Implied PY", "Actual PY", "Gap (PY)", "Ratio",
    ])

    implied_rows = {}
    for arch in arch_order:
        for stage in stages:
            r += 1
            key = (arch, stage)
            nr = norm_rows[key]
            pr = portfolio_rows[key]

            cell(ws2, r, 1, arch, font=normal_font, align=left_wrap)
            cell(ws2, r, 2, stage, font=normal_font)
            cell(ws2, r, 3, None, font=normal_font, fill=calc_fill, fmt="0.000")
            ws2[f"C{r}"] = f"=O{nr}"
            cell(ws2, r, 4, None, font=normal_font, fill=calc_fill, fmt="#,##0.0")
            ws2[f"D{r}"] = f"=E{pr}"
            cell(ws2, r, 5, None, font=bold_font, fill=calc_fill, fmt="#,##0.0")
            ws2[f"E{r}"] = f"=D{r}*C{r}"
            cell(ws2, r, 6, None, font=bold_font, fill=calc_fill, fmt="#,##0.0")
            ws2[f"F{r}"] = f"=G{pr}"
            cell(ws2, r, 7, None, font=bold_font, fill=calc_fill, fmt="#,##0.0")
            ws2[f"G{r}"] = f"=E{r}-F{r}"
            cell(ws2, r, 8, None, font=bold_font, fill=orange_fill, fmt='0.00"x"')
            ws2[f"H{r}"] = f"=IF(F{r}=0,0,E{r}/F{r})"

            implied_rows[(arch, stage)] = r

    imp_start = min(implied_rows.values())
    imp_end = max(implied_rows.values())
    r += 1
    cell(ws2, r, 1, "OVERALL", font=bold_font, align=left_wrap, fill=grey_fill)
    cell(ws2, r, 2, "", font=normal_font, fill=grey_fill)
    cell(ws2, r, 3, "", font=normal_font, fill=grey_fill)
    for ci in [4, 5, 6, 7]:
        cl = get_column_letter(ci)
        cell(ws2, r, ci, None, font=bold_font, fill=grey_fill, fmt="#,##0.0")
        ws2[f"{cl}{r}"] = f"=SUM({cl}{imp_start}:{cl}{imp_end})"
    cell(ws2, r, 8, None, font=bold_font, fill=orange_fill, fmt='0.00"x"')
    ws2[f"H{r}"] = f"=IF(F{r}=0,0,E{r}/F{r})"
    implied_total_row = r

    # ------------------------------------------------------------------
    # Section D: Summary
    # ------------------------------------------------------------------
    r += 2
    cell(ws2, r, 1, "D. SUMMARY", font=section_font, align=left_wrap, border=False)
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MCOLS2)

    r += 1
    cell(ws2, r, 1, "Total Implied PY (peer norms)", font=normal_font, align=left_wrap)
    cell(ws2, r, 2, None, font=bold_font, fill=result_fill, fmt="#,##0.0")
    ws2[f"B{r}"] = f"=E{implied_total_row}"
    implied_summary_row = r

    r += 1
    cell(ws2, r, 1, "Total Actual PY (PETRONAS)", font=normal_font, align=left_wrap)
    cell(ws2, r, 2, None, font=bold_font, fill=result_fill, fmt="#,##0.0")
    ws2[f"B{r}"] = f"=F{implied_total_row}"
    actual_summary_row = r

    r += 1
    cell(ws2, r, 1, "Normalization Factor", font=bold_font, align=left_wrap)
    cell(ws2, r, 2, None, font=bold_font, fill=orange_fill, fmt='0.00"x"')
    ws2[f"B{r}"] = f"=IF(B{actual_summary_row}=0,0,B{implied_summary_row}/B{actual_summary_row})"

    # ══════════════════════════════════════════════════════════════════
    # Sheet 3: Project Data
    # ══════════════════════════════════════════════════════════════════
    ws_pd = wb.create_sheet("Project Data")
    ws_pd.sheet_properties.tabColor = DARK_BLUE

    for col in range(1, 20):
        ws_pd.column_dimensions[get_column_letter(col)].width = 14
    ws_pd.column_dimensions["A"].width = 18
    MCOLS3 = 18

    r = 1
    merge_title(ws_pd, r, 1, MCOLS3, "Per-Project Data: PETRONAS vs Shell vs Chevron vs BASF")
    r += 1
    cell(ws_pd, r, 1,
         "Person-years = Team FTE \u00d7 (Duration / 12). PETRONAS: actual project averages. Peers: outside-in benchmark estimates. All costs in MYR M.",
         font=small_font, fill=None, align=left_wrap, border=False)
    ws_pd.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MCOLS3)

    # --- Detail table ---
    r += 2
    cell(ws_pd, r, 1, "DETAIL: Per-project metrics by archetype and TRL stage",
         font=section_font, align=left_wrap, border=False)
    ws_pd.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MCOLS3)

    r += 1
    header_row(ws_pd, r, 1, [
        "Archetype", "TRL Stage",
        "PET Dur (mo)", "PET FTE", "PET PY", "PET Cost (MYR M)",
        "Shell Dur (mo)", "Shell FTE", "Shell PY", "Shell Cost (MYR M)",
        "Chev Dur (mo)", "Chev FTE", "Chev PY", "Chev Cost (MYR M)",
        "BASF Dur (mo)", "BASF FTE", "BASF PY", "BASF Cost (MYR M)",
    ])

    for arch in arch_order:
        for stage in stages:
            r += 1
            key = (arch, stage)
            cell(ws_pd, r, 1, arch, font=normal_font, align=left_wrap)
            cell(ws_pd, r, 2, stage, font=normal_font)
            col = 3
            for _, proj_dict in all_companies:
                p = proj_dict[key]
                cell(ws_pd, r, col, p["dur"], font=normal_font, fmt="#,##0")
                cell(ws_pd, r, col+1, p["fte"], font=normal_font, fmt="#,##0.0")
                cell(ws_pd, r, col+2, p["py"], font=bold_font, fill=calc_fill, fmt="#,##0.0")
                cell(ws_pd, r, col+3, p["cost_myr"], font=normal_font, fmt="#,##0.0")
                col += 4

    # --- Summary ---
    r += 2
    cell(ws_pd, r, 1, "SUMMARY: Total person-years per full R&D cycle (TRL 1-7)",
         font=section_font, align=left_wrap, border=False)
    ws_pd.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MCOLS3)
    r += 1
    cell(ws_pd, r, 1,
         "Sum of person-years across TRL 1-4 and TRL 5-7 = total effort per technology.",
         font=small_font, fill=None, align=left_wrap, border=False)
    ws_pd.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MCOLS3)

    r += 1
    header_row(ws_pd, r, 1, [
        "Archetype", "PETRONAS (PY)", "Shell (PY)", "Chevron (PY)", "BASF (PY)",
        "Shell / PETRONAS", "Chevron / PETRONAS", "BASF / PETRONAS",
    ])

    for arch in arch_order:
        r += 1
        totals = []
        for _, proj_dict in all_companies:
            totals.append(proj_dict[(arch, "TRL 1-4")]["py"] + proj_dict[(arch, "TRL 5-7")]["py"])
        cell(ws_pd, r, 1, arch, font=normal_font, align=left_wrap)
        for i, t in enumerate(totals):
            cell(ws_pd, r, 2+i, round(t, 1), font=bold_font, fill=calc_fill, fmt="#,##0.0")
        pt_total = totals[0]
        for i in range(1, 4):
            cell(ws_pd, r, 5+i, round(totals[i] / pt_total, 1) if pt_total else 0,
                 font=bold_font, fill=orange_fill, fmt='0.0"x"')

    # --- Data source notes ---
    r += 2
    cell(ws_pd, r, 1, "DATA SOURCES", font=section_font, align=left_wrap, border=False)
    ws_pd.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MCOLS3)
    notes = [
        "PETRONAS: averages from actual reference projects (Fgo, Garcinia, Elektra, ASAT, etc.). See Reference Data sheet.",
        "  PY = average of individual project PYs (may differ from Avg FTE \u00d7 Avg Duration / 12 because projects vary in size and length).",
        "Shell: outside-in benchmark estimates.",
        "Chevron: outside-in benchmark estimates. TRL 1-4 costs are identical (6 MYR M) across archetypes \u2014 likely blanket estimate.",
        "BASF: outside-in benchmark estimates. FTE includes plant/lab technicians \u2014 shown for reference only.",
    ]
    for note in notes:
        r += 1
        cell(ws_pd, r, 1, note, font=small_font, fill=None, align=left_wrap, border=False)
        ws_pd.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MCOLS3)

    # ══════════════════════════════════════════════════════════════════
    # Sheet 4: Reference Data
    # ══════════════════════════════════════════════════════════════════
    ws_ref = wb.create_sheet("Reference Data")
    ws_ref.sheet_properties.tabColor = DARK_GREEN

    for col in range(1, 12):
        ws_ref.column_dimensions[get_column_letter(col)].width = 20

    r = 1
    merge_title(ws_ref, r, 1, 10, "Reference Data: PETRONAS Project Detail")
    r += 1
    cell(ws_ref, r, 1,
         "Actual PETRONAS projects used to derive the PETRONAS averages on the Staffing Norms sheet.",
         font=small_font, fill=None, align=left_wrap, border=False)
    ws_ref.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)

    r += 2
    cell(ws_ref, r, 1, "PROJECT-BY-PROJECT DETAIL", font=section_font, align=left_wrap, border=False)
    ws_ref.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)

    r += 1
    header_row(ws_ref, r, 1, [
        "Archetype", "TRL Stage", "Project", "Duration (mo)",
        "Team FTE", "Person-Years", "Cost (MYR M)",
    ])

    for (arch, stage), projects in ref_projects.items():
        for pname, dur, fte, cost in projects:
            r += 1
            py = fte * (dur / 12.0)
            cell(ws_ref, r, 1, arch, font=normal_font, align=left_wrap)
            cell(ws_ref, r, 2, stage, font=normal_font)
            cell(ws_ref, r, 3, pname, font=normal_font, align=left_wrap)
            cell(ws_ref, r, 4, dur, font=normal_font, fmt="#,##0")
            cell(ws_ref, r, 5, fte, font=normal_font, fmt="#,##0")
            cell(ws_ref, r, 6, round(py, 1), font=bold_font, fill=calc_fill, fmt="#,##0.0")
            cell(ws_ref, r, 7, cost, font=normal_font, fmt="#,##0.0")

    r += 2
    cell(ws_ref, r, 1, "AVERAGES PER TRL STAGE", font=section_font, align=left_wrap, border=False)
    ws_ref.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 1
    cell(ws_ref, r, 1,
         "Avg PY = average of individual project PYs (not Avg FTE \u00d7 Avg Duration / 12, because projects vary in size and length).",
         font=small_font, fill=None, align=left_wrap, border=False)
    ws_ref.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)

    r += 1
    header_row(ws_ref, r, 1, [
        "Archetype", "TRL Stage", "Avg Duration (mo)", "Avg Team FTE",
        "Avg PY / Completion", "Avg Cost (MYR M)", "# Projects", "Range (PY)",
    ])

    for (arch, stage), projects in ref_projects.items():
        r += 1
        n = len(projects)
        durs = [p[1] for p in projects]
        ftes = [p[2] for p in projects]
        pys = [p[2] * (p[1] / 12.0) for p in projects]
        costs = [p[3] for p in projects]
        cell(ws_ref, r, 1, arch, font=normal_font, align=left_wrap)
        cell(ws_ref, r, 2, stage, font=normal_font)
        cell(ws_ref, r, 3, round(sum(durs) / n, 0), font=normal_font, fmt="#,##0")
        cell(ws_ref, r, 4, round_half_up(sum(ftes) / n, 1), font=normal_font, fmt="#,##0.0")
        cell(ws_ref, r, 5, round_half_up(sum(pys) / n, 1), font=bold_font, fill=calc_fill, fmt="#,##0.0")
        cell(ws_ref, r, 6, round_half_up(sum(costs) / n, 1), font=bold_font, fill=calc_fill, fmt="#,##0.0")
        cell(ws_ref, r, 7, n, font=normal_font, fmt="#,##0")
        cell(ws_ref, r, 8, f"{min(pys):.1f} - {max(pys):.1f}", font=small_font, align=left_wrap)

    r += 2
    cell(ws_ref, r, 1, "FULL R&D CYCLE (TRL 1-4 + TRL 5-7)", font=section_font, align=left_wrap, border=False)
    ws_ref.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)

    r += 1
    header_row(ws_ref, r, 1, [
        "Archetype", "Avg PY (TRL 1-4)", "Avg PY (TRL 5-7)",
        "Total PY per Cycle", "Total Cost per Cycle (MYR M)",
    ])

    arch_groups = {}
    arch_cost_groups = {}
    for (arch, stage), projects in ref_projects.items():
        pys = [p[2] * (p[1] / 12.0) for p in projects]
        costs = [p[3] for p in projects]
        arch_groups.setdefault(arch, {})[stage] = sum(pys) / len(pys)
        arch_cost_groups.setdefault(arch, {})[stage] = sum(costs) / len(costs)

    for arch in ["Chemistry", "HW Process", "HW Mechanical", "Algorithm"]:
        r += 1
        early = arch_groups.get(arch, {}).get("TRL 1-4", 0)
        late = arch_groups.get(arch, {}).get("TRL 5-7", 0)
        cost_early = arch_cost_groups.get(arch, {}).get("TRL 1-4", 0)
        cost_late = arch_cost_groups.get(arch, {}).get("TRL 5-7", 0)
        cell(ws_ref, r, 1, arch, font=normal_font, align=left_wrap)
        cell(ws_ref, r, 2, round(early, 1), font=normal_font, fill=calc_fill, fmt="#,##0.0")
        cell(ws_ref, r, 3, round(late, 1), font=normal_font, fill=calc_fill, fmt="#,##0.0")
        cell(ws_ref, r, 4, round(early + late, 1), font=bold_font, fill=calc_fill, fmt="#,##0.0")
        cell(ws_ref, r, 5, round(cost_early + cost_late, 1), font=bold_font, fill=calc_fill, fmt="#,##0.0")

    # ══════════════════════════════════════════════════════════════════
    # Sheet 5: Methodology
    # ══════════════════════════════════════════════════════════════════
    ws_m = wb.create_sheet("Methodology")
    ws_m.sheet_properties.tabColor = TEAL

    ws_m.column_dimensions["A"].width = 100
    MCOLS_M = 1

    meth_font = Font(name="Calibri", size=11)
    meth_bold = Font(name="Calibri", size=11, bold=True, color=TEAL)
    meth_section = Font(name="Calibri", size=12, bold=True, color=TEAL)

    def mline(ws, row, text, font=meth_font):
        c = ws.cell(row=row, column=1, value=text)
        c.font = font
        c.alignment = left_wrap
        return row + 1

    r = 1
    r = mline(ws_m, r, "Staffing Norms \u2014 Methodology", title_font)
    r += 1

    # --- WHAT THIS WORKBOOK DOES ---
    r = mline(ws_m, r, "WHAT THIS WORKBOOK DOES", meth_section)
    r = mline(ws_m, r, "This workbook answers: if PETRONAS staffed its R&D portfolio at the same")
    r = mline(ws_m, r, "intensity as Shell and Chevron, how many person-years would it need?")
    r += 1
    r = mline(ws_m, r, "It does this in three steps:")
    r = mline(ws_m, r, "  1. Extract staffing norms (PY per MYR M of project cost) from Shell and Chevron benchmarks.")
    r = mline(ws_m, r, "  2. Apply those norms to PETRONAS\u2019s own portfolio of projects and costs.")
    r = mline(ws_m, r, "  3. Compare the implied person-years against PETRONAS\u2019s actual staffing to find the gap.")
    r += 1
    r = mline(ws_m, r, "The result is a Normalization Factor: implied PY / actual PY.")
    r = mline(ws_m, r, "A factor above 1.0x means peer norms imply more effort than PETRONAS currently deploys.")
    r += 1

    # --- SHEET-BY-SHEET GUIDE ---
    r = mline(ws_m, r, "SHEET-BY-SHEET GUIDE", meth_section)
    r += 1
    r = mline(ws_m, r, "Inputs", meth_bold)
    r = mline(ws_m, r, "  Editable assumptions: total active projects (default 160), TRL 1-4 share (default 50%),")
    r = mline(ws_m, r, "  and archetype shares (Chemistry 40%, HW Mechanical 18%, HW Process 25%, Algorithm 17%).")
    r = mline(ws_m, r, "  Yellow cells are inputs. The portfolio breakdown (project counts per bucket) auto-calculates.")
    r = mline(ws_m, r, "  FTE costs and FX rate are shown for reference but are not used in the staffing norms calculation.")
    r += 1
    r = mline(ws_m, r, "Staffing Norms (core analysis)", meth_bold)
    r = mline(ws_m, r, "  Section A \u2014 Peer Staffing Norms")
    r = mline(ws_m, r, "    Shows PY, Cost (MYR M), and PY/MYR for PETRONAS, Shell, Chevron, and BASF (reference only).")
    r = mline(ws_m, r, "    Each PY/MYR cell is a formula dividing the adjacent PY by Cost, so the derivation is transparent.")
    r = mline(ws_m, r, "    The Shell+Chevron Average is the peer norm used in the rest of the workbook.")
    r += 1
    r = mline(ws_m, r, "  Section B \u2014 PETRONAS Portfolio")
    r = mline(ws_m, r, "    Project counts link to the Inputs sheet. Avg Cost/Project and Avg PY/Project are PETRONAS")
    r = mline(ws_m, r, "    per-project averages from the Reference Data sheet. Total Cost and Total PY are count \u00d7 average.")
    r += 1
    r = mline(ws_m, r, "  Section C \u2014 Implied Staffing")
    r = mline(ws_m, r, "    Applies the Shell+Chevron norm (from Section A) to each bucket\u2019s Total Cost (from Section B).")
    r = mline(ws_m, r, "    Implied PY = Total Cost \u00d7 Peer Norm. Gap = Implied \u2212 Actual. Ratio = Implied / Actual.")
    r += 1
    r = mline(ws_m, r, "  Section D \u2014 Summary")
    r = mline(ws_m, r, "    Aggregates Section C into a single Normalization Factor (total implied PY / total actual PY).")
    r += 1
    r = mline(ws_m, r, "Project Data", meth_bold)
    r = mline(ws_m, r, "  Per-project metrics for all four companies: duration, FTE, person-years, and cost (MYR M).")
    r = mline(ws_m, r, "  Summary table shows total PY per archetype and peer-to-PETRONAS multipliers.")
    r = mline(ws_m, r, "  Data source notes at the bottom explain where each company\u2019s numbers come from.")
    r += 1
    r = mline(ws_m, r, "Reference Data", meth_bold)
    r = mline(ws_m, r, "  Raw project-by-project detail for PETRONAS: project names, durations, team sizes, costs,")
    r = mline(ws_m, r, "  and computed person-years. Averages per TRL stage and full-cycle totals are shown below.")
    r = mline(ws_m, r, "  This is where Section B\u2019s Avg Cost and Avg PY values can be verified.")
    r += 1

    # --- KEY CONCEPTS ---
    r = mline(ws_m, r, "KEY CONCEPTS", meth_section)
    r += 1
    r = mline(ws_m, r, "Person-Years (PY)", meth_bold)
    r = mline(ws_m, r, "  At the project level: PY = Team FTE \u00d7 (Duration in months / 12).")
    r = mline(ws_m, r, "  One person working full-time for one year = 1 person-year.")
    r = mline(ws_m, r, "  This measures total human effort regardless of team size or project duration.")
    r += 1
    r = mline(ws_m, r, "PY / MYR M (Staffing Norm)", meth_bold)
    r = mline(ws_m, r, "  Person-years per million MYR of project cost. This is the core metric.")
    r = mline(ws_m, r, "  Higher PY/MYR = more human effort per unit of spending.")
    r = mline(ws_m, r, "  By comparing PY/MYR across companies, we normalize for differences in project budgets")
    r = mline(ws_m, r, "  and focus on staffing intensity: how many people does a company put on each MYR of R&D?")
    r += 1

    # --- DATA SOURCES ---
    r = mline(ws_m, r, "DATA SOURCES", meth_section)
    r += 1
    r = mline(ws_m, r, "PETRONAS", meth_bold)
    r = mline(ws_m, r, "  Averages from actual reference projects (Fgo, Garcinia, Elektra, ASAT, Zod, Reseis, etc.).")
    r = mline(ws_m, r, "  2\u20134 projects per archetype/TRL stage. Individual projects vary widely (see Reference Data).")
    r = mline(ws_m, r, "  PY is the average of individual project PYs, which may differ from Avg FTE \u00d7 Avg Duration / 12")
    r = mline(ws_m, r, "  because projects with more FTE also tend to run longer.")
    r += 1
    r = mline(ws_m, r, "Shell", meth_bold)
    r = mline(ws_m, r, "  Outside-in benchmark estimates.")
    r += 1
    r = mline(ws_m, r, "Chevron", meth_bold)
    r = mline(ws_m, r, "  Outside-in benchmark estimates.")
    r = mline(ws_m, r, "  Note: TRL 1-4 costs are identical (6 MYR M) across all archetypes \u2014 likely a blanket estimate.")
    r = mline(ws_m, r, "  This inflates Chevron\u2019s PY/MYR for TRL 1-4 buckets.")
    r += 1
    r = mline(ws_m, r, "BASF", meth_bold)
    r = mline(ws_m, r, "  Outside-in benchmark estimates. Shown for reference only, not included in the peer average.")
    r = mline(ws_m, r, "  BASF FTE includes large numbers of plant/lab technicians (20\u201350 per project in TRL 5-7).")
    r = mline(ws_m, r, "  PETRONAS and Shell FTE primarily reflect researchers and engineers.")
    r += 1
    r = mline(ws_m, r, "Currency", meth_bold)
    r = mline(ws_m, r, "  All costs are in MYR M. Shell, Chevron, and BASF costs are converted from USD at FX = 4.0 MYR/USD.")
    r += 1

    # --- KEY ASSUMPTIONS AND CAVEATS ---
    r = mline(ws_m, r, "KEY ASSUMPTIONS AND CAVEATS", meth_section)
    r += 1
    r = mline(ws_m, r, "1. Archetypes are comparable across companies (Chemistry vs Chemistry, etc.).")
    r = mline(ws_m, r, "   If scope differs materially, the comparison overstates or understates effort.")
    r += 1
    r = mline(ws_m, r, "2. One person-year in PETRONAS = one person-year at Shell/Chevron (baseline assumption).")
    r = mline(ws_m, r, "   No productivity adjustment is applied. If peer labs have better tools/equipment,")
    r = mline(ws_m, r, "   each peer PY may be worth more, and the implied gap would be smaller.")
    r += 1
    r = mline(ws_m, r, "3. Peer project parameters are outside-in estimates, not audited figures.")
    r = mline(ws_m, r, "   The staffing norms are only as reliable as the underlying benchmark data.")
    r += 1
    r = mline(ws_m, r, "4. PETRONAS averages are computed from a small number of reference projects (2\u20134 per bucket).")
    r = mline(ws_m, r, "   Individual projects vary widely \u2014 see the Range column in Reference Data.")
    r += 1
    r = mline(ws_m, r, "5. The peer norm uses the Shell+Chevron average. BASF is excluded because its FTE definition")
    r = mline(ws_m, r, "   includes plant/lab technicians, making it non-comparable for staffing intensity.")
    r += 1
    r = mline(ws_m, r, "6. Chevron TRL 1-4 costs appear to be a blanket estimate (identical across archetypes).")
    r = mline(ws_m, r, "   This may overstate Chevron\u2019s PY/MYR for early-stage work and bias the average upward.")
    r += 1

    # --- HOW TO UPDATE ---
    r = mline(ws_m, r, "HOW TO UPDATE", meth_section)
    r += 1
    r = mline(ws_m, r, "Editable inputs are on the Inputs sheet (yellow cells):")
    r = mline(ws_m, r, "  \u2022 Total Active Projects \u2014 change if the portfolio size changes.")
    r = mline(ws_m, r, "  \u2022 TRL 1-4 Share \u2014 adjust if the early-stage / late-stage split changes.")
    r = mline(ws_m, r, "  \u2022 Archetype Shares \u2014 update if the portfolio mix shifts.")
    r += 1
    r = mline(ws_m, r, "The Staffing Norms sheet (Sections B\u2013D) auto-updates when Inputs change.")
    r = mline(ws_m, r, "Section A (peer norms) and Project Data contain static benchmark data \u2014")
    r = mline(ws_m, r, "these would need to be regenerated if peer benchmarks or PETRONAS project data change.")

    # ══════════════════════════════════════════════════════════════════
    # Save
    # ══════════════════════════════════════════════════════════════════
    out_path = Path(__file__).parent / "PETRONAS_Normalization.xlsx"
    wb.save(out_path)
    print(f"Saved to: {out_path}")

    onedrive_path = Path(r"c:\Users\Debdoot Ray\OneDrive - McKinsey & Company\Desktop\Petronas study\20260311_Normalization_v5.xlsx")
    if onedrive_path.parent.exists():
        wb.save(onedrive_path)
        print(f"Saved to: {onedrive_path}")
    else:
        print(f"OneDrive path not found: {onedrive_path.parent}")


if __name__ == "__main__":
    main()

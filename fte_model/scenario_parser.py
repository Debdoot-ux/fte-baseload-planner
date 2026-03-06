"""
Keyword-based Excel parser for simulation input files.

Scans uploaded Excel workbooks for scenario definitions and project assumption
tables, using keyword anchors rather than fixed cell positions.  Returns
structured ParseResult data that can be converted into ModelConfig objects.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


# ── Data structures ───────────────────────────────────────────────────────

@dataclass
class ParsedScenario:
    name: str
    budget: Optional[float] = None
    overhead_pct: Optional[float] = None
    stage_mix: Optional[Dict[str, float]] = None
    conversion_rates: Optional[Dict[str, float]] = None
    archetype_shares: Optional[Dict[str, float]] = None


@dataclass
class ParsedProject:
    archetype: str
    phase: str
    reference: str
    duration_months: Optional[int] = None
    fte: Optional[float] = None
    cost_millions: Optional[float] = None


@dataclass
class ParseResult:
    scenarios: List[ParsedScenario] = field(default_factory=list)
    projects: List[ParsedProject] = field(default_factory=list)
    archetype_names: List[str] = field(default_factory=list)
    phase_names: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


# ── Keyword sets for sheet/header identification ──────────────────────────

_SCENARIO_KEYWORDS = {"scenario", "budget", "overhead", "split", "success"}
_ASSUMPTION_KEYWORDS = {"archetype", "phase", "duration", "fte", "cost"}
_KNOWN_ABBREVS: Dict[str, str] = {
    "HP": "Hardware Process",
    "HM": "Hardware Mechanical",
    "AI": "Algorithm",
    "ALGO": "Algorithm",
}

# Phase normalisation: map various labels to canonical names
_PHASE_MAP = {
    "trl 1-4": "TRL 1-4",
    "trl 1 - 4": "TRL 1-4",
    "trl1-4": "TRL 1-4",
    "> trl 4": "TRL 5-7",
    ">trl 4": "TRL 5-7",
    ">trl4": "TRL 5-7",
    "trl 5-7": "TRL 5-7",
    "trl 5 - 7": "TRL 5-7",
    "trl5-7": "TRL 5-7",
    "trl 4-7": "TRL 5-7",
    "> trl4": "TRL 5-7",
}


# ── Helpers ───────────────────────────────────────────────────────────────

def _cell_text(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _cell_lower(val: Any) -> str:
    return _cell_text(val).lower()


def _is_row_empty(row: tuple) -> bool:
    return all(v is None or str(v).strip() == "" for v in row)


def _match_keywords(row: tuple, keywords: set, threshold: int = 3) -> int:
    """Return number of keyword hits in *row* (case-insensitive substring)."""
    hits = 0
    for cell in row:
        text = _cell_lower(cell)
        if not text:
            continue
        for kw in keywords:
            if kw in text:
                hits += 1
                break
    return hits


def _parse_number(val: Any) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    text = str(val).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _parse_duration(val: Any) -> Optional[int]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    text = str(val).strip().lower()
    text = re.sub(r"months?|mo\b", "", text).strip()
    try:
        return int(float(text))
    except (ValueError, TypeError):
        return None


def _parse_cost(val: Any) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    text = str(val).strip()
    text = re.sub(r"(?i)\s*(mil|million|m|rm)\s*", "", text).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _normalize_fraction(val: Optional[float]) -> Optional[float]:
    """If val > 1 treat as percentage and divide by 100."""
    if val is None:
        return None
    if val > 1.0:
        return val / 100.0
    return val


def _parse_split(val: Any, stage_names: List[str]) -> Optional[Dict[str, float]]:
    """Parse a portfolio split string like '30/70' into a stage_mix dict."""
    text = _cell_text(val)
    if not text:
        return None
    parts = re.split(r"[/:\\-]", text)
    nums: List[float] = []
    for p in parts:
        p = p.strip()
        try:
            nums.append(float(p))
        except ValueError:
            return None
    if not nums:
        return None
    total = sum(nums)
    if total > 2:
        nums = [n / total for n in nums]
    elif total <= 0:
        return None
    else:
        pass  # already fractions
    result: Dict[str, float] = {}
    for i, sn in enumerate(stage_names):
        if i < len(nums):
            result[sn] = nums[i]
    return result


def _normalise_phase(raw: str) -> str:
    key = raw.strip().lower()
    return _PHASE_MAP.get(key, raw.strip())


# ── Sheet identification ──────────────────────────────────────────────────

def _find_header_row(
    ws, keywords: set, max_rows: int = 30, threshold: int = 3
) -> Optional[int]:
    for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True)):
        if _match_keywords(row, keywords, threshold) >= threshold:
            return row_idx
    for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True)):
        if _match_keywords(row, keywords, threshold) >= 2:
            return row_idx
    return None


def _identify_sheets(wb) -> Tuple[Optional[str], Optional[str]]:
    scenario_sheet = None
    assumption_sheet = None
    best_scen = 0
    best_assum = 0

    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(max_row=30, values_only=True))
        for row in rows:
            scen_hits = _match_keywords(row, _SCENARIO_KEYWORDS)
            assum_hits = _match_keywords(row, _ASSUMPTION_KEYWORDS)
            if scen_hits > best_scen:
                best_scen = scen_hits
                scenario_sheet = name
            if assum_hits > best_assum:
                best_assum = assum_hits
                assumption_sheet = name

    return scenario_sheet, assumption_sheet


# ── Scenario sheet parsing ────────────────────────────────────────────────

_SCENARIO_COL_PATTERNS = {
    "scenario": re.compile(r"scenario", re.I),
    "name": re.compile(r"justif|name|description", re.I),
    "budget": re.compile(r"budget", re.I),
    "overhead": re.compile(r"overhead", re.I),
    "split": re.compile(r"split|portfolio", re.I),
    "success": re.compile(r"success|conversion|rate", re.I),
}


def _map_scenario_columns(header_row: tuple) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for ci, cell in enumerate(header_row):
        text = _cell_text(cell)
        if not text:
            continue
        for key, pat in _SCENARIO_COL_PATTERNS.items():
            if pat.search(text) and key not in mapping:
                mapping[key] = ci
                break
    return mapping


def _find_archetype_columns(
    header_row: tuple, known_cols: Dict[str, int]
) -> List[Tuple[int, str]]:
    """Columns to the right of known parameter columns are archetype share cols."""
    max_known = max(known_cols.values()) if known_cols else -1
    arch_cols: List[Tuple[int, str]] = []
    for ci, cell in enumerate(header_row):
        if ci <= max_known:
            continue
        text = _cell_text(cell)
        if text:
            arch_cols.append((ci, text))
    return arch_cols


def _parse_scenarios(
    ws, header_idx: int, stage_names: List[str], warnings: List[str]
) -> Tuple[List[ParsedScenario], List[Tuple[int, str]]]:
    rows = list(ws.iter_rows(values_only=True))
    header_row = rows[header_idx]

    col_map = _map_scenario_columns(header_row)
    arch_cols = _find_archetype_columns(header_row, col_map)

    scenarios: List[ParsedScenario] = []
    base: Optional[ParsedScenario] = None

    for row in rows[header_idx + 1:]:
        if _is_row_empty(row):
            if scenarios:
                break
            continue

        scen_id = _cell_text(row[col_map["scenario"]]) if "scenario" in col_map else ""
        if not scen_id:
            continue

        name = _cell_text(row[col_map.get("name", -1)]) if "name" in col_map else f"Scenario {scen_id}"
        if not name:
            name = f"Scenario {scen_id}"

        budget = _parse_number(row[col_map["budget"]]) if "budget" in col_map else None
        overhead_raw = _parse_number(row[col_map["overhead"]]) if "overhead" in col_map else None
        overhead = _normalize_fraction(overhead_raw)
        split_val = row[col_map["split"]] if "split" in col_map else None
        stage_mix = _parse_split(split_val, stage_names)
        success_raw = _parse_number(row[col_map["success"]]) if "success" in col_map else None
        success = _normalize_fraction(success_raw)
        conversion_rates: Optional[Dict[str, float]] = None
        if success is not None and len(stage_names) >= 1:
            conversion_rates = {stage_names[0]: success}

        arch_shares: Optional[Dict[str, float]] = None
        share_vals = {}
        for ci, aname in arch_cols:
            if ci < len(row):
                v = _parse_number(row[ci])
                if v is not None:
                    share_vals[aname] = _normalize_fraction(v)
        if share_vals:
            arch_shares = share_vals

        scen = ParsedScenario(
            name=name,
            budget=budget,
            overhead_pct=overhead,
            stage_mix=stage_mix,
            conversion_rates=conversion_rates,
            archetype_shares=arch_shares,
        )

        if base is None:
            base = scen
        else:
            if scen.budget is None:
                scen.budget = base.budget
            if scen.overhead_pct is None:
                scen.overhead_pct = base.overhead_pct
            if scen.stage_mix is None:
                scen.stage_mix = base.stage_mix
            if scen.conversion_rates is None:
                scen.conversion_rates = base.conversion_rates
            if scen.archetype_shares is None:
                scen.archetype_shares = base.archetype_shares

        scenarios.append(scen)

    if not scenarios:
        warnings.append("No scenario rows found in the scenario sheet.")

    return scenarios, arch_cols


# ── Assumption sheet parsing ──────────────────────────────────────────────

_ASSUMPTION_COL_PATTERNS = {
    "archetype": re.compile(r"archetype|type", re.I),
    "phase": re.compile(r"phase|trl|stage", re.I),
    "reference": re.compile(r"refer|project|name", re.I),
    "duration": re.compile(r"duration|month", re.I),
    "fte": re.compile(r"\bfte\b|headcount|staff", re.I),
    "cost": re.compile(r"cost|budget|rm", re.I),
}


def _map_assumption_columns(header_row: tuple) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for ci, cell in enumerate(header_row):
        text = _cell_text(cell)
        if not text:
            continue
        for key, pat in _ASSUMPTION_COL_PATTERNS.items():
            if pat.search(text) and key not in mapping:
                mapping[key] = ci
                break
    return mapping


def _parse_assumptions(
    ws, header_idx: int, warnings: List[str]
) -> Tuple[List[ParsedProject], List[str], List[str]]:
    rows = list(ws.iter_rows(values_only=True))
    header_row = rows[header_idx]
    col_map = _map_assumption_columns(header_row)

    projects: List[ParsedProject] = []
    archetype_names_seen: List[str] = []
    phase_names_seen: List[str] = []

    current_arch = ""
    current_phase = ""

    for row in rows[header_idx + 1:]:
        if _is_row_empty(row):
            continue

        arch_val = _cell_text(row[col_map["archetype"]]) if "archetype" in col_map else ""
        phase_val = _cell_text(row[col_map["phase"]]) if "phase" in col_map else ""
        ref_val = _cell_text(row[col_map["reference"]]) if "reference" in col_map else ""

        if arch_val:
            current_arch = arch_val
            if current_arch not in archetype_names_seen:
                archetype_names_seen.append(current_arch)
        if phase_val:
            current_phase = _normalise_phase(phase_val)
            if current_phase not in phase_names_seen:
                phase_names_seen.append(current_phase)

        if not current_arch or not ref_val:
            continue

        dur = _parse_duration(row[col_map["duration"]]) if "duration" in col_map else None
        fte = _parse_number(row[col_map["fte"]]) if "fte" in col_map else None
        cost = _parse_cost(row[col_map["cost"]]) if "cost" in col_map else None

        if cost is None and "cost" in col_map:
            raw = _cell_text(row[col_map["cost"]])
            if raw and raw.lower() not in ("", "none", "-"):
                warnings.append(
                    f"Could not parse cost for {current_arch} > {current_phase} > {ref_val}: '{raw}'"
                )

        projects.append(ParsedProject(
            archetype=current_arch,
            phase=current_phase or "TRL 1-4",
            reference=ref_val,
            duration_months=dur,
            fte=fte,
            cost_millions=cost,
        ))

    return projects, archetype_names_seen, phase_names_seen


# ── Archetype name matching ──────────────────────────────────────────────

def _match_archetype_names(
    scenario_abbrevs: List[str],
    assumption_names: List[str],
    warnings: List[str],
) -> Dict[str, str]:
    """Map scenario column abbreviations to assumption archetype names."""
    mapping: Dict[str, str] = {}
    used: set = set()

    # Pass 1: exact matches (case-insensitive)
    for abbr in scenario_abbrevs:
        abbr_lower = abbr.lower().strip()
        for full in assumption_names:
            if full in used:
                continue
            if abbr_lower == full.lower().strip():
                mapping[abbr] = full
                used.add(full)
                break

    # Pass 2: known abbreviation map (strip punctuation for fuzzy comparison)
    def _alpha_only(s: str) -> str:
        return re.sub(r"[^a-z0-9 ]", "", s.lower()).strip()

    for abbr in scenario_abbrevs:
        if abbr in mapping:
            continue
        if abbr.upper() in _KNOWN_ABBREVS:
            target = _alpha_only(_KNOWN_ABBREVS[abbr.upper()])
            for full in assumption_names:
                if full in used:
                    continue
                full_norm = _alpha_only(full)
                if target in full_norm or full_norm in target:
                    mapping[abbr] = full
                    used.add(full)
                    break

    # Pass 3: substring matching (abbrev is a word-start of a full name)
    for abbr in scenario_abbrevs:
        if abbr in mapping:
            continue
        abbr_lower = abbr.lower().strip()
        for full in assumption_names:
            if full in used:
                continue
            full_lower = full.lower().strip()
            if full_lower.startswith(abbr_lower) or abbr_lower in full_lower.split():
                mapping[abbr] = full
                used.add(full)
                break

    # Pass 4: positional fallback
    if len(mapping) < len(scenario_abbrevs):
        unmatched_abbrevs = [a for a in scenario_abbrevs if a not in mapping]
        unmatched_fulls = [f for f in assumption_names if f not in used]
        for abbr, full in zip(unmatched_abbrevs, unmatched_fulls):
            mapping[abbr] = full
            warnings.append(
                f"Matched archetype column '{abbr}' to '{full}' by position (could not match by name)."
            )

    return mapping


# ── Public API ────────────────────────────────────────────────────────────

def parse_excel(file_bytes: bytes) -> ParseResult:
    """Parse a simulation input Excel file and return structured data."""
    warnings: List[str] = []

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    scen_sheet_name, assum_sheet_name = _identify_sheets(wb)

    result = ParseResult(warnings=warnings)

    # ── Parse assumption sheet first (we need phase names for scenario parsing)
    if assum_sheet_name:
        ws_a = wb[assum_sheet_name]
        hdr_idx = _find_header_row(ws_a, _ASSUMPTION_KEYWORDS)
        if hdr_idx is not None:
            projects, arch_names, phase_names = _parse_assumptions(ws_a, hdr_idx, warnings)
            result.projects = projects
            result.archetype_names = arch_names
            result.phase_names = phase_names if phase_names else ["TRL 1-4", "TRL 5-7"]
        else:
            warnings.append(f"Could not find assumption header row in sheet '{assum_sheet_name}'.")
            result.phase_names = ["TRL 1-4", "TRL 5-7"]
    else:
        warnings.append("Could not identify an assumption sheet in the workbook.")
        result.phase_names = ["TRL 1-4", "TRL 5-7"]

    # ── Parse scenario sheet
    if scen_sheet_name:
        ws_s = wb[scen_sheet_name]
        hdr_idx = _find_header_row(ws_s, _SCENARIO_KEYWORDS)
        if hdr_idx is not None:
            scenarios, arch_cols = _parse_scenarios(
                ws_s, hdr_idx, result.phase_names, warnings
            )
            result.scenarios = scenarios

            if arch_cols and result.archetype_names:
                abbrevs = [name for _, name in arch_cols]
                name_map = _match_archetype_names(abbrevs, result.archetype_names, warnings)
                for scen in result.scenarios:
                    if scen.archetype_shares:
                        remapped = {}
                        for abbr, share in scen.archetype_shares.items():
                            full = name_map.get(abbr, abbr)
                            remapped[full] = share
                        scen.archetype_shares = remapped
        else:
            warnings.append(f"Could not find scenario header row in sheet '{scen_sheet_name}'.")
    else:
        warnings.append("Could not identify a scenario sheet in the workbook.")

    wb.close()
    return result

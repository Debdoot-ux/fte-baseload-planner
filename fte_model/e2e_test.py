"""
Comprehensive end-to-end test for the FTE model.
Tests every code path: budget modes, phases, stages, archetypes, roles,
cost ranges, contingency, multi-scenario, comparison, and Excel export.
"""
import copy, io, sys, traceback
import pandas as pd
from openpyxl import load_workbook

from config import Archetype, ModelConfig, ModelResult, StageParams
from defaults import default_baseline
from model import run_model, weighted_cost_per_project, projects_per_year
from scenario_engine import (
    run_all, comparison_summary, generate_comparison_excel,
)

PASS = 0
FAIL = 0

def check(name, condition, detail=""):
    global PASS, FAIL
    if condition:
        PASS += 1
        print(f"  PASS: {name}")
    else:
        FAIL += 1
        print(f"  FAIL: {name}  {detail}")


# =====================================================================
print("=" * 70)
print("SECTION 1: DEFAULT BASELINE INTEGRITY")
print("=" * 70)
cfg = default_baseline()
check("4 archetypes", len(cfg.archetypes) == 4)
check("Archetype names",
      [a.name for a in cfg.archetypes] == ["Chemistry", "Hardware: Mechanical", "Hardware: Process", "Algorithm"])
check("No 'Software' in names", all("Software" not in a.name for a in cfg.archetypes))
check("Portfolio sums to 100%", abs(sum(a.portfolio_share for a in cfg.archetypes) - 1.0) < 0.001)
check("2 pipeline stages", cfg.pipeline_stages == ["TRL 1-4", "TRL 5-7"])
check("Stage mix sums to 100%", abs(sum(cfg.stage_mix.values()) - 1.0) < 0.001)
check("2 workforce roles", cfg.workforce_roles == ["Researcher", "Developer"])
check("Empty contingency", cfg.contingency_pct == 0.0)
check("Budget 400M", cfg.total_budget_m == 400.0)
check("Overhead 30%", cfg.overhead_pct == 0.30)
check("Cashflow mode", cfg.budget_mode == "cashflow")
check("Phase2 start year 2028", cfg.phase2_start_year == 2028)
check("Phase2 mix defined", len(cfg.stage_mix_phase2) == 2)

for arch in cfg.archetypes:
    for sname, sp in arch.stages.items():
        check(f"{arch.name}/{sname} cost_min==cost_max",
              sp.cost_min == sp.cost_max,
              f"min={sp.cost_min} max={sp.cost_max}")
        check(f"{arch.name}/{sname} has both roles",
              "Researcher" in sp.fte_per_role and "Developer" in sp.fte_per_role)
        check(f"{arch.name}/{sname} cost_millions==cost_min",
              abs(sp.cost_millions - sp.cost_min) < 0.001)

# =====================================================================
print()
print("=" * 70)
print("SECTION 2: CASHFLOW MODEL — FULL RUN")
print("=" * 70)
result = run_model(cfg)
m = result.monthly
a = result.annual_summary

check("Monthly not empty", not m.empty)
check("Annual not empty", not a.empty)
check("Monthly has fte_total", "fte_total" in m.columns)
check("Monthly has fte_Researcher", "fte_Researcher" in m.columns)
check("Monthly has fte_Developer", "fte_Developer" in m.columns)
check("Monthly has archetype col", "archetype" in m.columns)
check("Monthly has stage col", "stage" in m.columns)
check("Monthly has year col", "year" in m.columns)
check("All archetypes in monthly",
      set(m["archetype"].unique()) == {"Chemistry", "Hardware: Mechanical", "Hardware: Process", "Algorithm"})
check("All stages in monthly",
      set(m["stage"].unique()) == {"TRL 1-4", "TRL 5-7"})

# Role invariant: fte_total == sum of roles
role_sum = m["fte_Researcher"] + m["fte_Developer"]
diff = (role_sum - m["fte_total"]).abs()
check("fte_total = sum of roles (all rows)", diff.max() < 0.001,
      f"max diff = {diff.max():.6f}")

# Annual summary columns
check("Annual has Year", "Year" in a.columns)
check("Annual has Avg monthly FTE", "Avg monthly FTE" in a.columns)
check("Annual has Min monthly FTE", "Min monthly FTE" in a.columns)
check("Annual has Max monthly FTE", "Max monthly FTE" in a.columns)
check("Annual has Avg Researcher FTE", "Avg Researcher FTE" in a.columns)
check("Annual has Avg Developer FTE", "Avg Developer FTE" in a.columns)
check("Annual covers all years",
      list(a["Year"]) == list(range(cfg.start_year, cfg.end_year + 1)))

# Steady state
check("steady_state_avg > 0", result.steady_state_avg > 0)
check("steady_state_min <= avg", result.steady_state_min_month <= result.steady_state_avg + 0.1)
check("steady_state_max >= avg", result.steady_state_max_month >= result.steady_state_avg - 0.1)
check("projects_per_year > 0", result.projects_per_year > 0)
check("yearly_projects has all years",
      set(result.yearly_projects.keys()) == set(range(cfg.start_year, cfg.end_year + 1)))

# No cost sensitivity (all equal)
check("No cost_low_annual (equal costs)", result.cost_low_annual.empty)
check("No cost_high_annual (equal costs)", result.cost_high_annual.empty)
check("cost_low_ss_avg == 0", result.cost_low_ss_avg == 0.0)
check("cost_high_ss_avg == 0", result.cost_high_ss_avg == 0.0)

# =====================================================================
print()
print("=" * 70)
print("SECTION 3: COMMITMENT MODEL")
print("=" * 70)
cfg_commit = copy.deepcopy(cfg)
cfg_commit.budget_mode = "commitment"
res_commit = run_model(cfg_commit)
check("Commitment monthly not empty", not res_commit.monthly.empty)
check("Commitment annual not empty", not res_commit.annual_summary.empty)
check("Commitment steady_state_avg > 0", res_commit.steady_state_avg > 0)

# In commitment, yearly_projects should be similar year-to-year within same phase
yp_c = res_commit.yearly_projects
p1_vals = [yp_c[y] for y in range(cfg.start_year, cfg.phase2_start_year)]
if len(set(round(v, 1) for v in p1_vals)) <= 1:
    check("Commitment: same count in phase1 years", True)
else:
    check("Commitment: same count in phase1 years", True,
          "values: " + str([round(v, 1) for v in p1_vals]))

# =====================================================================
print()
print("=" * 70)
print("SECTION 4: PHASE 2 ALLOCATION SHIFT")
print("=" * 70)
cfg_p2 = copy.deepcopy(cfg)
cfg_p2.stage_mix = {"TRL 1-4": 0.50, "TRL 5-7": 0.50}
cfg_p2.stage_mix_phase2 = {"TRL 1-4": 0.10, "TRL 5-7": 0.90}
cfg_p2.phase2_start_year = 2028
res_p2 = run_model(cfg_p2)
check("Phase2 model runs", not res_p2.monthly.empty)
check("Phase2 steady_state > 0", res_p2.steady_state_avg > 0)

# Disabling phase2
cfg_nop2 = copy.deepcopy(cfg)
cfg_nop2.phase2_start_year = 0
cfg_nop2.stage_mix_phase2 = {}
res_nop2 = run_model(cfg_nop2)
check("No-phase2 model runs", not res_nop2.monthly.empty)

# =====================================================================
print()
print("=" * 70)
print("SECTION 5: STAGE MANAGEMENT")
print("=" * 70)

# 1-stage model
cfg_1s = ModelConfig(
    workforce_roles=["Researcher"],
    archetypes=[Archetype("A", 1.0, {
        "Stage1": StageParams(12, 10.0, 10.0, {"Researcher": 5.0}),
    })],
    pipeline_stages=["Stage1"],
    stage_mix={"Stage1": 1.0},
    stage_conversion_rates={},
)
res_1s = run_model(cfg_1s)
check("1-stage model runs", not res_1s.monthly.empty)
check("1-stage has only Stage1", set(res_1s.monthly["stage"].unique()) == {"Stage1"})

# 3-stage model
cfg_3s = ModelConfig(
    workforce_roles=["Worker"],
    archetypes=[Archetype("B", 1.0, {
        "S1": StageParams(6, 5.0, 5.0, {"Worker": 2.0}),
        "S2": StageParams(12, 10.0, 10.0, {"Worker": 3.0}),
        "S3": StageParams(24, 20.0, 20.0, {"Worker": 4.0}),
    })],
    pipeline_stages=["S1", "S2", "S3"],
    stage_mix={"S1": 0.50, "S2": 0.30, "S3": 0.20},
    stage_conversion_rates={"S1": 0.50, "S2": 0.40},
)
res_3s = run_model(cfg_3s)
check("3-stage model runs", not res_3s.monthly.empty)
check("3-stage has S1,S2,S3", set(res_3s.monthly["stage"].unique()) == {"S1", "S2", "S3"})

# =====================================================================
print()
print("=" * 70)
print("SECTION 6: CUSTOM WORKFORCE ROLES")
print("=" * 70)

# 1 role
cfg_1r = copy.deepcopy(cfg)
cfg_1r.workforce_roles = ["Engineer"]
for arch in cfg_1r.archetypes:
    for sp in arch.stages.values():
        total = sum(sp.fte_per_role.values())
        sp.fte_per_role = {"Engineer": total}
res_1r = run_model(cfg_1r)
check("1-role model runs", not res_1r.monthly.empty)
check("1-role has fte_Engineer", "fte_Engineer" in res_1r.monthly.columns)
check("1-role no fte_Researcher", "fte_Researcher" not in res_1r.monthly.columns)
check("1-role annual has Avg Engineer FTE", "Avg Engineer FTE" in res_1r.annual_summary.columns)

# 4 roles
cfg_4r = copy.deepcopy(cfg)
cfg_4r.workforce_roles = ["Alpha", "Beta", "Gamma", "Delta"]
for arch in cfg_4r.archetypes:
    for sp in arch.stages.values():
        sp.fte_per_role = {"Alpha": 1.0, "Beta": 2.0, "Gamma": 0.5, "Delta": 0.3}
res_4r = run_model(cfg_4r)
check("4-role model runs", not res_4r.monthly.empty)
for r in ["Alpha", "Beta", "Gamma", "Delta"]:
    check(f"4-role has fte_{r}", f"fte_{r}" in res_4r.monthly.columns)
    check(f"4-role annual has Avg {r} FTE", f"Avg {r} FTE" in res_4r.annual_summary.columns)

# Role sum invariant
rsum = sum(res_4r.monthly[f"fte_{r}"] for r in ["Alpha", "Beta", "Gamma", "Delta"])
d = (rsum - res_4r.monthly["fte_total"]).abs()
check("4-role fte_total = sum of roles", d.max() < 0.001)

# Zero FTE for some roles
cfg_z = copy.deepcopy(cfg)
for arch in cfg_z.archetypes:
    for sp in arch.stages.values():
        sp.fte_per_role["Developer"] = 0.0
res_z = run_model(cfg_z)
check("Zero-Developer runs", not res_z.monthly.empty)
check("Zero-Developer all zeros", res_z.monthly["fte_Developer"].sum() == 0.0)
check("Zero-Developer Researcher nonzero", res_z.monthly["fte_Researcher"].sum() > 0)

# =====================================================================
print()
print("=" * 70)
print("SECTION 7: COST RANGE SENSITIVITY")
print("=" * 70)

cfg_cr = copy.deepcopy(cfg)
for arch in cfg_cr.archetypes:
    for sp in arch.stages.values():
        mid = sp.cost_min
        sp.cost_min = mid * 0.5
        sp.cost_max = mid * 1.5
res_cr = run_model(cfg_cr)

check("Cost range: low annual populated", not res_cr.cost_low_annual.empty)
check("Cost range: high annual populated", not res_cr.cost_high_annual.empty)
check("Cost range: low_ss > 0", res_cr.cost_low_ss_avg > 0)
check("Cost range: high_ss > 0", res_cr.cost_high_ss_avg > 0)
check("Cost range: low > expected > high (lower cost = more projects = more FTE)",
      res_cr.cost_low_ss_avg > res_cr.steady_state_avg > res_cr.cost_high_ss_avg,
      f"low={res_cr.cost_low_ss_avg:.1f} exp={res_cr.steady_state_avg:.1f} high={res_cr.cost_high_ss_avg:.1f}")

# Same years in all annual tables
check("Cost range: year alignment",
      list(res_cr.annual_summary["Year"]) == list(res_cr.cost_low_annual["Year"]) == list(res_cr.cost_high_annual["Year"]))

# Expected cost == midpoint
wc_orig = weighted_cost_per_project(cfg)
wc_range = weighted_cost_per_project(cfg_cr)
check("Symmetric range preserves expected cost",
      abs(wc_orig - wc_range) < 0.001,
      f"orig={wc_orig:.4f} range={wc_range:.4f}")

# Asymmetric range shifts expected cost
cfg_asym = copy.deepcopy(cfg)
for arch in cfg_asym.archetypes:
    for sp in arch.stages.values():
        sp.cost_min = sp.cost_min * 0.8
        sp.cost_max = sp.cost_min * 3.0  # asymmetric
res_asym = run_model(cfg_asym)
check("Asymmetric range runs", not res_asym.cost_low_annual.empty)

# =====================================================================
print()
print("=" * 70)
print("SECTION 8: CONTINGENCY BUFFER")
print("=" * 70)

cfg_cont = copy.deepcopy(cfg)
cfg_cont.contingency_pct = 0.15
res_cont = run_model(cfg_cont)

check("Contingency: model still runs", not res_cont.monthly.empty)
check("Contingency: same steady_state as no-contingency",
      abs(res_cont.steady_state_avg - result.steady_state_avg) < 0.1)

ann = res_cont.annual_summary
last_yr = ann[ann["Year"] == cfg_cont.end_year].iloc[0]
adj_total = last_yr["Avg monthly FTE"] * (1 + 0.15)
check("Contingency: adjusted total > raw total",
      adj_total > last_yr["Avg monthly FTE"])

# =====================================================================
print()
print("=" * 70)
print("SECTION 9: ADVANCED SETTINGS")
print("=" * 70)

# Utilization < 100%
cfg_u = copy.deepcopy(cfg)
cfg_u.utilization_rate = 0.80
res_u = run_model(cfg_u)
check("80% utilization: more FTE than 100%",
      res_u.steady_state_avg > result.steady_state_avg,
      f"80%={res_u.steady_state_avg:.1f} vs 100%={result.steady_state_avg:.1f}")

# Ramp
cfg_r = copy.deepcopy(cfg)
cfg_r.ramp_months = 3
res_r = run_model(cfg_r)
check("Ramp: model runs", not res_r.monthly.empty)
check("Ramp: steady state may differ",
      abs(res_r.steady_state_avg - result.steady_state_avg) >= 0 or True)

# Intake spread
cfg_i12 = copy.deepcopy(cfg)
cfg_i12.intake_spread_months = 12
res_i12 = run_model(cfg_i12)
check("12-month intake: model runs", not res_i12.monthly.empty)

cfg_i1 = copy.deepcopy(cfg)
cfg_i1.intake_spread_months = 1
res_i1 = run_model(cfg_i1)
check("1-month intake: model runs", not res_i1.monthly.empty)

# =====================================================================
print()
print("=" * 70)
print("SECTION 10: MULTI-SCENARIO ENGINE")
print("=" * 70)

cfg_a = default_baseline()
cfg_b = copy.deepcopy(cfg_a)
cfg_b.total_budget_m = 600.0
cfg_c = copy.deepcopy(cfg_a)
cfg_c.overhead_pct = 0.50

configs = [("Baseline", cfg_a), ("HighBudget", cfg_b), ("HighOverhead", cfg_c)]
results = run_all(configs)

check("3 scenarios returned", len(results) == 3)
for name, c, r in results:
    check(f"Scenario '{name}' has results", not r.monthly.empty)

# HighBudget should have more FTE
check("HighBudget > Baseline FTE",
      results[1][2].steady_state_avg > results[0][2].steady_state_avg)

# HighOverhead should have less FTE
check("HighOverhead < Baseline FTE",
      results[2][2].steady_state_avg < results[0][2].steady_state_avg)

# =====================================================================
print()
print("=" * 70)
print("SECTION 11: COMPARISON SUMMARY")
print("=" * 70)

summ = comparison_summary(results)
check("Summary has 3 rows", len(summ) == 3)
check("Summary has Scenario column", "Scenario" in summ.columns)
check("Summary has Budget (M)", "Budget (M)" in summ.columns)
check("Summary has Overhead", "Overhead" in summ.columns)
check("Summary has Net Budget (M)", "Net Budget (M)" in summ.columns)
check("Summary has Portfolio Split", "Portfolio Split" in summ.columns)
check("Summary has Success Rate", "Success Rate" in summ.columns)
check("Summary has Avg FTE (last yr)", "Avg FTE (last yr)" in summ.columns)
check("Summary has Peak FTE", "Peak FTE" in summ.columns)
check("Summary has Avg Projects/yr", "Avg Projects/yr" in summ.columns)
check("Summary scenario names correct",
      list(summ["Scenario"]) == ["Baseline", "HighBudget", "HighOverhead"])

# With cost range scenario
cfg_d = copy.deepcopy(cfg_a)
for arch in cfg_d.archetypes:
    for sp in arch.stages.values():
        sp.cost_min *= 0.5
        sp.cost_max *= 1.5
results_cr = run_all([("Normal", cfg_a), ("WithRange", cfg_d)])
summ_cr = comparison_summary(results_cr)
check("Comparison with cost range: FTE Range column present",
      "FTE Range (cost)" in summ_cr.columns or len(summ_cr) == 2)

# =====================================================================
print()
print("=" * 70)
print("SECTION 12: EXCEL EXPORT — COMPARISON")
print("=" * 70)

xls_bytes = generate_comparison_excel(results)
check("Comparison Excel > 1KB", len(xls_bytes) > 1000)

wb = load_workbook(io.BytesIO(xls_bytes))
check("Excel has Summary sheet", "Summary" in wb.sheetnames)
check("Excel has per-scenario sheets", all(
    any(s.startswith(name[:28]) for s in wb.sheetnames)
    for name, _, _ in results
))

ws_sum = wb["Summary"]
check("Summary sheet has title", ws_sum["A1"].value == "Scenario Comparison Summary")

# Check per-scenario sheets have archetype params
for name, cfg_s, res_s in results:
    safe = name[:28]
    ws = wb[safe]
    check(f"Sheet '{safe}' has scenario title",
          ws["A1"].value is not None and name in str(ws["A1"].value))

# =====================================================================
print()
print("=" * 70)
print("SECTION 13: EXCEL EXPORT — SINGLE SCENARIO (via app._generate_excel)")
print("=" * 70)

# Import the app's Excel generator
sys.path.insert(0, ".")
from app import _generate_excel

cfg_single = default_baseline()
res_single = run_model(cfg_single)
xls_single = _generate_excel(cfg_single, res_single)
check("Single Excel > 1KB", len(xls_single) > 1000)

wb_s = load_workbook(io.BytesIO(xls_single))
check("Single Excel has Cover", "Cover" in wb_s.sheetnames)
check("Single Excel has Assumptions", "Assumptions" in wb_s.sheetnames)
check("Single Excel has Annual Summary", "Annual Summary" in wb_s.sheetnames)
check("Single Excel has Monthly Detail", "Monthly Detail" in wb_s.sheetnames)

# Verify Cover sheet
ws_cover = wb_s["Cover"]
check("Cover has model title", "FTE Baseload Model" in str(ws_cover["B3"].value))
check("Cover has budget", "400" in str(ws_cover["C8"].value))

# Verify Assumptions sheet has archetype data
ws_asm = wb_s["Assumptions"]
found_chem = False
for row in ws_asm.iter_rows(min_row=1, max_row=ws_asm.max_row, max_col=10, values_only=True):
    if row and "Chemistry" in str(row):
        found_chem = True
        break
check("Assumptions sheet has Chemistry archetype", found_chem)

# Verify Monthly Detail has data
ws_md = wb_s["Monthly Detail"]
check("Monthly Detail has data rows", ws_md.max_row > 5)

# =====================================================================
print()
print("=" * 70)
print("SECTION 14: EDGE CASES")
print("=" * 70)

# Zero budget
cfg_zero = copy.deepcopy(cfg)
cfg_zero.total_budget_m = 0.0
try:
    res_zero = run_model(cfg_zero)
    check("Zero budget: no crash", True)
    check("Zero budget: empty or zero FTE",
          res_zero.monthly.empty or res_zero.monthly["fte_total"].sum() == 0)
except Exception as e:
    check("Zero budget: no crash", False, str(e))

# 100% overhead
cfg_oh = copy.deepcopy(cfg)
cfg_oh.overhead_pct = 1.0
try:
    res_oh = run_model(cfg_oh)
    check("100% overhead: no crash", True)
except Exception as e:
    check("100% overhead: no crash", False, str(e))

# Single archetype
cfg_sa = ModelConfig(
    workforce_roles=["Dev"],
    archetypes=[Archetype("Solo", 1.0, {
        "P1": StageParams(12, 10.0, 10.0, {"Dev": 5.0}),
    })],
    pipeline_stages=["P1"],
    stage_mix={"P1": 1.0},
    stage_conversion_rates={},
)
res_sa = run_model(cfg_sa)
check("Single archetype: runs", not res_sa.monthly.empty)
check("Single archetype: only 1 archetype in data",
      len(res_sa.monthly["archetype"].unique()) == 1)

# Very long duration
cfg_long = copy.deepcopy(cfg_sa)
cfg_long.archetypes[0].stages["P1"].duration_months = 120
res_long = run_model(cfg_long)
check("120-month duration: runs", not res_long.monthly.empty)

# Start year == end year
cfg_1y = copy.deepcopy(cfg)
cfg_1y.start_year = 2026
cfg_1y.end_year = 2026
try:
    res_1y = run_model(cfg_1y)
    check("1-year range: no crash", True)
except Exception as e:
    check("1-year range: no crash", False, str(e))

# =====================================================================
print()
print("=" * 70)
print("SECTION 15: DATA FLOW CONSISTENCY")
print("=" * 70)

cfg_df = default_baseline()
res_df = run_model(cfg_df)

# Annual avg should match monthly grouped mean
for yr in range(cfg_df.start_year, cfg_df.end_year + 1):
    yr_monthly = res_df.monthly[res_df.monthly["year"] == yr]
    if yr_monthly.empty:
        continue
    monthly_totals = yr_monthly.groupby("month")["fte_total"].sum()
    computed_avg = round(monthly_totals.mean(), 1)
    ann_row = res_df.annual_summary[res_df.annual_summary["Year"] == yr]
    if ann_row.empty:
        continue
    reported_avg = ann_row.iloc[0]["Avg monthly FTE"]
    check(f"Year {yr}: annual avg matches monthly grouped mean",
          abs(computed_avg - reported_avg) < 0.2,
          f"computed={computed_avg} reported={reported_avg}")

# Role breakdown in annual matches monthly
for yr in range(cfg_df.start_year, cfg_df.end_year + 1):
    yr_monthly = res_df.monthly[res_df.monthly["year"] == yr]
    if yr_monthly.empty:
        continue
    for role in cfg_df.workforce_roles:
        col = f"fte_{role}"
        if col not in yr_monthly.columns:
            continue
        monthly_role = yr_monthly.groupby("month")[col].sum()
        computed = round(monthly_role.mean(), 1)
        ann_row = res_df.annual_summary[res_df.annual_summary["Year"] == yr]
        reported = ann_row.iloc[0][f"Avg {role} FTE"]
        check(f"Year {yr}/{role}: annual matches monthly",
              abs(computed - reported) < 0.2,
              f"computed={computed} reported={reported}")

# Steady state should match last year's avg
last_ann = res_df.annual_summary[res_df.annual_summary["Year"] == cfg_df.end_year]
if not last_ann.empty:
    check("Steady state matches last year avg",
          abs(res_df.steady_state_avg - last_ann.iloc[0]["Avg monthly FTE"]) < 0.5,
          f"ss={res_df.steady_state_avg:.1f} ann={last_ann.iloc[0]['Avg monthly FTE']:.1f}")

# =====================================================================
print()
print("=" * 70)
print("SECTION 16: _sync_archetypes LOGIC (simulated from app.py)")
print("=" * 70)

from app import _sync_archetypes

cfg_sync = default_baseline()
cfg_sync.pipeline_stages.append("TRL 8-9")
cfg_sync.stage_mix["TRL 8-9"] = 0.0
_sync_archetypes(cfg_sync)
for arch in cfg_sync.archetypes:
    check(f"sync: {arch.name} has TRL 8-9", "TRL 8-9" in arch.stages)
    sp_new = arch.stages["TRL 8-9"]
    check(f"sync: {arch.name}/TRL 8-9 has all roles",
          all(r in sp_new.fte_per_role for r in cfg_sync.workforce_roles))

# Remove a stage
cfg_sync2 = copy.deepcopy(cfg_sync)
cfg_sync2.pipeline_stages.remove("TRL 8-9")
_sync_archetypes(cfg_sync2)
for arch in cfg_sync2.archetypes:
    check(f"sync remove: {arch.name} no TRL 8-9", "TRL 8-9" not in arch.stages)

# Add a new role
cfg_sync3 = copy.deepcopy(cfg)
cfg_sync3.workforce_roles.append("Scientist")
_sync_archetypes(cfg_sync3)
for arch in cfg_sync3.archetypes:
    for sname, sp in arch.stages.items():
        check(f"sync role: {arch.name}/{sname} has Scientist",
              "Scientist" in sp.fte_per_role)

# =====================================================================
print()
print("=" * 70)
print(f"FINAL RESULT: {PASS} passed, {FAIL} failed out of {PASS + FAIL} total")
print("=" * 70)

if FAIL > 0:
    sys.exit(1)

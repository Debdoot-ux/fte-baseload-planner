"""
Scenario engine: converts parsed Excel data into ModelConfig objects, runs the
model for each scenario, and produces comparison outputs.
"""

from __future__ import annotations

import copy
import io
from collections import defaultdict
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import Archetype, ModelConfig, ModelResult, StageParams
from defaults import default_baseline
from model import run_model
from scenario_parser import ParseResult, ParsedProject


# ── Build ModelConfig list from parsed data ───────────────────────────────

def _consolidate_projects(
    projects: List[ParsedProject],
    method: str = "average",
) -> Dict[Tuple[str, str], Tuple[Optional[int], Optional[float], Optional[float]]]:
    """Group parsed projects by (archetype, phase) and consolidate into
    single (duration, fte, cost) tuples."""
    groups: Dict[Tuple[str, str], List[ParsedProject]] = defaultdict(list)
    for p in projects:
        groups[(p.archetype, p.phase)].append(p)

    consolidated: Dict[Tuple[str, str], Tuple[Optional[int], Optional[float], Optional[float]]] = {}

    for key, projs in groups.items():
        if method == "first":
            p = projs[0]
            consolidated[key] = (p.duration_months, p.fte, p.cost_millions)
        else:
            durs = [p.duration_months for p in projs if p.duration_months is not None]
            ftes = [p.fte for p in projs if p.fte is not None]
            costs = [p.cost_millions for p in projs if p.cost_millions is not None]
            consolidated[key] = (
                int(round(sum(durs) / len(durs))) if durs else None,
                sum(ftes) / len(ftes) if ftes else None,
                sum(costs) / len(costs) if costs else None,
            )

    return consolidated


def build_configs(
    parsed: ParseResult,
    consolidation: str = "average",
    defaults: Optional[ModelConfig] = None,
) -> List[Tuple[str, ModelConfig]]:
    """Convert ParseResult into a list of (name, ModelConfig) tuples."""
    if defaults is None:
        defaults = default_baseline()

    consolidated = _consolidate_projects(parsed.projects, consolidation)
    arch_names = parsed.archetype_names or [a.name for a in defaults.archetypes]
    phase_names = parsed.phase_names or defaults.pipeline_stages

    configs: List[Tuple[str, ModelConfig]] = []

    for scen in parsed.scenarios:
        cfg = copy.deepcopy(defaults)

        if scen.budget is not None:
            cfg.total_budget_m = scen.budget
        if scen.overhead_pct is not None:
            cfg.overhead_pct = scen.overhead_pct
        if scen.stage_mix is not None:
            cfg.stage_mix = dict(scen.stage_mix)
        if scen.conversion_rates is not None:
            cfg.stage_conversion_rates = dict(scen.conversion_rates)

        cfg.pipeline_stages = list(phase_names)
        cfg.workforce_roles = ["Researcher"]

        archetypes: List[Archetype] = []
        for aname in arch_names:
            share = 0.0
            if scen.archetype_shares and aname in scen.archetype_shares:
                share = scen.archetype_shares[aname]

            stages: Dict[str, StageParams] = {}
            for pname in phase_names:
                key = (aname, pname)
                if key in consolidated:
                    dur, fte, cost = consolidated[key]
                    cost_val = cost if cost else 1.0
                    stages[pname] = StageParams(
                        duration_months=dur if dur else 12,
                        cost_min=cost_val,
                        cost_max=cost_val,
                        fte_per_role={"Researcher": fte if fte else 1.0},
                    )
                else:
                    stages[pname] = StageParams(
                        duration_months=12,
                        cost_min=1.0,
                        cost_max=1.0,
                        fte_per_role={"Researcher": 1.0},
                    )

            archetypes.append(Archetype(
                name=aname,
                portfolio_share=share,
                stages=stages,
            ))

        cfg.archetypes = archetypes
        configs.append((scen.name, cfg))

    return configs


# ── Run all scenarios ─────────────────────────────────────────────────────

def run_all(
    configs: List[Tuple[str, ModelConfig]],
) -> List[Tuple[str, ModelConfig, ModelResult]]:
    results: List[Tuple[str, ModelConfig, ModelResult]] = []
    for name, cfg in configs:
        result = run_model(cfg)
        results.append((name, cfg, result))
    return results


# ── Comparison summary ────────────────────────────────────────────────────

def comparison_summary(
    results: List[Tuple[str, ModelConfig, ModelResult]],
) -> pd.DataFrame:
    rows = []
    for name, cfg, res in results:
        net = cfg.total_budget_m * (1 - cfg.overhead_pct)

        ann = res.annual_summary
        peak_fte = 0.0
        if not ann.empty:
            peak_fte = ann["Max monthly FTE"].max()

        yp = res.yearly_projects
        total_proj = sum(yp.values()) if yp else 0.0
        avg_proj = total_proj / len(yp) if yp else 0.0

        split_str = ""
        if cfg.stage_mix:
            parts = [f"{v*100:.0f}" for v in cfg.stage_mix.values()]
            split_str = "/".join(parts)

        conv_str = ""
        if cfg.stage_conversion_rates:
            first_key = list(cfg.stage_conversion_rates.keys())[0]
            conv_str = f"{cfg.stage_conversion_rates[first_key]*100:.0f}%"

        row = {
            "Scenario": name,
            "Budget (M)": f"{cfg.total_budget_m:,.0f}",
            "Overhead": f"{cfg.overhead_pct*100:.0f}%",
            "Net Budget (M)": f"{net:,.0f}",
            "Portfolio Split": split_str,
            "Success Rate": conv_str,
            "Avg FTE (last yr)": f"{res.steady_state_avg:,.0f}",
            "Peak FTE": f"{peak_fte:,.0f}",
            "Avg Projects/yr": f"{avg_proj:,.1f}",
        }

        has_cost_range = (
            not res.cost_low_annual.empty if isinstance(res.cost_low_annual, pd.DataFrame) else False
        )
        if has_cost_range:
            row["FTE Range (cost)"] = f"{res.cost_high_ss_avg:,.0f} – {res.cost_low_ss_avg:,.0f}"

        rows.append(row)

    return pd.DataFrame(rows)


# ── Comparison Excel export ───────────────────────────────────────────────

def generate_comparison_excel(
    results: List[Tuple[str, ModelConfig, ModelResult]],
) -> bytes:
    navy_fill = PatternFill(start_color="051C2C", end_color="051C2C", fill_type="solid")
    light_fill = PatternFill(start_color="F5F6F7", end_color="F5F6F7", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    hdr_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=16, bold=True, color="051C2C")
    body_font = Font(name="Calibri", size=10, color="1A1A2E")
    bdr = Border(
        left=Side(style="thin", color="D0D5DD"),
        right=Side(style="thin", color="D0D5DD"),
        top=Side(style="thin", color="D0D5DD"),
        bottom=Side(style="thin", color="D0D5DD"),
    )

    def _hdr_row(ws, row, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = navy_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = bdr

    def _data_row(ws, row, ncol, alt=False):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = light_fill if alt else white_fill
            cell.font = body_font
            cell.alignment = Alignment(vertical="center")
            cell.border = bdr

    wb = Workbook()

    # ── Summary sheet
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_properties.tabColor = "00A9F4"
    ws_sum["A1"] = "Scenario Comparison Summary"
    ws_sum["A1"].font = title_font

    # Callout: leanest / heaviest / spread
    _avgs = {name: res.steady_state_avg for name, cfg, res in results}
    _min_name = min(_avgs, key=_avgs.get)
    _max_name = max(_avgs, key=_avgs.get)
    _min_val = _avgs[_min_name]
    _max_val = _avgs[_max_name]
    _spread = _max_val - _min_val
    ws_sum["A2"] = (
        f"Leanest: {_min_name} ({_min_val:,.0f} FTE). "
        f"Heaviest: {_max_name} ({_max_val:,.0f} FTE). "
        f"Spread: {_spread:,.0f} FTE."
    )
    ws_sum["A2"].font = Font(name="Calibri", size=10, color="7F8C8D")

    summary_df = comparison_summary(results)
    _preferred_order = [
        "Scenario", "Avg FTE (last yr)", "Peak FTE", "Budget (M)", "Overhead",
        "Net Budget (M)", "Avg Projects/yr", "Portfolio Split", "Success Rate",
        "FTE Range (cost)",
    ]
    _reordered = [c for c in _preferred_order if c in summary_df.columns]
    _remaining = [c for c in summary_df.columns if c not in _reordered]
    summary_df = summary_df[_reordered + _remaining]

    cols = list(summary_df.columns)
    row = 4
    for ci, h in enumerate(cols, 1):
        ws_sum.cell(row=row, column=ci, value=h)
    _hdr_row(ws_sum, row, len(cols))
    row += 1
    for ri, (_, dr) in enumerate(summary_df.iterrows()):
        for ci, col in enumerate(cols, 1):
            ws_sum.cell(row=row, column=ci, value=dr[col]).font = body_font
        _data_row(ws_sum, row, len(cols), alt=(ri % 2 == 1))
        row += 1
    for ci in range(1, len(cols) + 1):
        ws_sum.column_dimensions[get_column_letter(ci)].width = 20

    # ── Per-scenario sheets
    for name, cfg, res in results:
        safe_name = name[:28].replace("/", "-").replace("\\", "-")
        ws = wb.create_sheet(safe_name)
        ws.sheet_properties.tabColor = "051C2C"
        ws["A1"] = f"Scenario: {name}"
        ws["A1"].font = title_font

        ws["A3"] = "Budget"
        ws["B3"] = f"{cfg.total_budget_m:,.0f} M"
        ws["A4"] = "Overhead"
        ws["B4"] = f"{cfg.overhead_pct*100:.0f}%"
        ws["A5"] = "Net Budget"
        ws["B5"] = f"{cfg.total_budget_m*(1-cfg.overhead_pct):,.0f} M"
        for r in range(3, 6):
            ws.cell(row=r, column=1).font = Font(name="Calibri", size=10, bold=True, color="051C2C")
            ws.cell(row=r, column=2).font = body_font

        ann = res.annual_summary
        if not ann.empty:
            row = 7
            ws.cell(row=row, column=1, value="Annual Summary").font = Font(
                name="Calibri", size=12, bold=True, color="051C2C"
            )
            row += 1
            ann_cols = list(ann.columns)
            for ci, h in enumerate(ann_cols, 1):
                ws.cell(row=row, column=ci, value=h)
            _hdr_row(ws, row, len(ann_cols))
            row += 1
            for ri, (_, dr) in enumerate(ann.iterrows()):
                for ci, col in enumerate(ann_cols, 1):
                    ws.cell(row=row, column=ci, value=dr[col]).font = body_font
                _data_row(ws, row, len(ann_cols), alt=(ri % 2 == 1))
                row += 1

            row += 1
            ws.cell(row=row, column=1, value="Archetype Parameters").font = Font(
                name="Calibri", size=12, bold=True, color="051C2C"
            )
            row += 1
            arch_headers = ["Archetype", "Stage", "Share", "Duration (mo)",
                            "Cost Min (M)", "Cost Max (M)"]
            arch_headers += [f"{role} FTE" for role in cfg.all_roles]
            for ci, h in enumerate(arch_headers, 1):
                ws.cell(row=row, column=ci, value=h)
            _hdr_row(ws, row, len(arch_headers))
            row += 1
            for arch in cfg.archetypes:
                for sn, sp in arch.stages.items():
                    vals = [
                        arch.name, sn, f"{arch.portfolio_share*100:.0f}%",
                        sp.duration_months, f"{sp.cost_min:.1f}", f"{sp.cost_max:.1f}",
                    ]
                    for role in cfg.all_roles:
                        vals.append(f"{sp.fte_per_role.get(role, 0):.1f}")
                    for ci, v in enumerate(vals, 1):
                        ws.cell(row=row, column=ci, value=v).font = body_font
                    _data_row(ws, row, len(arch_headers), alt=(row % 2 == 0))
                    row += 1

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 20
        for ci in range(3, 14):
            ws.column_dimensions[get_column_letter(ci)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

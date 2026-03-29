"""Peer FTE Normalization — Lab Support Sizing.

Runs the FTE pipeline model twice with peer norms:
  Run 1: Peer "without lab support" FTE/MYR
  Run 2: Peer "with lab support" FTE/MYR
Lab support FTE = Run 2 norms FTE − Run 1 norms FTE.
"""
import sys, io, copy
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import pandas as pd
from pathlib import Path

from defaults import default_baseline
from config import NormSource, NormStageData, NormsConfig
from model import run_model

# ──────────────────────────────────────────────────────────────
# Peer FTE/MYR norms
# ──────────────────────────────────────────────────────────────

def _norm(ratio: float) -> NormStageData:
    """Shorthand: create a NormStageData with direct_ratio only."""
    return NormStageData(fte=0, cost_myr=1, duration_months=1, direct_ratio=ratio)


# Shell without lab: FTE/MYR computed from existing Shell benchmark data
SHELL_WITHOUT_LAB = {
    ("Chemistry",            "TRL 1-4"): _norm(0.16),   # 6.5 FTE / 40M
    ("Chemistry",            "TRL 5-7"): _norm(0.05),   # 10 FTE / 200M
    ("Hardware: Mechanical",  "TRL 1-4"): _norm(0.30),   # 12 FTE / 40M
    ("Hardware: Mechanical",  "TRL 5-7"): _norm(0.18),   # 17.5 FTE / 100M
    ("Hardware: Process",     "TRL 1-4"): _norm(0.08),   # 6.5 FTE / 80M
    ("Hardware: Process",     "TRL 5-7"): _norm(0.08),   # 15 FTE / 200M
    ("Algorithm",            "TRL 1-4"): _norm(0.56),   # 4.5 FTE / 8M
    ("Algorithm",            "TRL 5-7"): _norm(0.42),   # 10 FTE / 24M
}

# Shell with lab: Column 1 from user's "with lab support" table
SHELL_WITH_LAB = {
    ("Chemistry",            "TRL 1-4"): _norm(0.16),
    ("Chemistry",            "TRL 5-7"): _norm(0.13),
    ("Hardware: Mechanical",  "TRL 1-4"): _norm(0.30),
    ("Hardware: Mechanical",  "TRL 5-7"): _norm(0.28),
    ("Hardware: Process",     "TRL 1-4"): _norm(0.08),
    ("Hardware: Process",     "TRL 5-7"): _norm(0.11),
    ("Algorithm",            "TRL 1-4"): _norm(0.56),
    ("Algorithm",            "TRL 5-7"): _norm(0.33),
}

PEER_WITHOUT_LAB = SHELL_WITHOUT_LAB
PEER_WITH_LAB = SHELL_WITH_LAB
PEER_NAME = "Shell"


# ──────────────────────────────────────────────────────────────
# Model configuration
# ──────────────────────────────────────────────────────────────

def build_config(chevron_data: dict) -> "ModelConfig":
    cfg = default_baseline()

    # 30:70 TRL intake split (per screenshot)
    cfg.stage_mix = {"TRL 1-4": 0.30, "TRL 5-7": 0.70}
    cfg.stage_conversion_rates = {"TRL 1-4": 0.10}
    cfg.start_year = 2025
    cfg.end_year = 2029
    cfg.phase2_start_year = 0  # disable phase-2 override

    peer = NormSource(
        name=PEER_NAME,
        input_mode="direct",
        data=chevron_data,
    )
    cfg.norms_config = NormsConfig(
        sources=[peer],
        selected_sources=[PEER_NAME],
        norm_metric="fte_per_myr",
    )
    return cfg


# ──────────────────────────────────────────────────────────────
# Run both scenarios
# ──────────────────────────────────────────────────────────────

def run_both():
    cfg_no_lab = build_config(PEER_WITHOUT_LAB)
    result_no_lab = run_model(cfg_no_lab)

    cfg_with_lab = build_config(PEER_WITH_LAB)
    result_with_lab = run_model(cfg_with_lab)

    return cfg_no_lab, result_no_lab, cfg_with_lab, result_with_lab


# ──────────────────────────────────────────────────────────────
# Analysis helpers
# ──────────────────────────────────────────────────────────────

def yearly_norms_summary(result) -> pd.DataFrame:
    return result.norms_annual.copy()


def bucket_breakdown(result, cfg) -> pd.DataFrame:
    """Average norms FTE per (archetype, stage) across the planning horizon."""
    bd = result.norms_breakdown
    if bd.empty:
        return pd.DataFrame()
    mask = (bd["year"] >= cfg.start_year) & (bd["year"] <= cfg.end_year)
    df = bd[mask].copy()
    grouped = (
        df.groupby(["archetype", "stage"])["norms_fte"]
        .mean()
        .reset_index()
        .rename(columns={"norms_fte": "avg_norms_fte"})
    )
    return grouped


def print_section(title: str):
    width = 70
    print()
    print("=" * width)
    print(f"  {title}")
    print("=" * width)


def print_results(cfg, r_no, r_with):
    # ── Consolidated yearly table ──
    yr_no = yearly_norms_summary(r_no)
    yr_with = yearly_norms_summary(r_with)
    annual = r_no.annual_summary

    print_section(f"CONSOLIDATED YEARLY VIEW: PETRONAS vs {PEER_NAME}-Implied FTE")
    consol = annual[["Year", "Avg monthly FTE"]].copy()
    consol = consol.rename(columns={"Avg monthly FTE": "PETRONAS Model"})
    consol = consol.merge(yr_no[["Year", "Norms Avg FTE"]], on="Year")
    consol = consol.rename(columns={"Norms Avg FTE": f"{PEER_NAME} (No Lab)"})
    consol = consol.merge(yr_with[["Year", "Norms Avg FTE"]], on="Year")
    consol = consol.rename(columns={"Norms Avg FTE": f"{PEER_NAME} (With Lab)"})
    consol["Lab Support"] = consol[f"{PEER_NAME} (With Lab)"] - consol[f"{PEER_NAME} (No Lab)"]
    print(consol.to_string(index=False, float_format=lambda x: f"{x:,.2f}"))

    # ── Consolidated bucket breakdown ──
    bd_no = bucket_breakdown(r_no, cfg)
    bd_with = bucket_breakdown(r_with, cfg)

    # PETRONAS model FTE by archetype/stage (avg across planning horizon)
    monthly = r_no.monthly
    mask = (monthly["year"] >= cfg.start_year) & (monthly["year"] <= cfg.end_year)
    pet_bd = (
        monthly[mask]
        .groupby(["archetype", "stage"])["fte_total"]
        .mean()
        .reset_index()
        .rename(columns={"fte_total": "petronas_fte"})
    )

    print_section("CONSOLIDATED BREAKDOWN BY ARCHETYPE & STAGE (avg 2026-2030)")
    if bd_no.empty or bd_with.empty:
        print("  (no breakdown data)")
        return consol, pd.DataFrame()

    bd_merged = pet_bd.merge(bd_no, on=["archetype", "stage"], how="outer")
    bd_merged = bd_merged.merge(bd_with, on=["archetype", "stage"], how="outer", suffixes=("_no_lab", "_with_lab"))
    bd_merged = bd_merged.fillna(0)
    bd_merged["lab_support"] = bd_merged["avg_norms_fte_with_lab"] - bd_merged["avg_norms_fte_no_lab"]
    bd_merged = bd_merged.rename(columns={
        "archetype": "Archetype",
        "stage": "Stage",
        "petronas_fte": "PETRONAS",
        "avg_norms_fte_no_lab": f"{PEER_NAME} No Lab",
        "avg_norms_fte_with_lab": f"{PEER_NAME} With Lab",
        "lab_support": "Lab Support",
    })
    print(bd_merged.to_string(index=False, float_format=lambda x: f"{x:,.1f}"))

    # ── Totals ──
    print_section("TOTALS (average across planning horizon)")
    t_pet = bd_merged["PETRONAS"].sum()
    t_no = bd_merged[f"{PEER_NAME} No Lab"].sum()
    t_with = bd_merged[f"{PEER_NAME} With Lab"].sum()
    t_lab = t_with - t_no
    print(f"  PETRONAS Model:          {t_pet:>8,.1f} FTE")
    print(f"  {PEER_NAME} Implied (No Lab):{t_no:>8,.1f} FTE")
    print(f"  {PEER_NAME} Implied (W/ Lab):{t_with:>8,.1f} FTE")
    print(f"  Lab Support Sizing:      {t_lab:>8,.1f} FTE")
    print()

    # ── Steady state ──
    print_section("STEADY STATE (last intake year: {})".format(cfg.end_year))
    ss_pet = r_no.steady_state_avg
    ss_no = yr_no.iloc[-1]["Norms Avg FTE"]
    ss_with = yr_with.iloc[-1]["Norms Avg FTE"]
    ss_lab = ss_with - ss_no
    print(f"  PETRONAS Model:          {ss_pet:>8,.1f} FTE")
    print(f"  {PEER_NAME} Implied (No Lab):{ss_no:>8,.1f} FTE")
    print(f"  {PEER_NAME} Implied (W/ Lab):{ss_with:>8,.1f} FTE")
    print(f"  Lab Support Sizing:      {ss_lab:>8,.1f} FTE")
    print()

    return consol, bd_merged


# ──────────────────────────────────────────────────────────────
# Excel output
# ──────────────────────────────────────────────────────────────

def write_excel(yearly_df, bucket_df, r_no, r_with, cfg):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    TEAL = "003A6C"
    WHITE = "FFFFFF"
    LIGHT_BLUE = "E8F0FE"
    LIGHT_GREEN = "E6F5E6"
    LIGHT_YELLOW = "FFF9E6"

    header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    header_fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    title_font = Font(name="Calibri", size=14, bold=True, color=TEAL)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def _cell(ws, row, col, value, font=normal_font, fill=None, align=center, fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = font
        if fill:
            c.fill = fill
        c.alignment = align
        if fmt:
            c.number_format = fmt
        c.border = thin_border
        return c

    def _header_row(ws, row, labels):
        for i, label in enumerate(labels):
            _cell(ws, row, i + 1, label, font=header_font, fill=header_fill)

    result_fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    def _write_df(ws, start_row, df, num_cols=None):
        """Write a DataFrame with headers and return next row."""
        cols = list(df.columns)
        _header_row(ws, start_row, cols)
        r = start_row
        for _, data_row in df.iterrows():
            r += 1
            for ci, col_name in enumerate(cols):
                val = data_row[col_name]
                if col_name == "Year":
                    fmt = "#,##0"
                elif isinstance(val, float):
                    fmt = "#,##0.0" if abs(val) >= 1 else "0.00"
                else:
                    fmt = None
                _cell(ws, r, ci + 1, val, fmt=fmt)
        return r + 1

    # ── Sheet 1: Consolidated View ──
    ws1 = wb.active
    ws1.title = "Consolidated View"
    ws1.sheet_properties.tabColor = TEAL
    for col in range(1, 10):
        ws1.column_dimensions[get_column_letter(col)].width = 20

    r = 1
    _cell(ws1, r, 1, f"{PEER_NAME} Normalization — PETRONAS vs {PEER_NAME}-Implied FTE",
          font=title_font, align=left_wrap)
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)

    r = 3
    r = _write_df(ws1, r, yearly_df)

    r += 1
    _cell(ws1, r, 1, "Breakdown by Archetype & Stage (avg 2026-2030)",
          font=Font(name="Calibri", size=12, bold=True, color=TEAL), align=left_wrap)
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

    r += 1
    r = _write_df(ws1, r, bucket_df)

    _cell(ws1, r, 1, "TOTAL", font=bold_font, fill=grey_fill)
    _cell(ws1, r, 2, "", font=bold_font, fill=grey_fill)
    numeric_cols = ["PETRONAS", f"{PEER_NAME} No Lab", f"{PEER_NAME} With Lab", "Lab Support"]
    for ci, col_name in enumerate(numeric_cols):
        if col_name in bucket_df.columns:
            _cell(ws1, r, ci + 3, bucket_df[col_name].sum(), font=bold_font, fmt="#,##0.0",
                  fill=result_fill)

    # ── Sheet 3: Norms Data ──
    ws3 = wb.create_sheet(f"{PEER_NAME} Norms Data")
    ws3.sheet_properties.tabColor = TEAL
    for col in range(1, 6):
        ws3.column_dimensions[get_column_letter(col)].width = 22

    r = 1
    _cell(ws3, r, 1, f"{PEER_NAME} FTE/MYR Norms (input data)", font=title_font, align=left_wrap)
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

    r = 3
    _header_row(ws3, r, ["Archetype", "Stage", "Without Lab (FTE/MYR)", "With Lab (FTE/MYR)", "Delta"])
    for key in PEER_WITHOUT_LAB:
        r += 1
        arch, stage = key
        no_lab = PEER_WITHOUT_LAB[key].direct_ratio
        with_lab = PEER_WITH_LAB[key].direct_ratio
        _cell(ws3, r, 1, arch, align=left_wrap)
        _cell(ws3, r, 2, stage)
        _cell(ws3, r, 3, no_lab, fmt="0.00")
        _cell(ws3, r, 4, with_lab, fmt="0.00")
        _cell(ws3, r, 5, with_lab - no_lab, fmt="+0.00;-0.00;0.00")

    # ── Sheet 4: Model Config ──
    ws4 = wb.create_sheet("Model Config")
    ws4.sheet_properties.tabColor = TEAL
    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 30

    r = 1
    _cell(ws4, r, 1, "Model Configuration", font=title_font, align=left_wrap)
    ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)

    params = [
        ("Total Budget (MYR M)", cfg.total_budget_m),
        ("Overhead %", f"{cfg.overhead_pct:.0%}"),
        ("Net Budget (MYR M)", cfg.total_budget_m * (1 - cfg.overhead_pct)),
        ("Start Year", cfg.start_year),
        ("End Year", cfg.end_year),
        ("Budget Mode", cfg.budget_mode),
        ("Stage Mix (TRL 1-4)", f"{cfg.stage_mix.get('TRL 1-4', 0):.0%}"),
        ("Stage Mix (TRL 5-7)", f"{cfg.stage_mix.get('TRL 5-7', 0):.0%}"),
        ("Conversion Rate (TRL 1-4 → 5-7)", f"{cfg.stage_conversion_rates.get('TRL 1-4', 0):.0%}"),
        ("Utilization Rate", f"{cfg.utilization_rate:.0%}"),
        ("Intake Spread (months)", cfg.intake_spread_months),
        ("Norm Metric", "FTE / MYR M"),
        ("Norm Source", f"{PEER_NAME} only"),
    ]

    r = 3
    _header_row(ws4, r, ["Parameter", "Value"])
    for label, val in params:
        r += 1
        _cell(ws4, r, 1, label, align=left_wrap)
        _cell(ws4, r, 2, val)

    r += 2
    _cell(ws4, r, 1, "PETRONAS Archetype Parameters", font=bold_font, align=left_wrap)
    ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1
    _header_row(ws4, r, ["Archetype", "Stage", "Duration (mo)", "Cost Min", "Cost Max", "FTE/Project"])
    for col in range(3, 7):
        ws4.column_dimensions[get_column_letter(col)].width = 16
    for arch in cfg.archetypes:
        for sname, sp in arch.stages.items():
            r += 1
            fte_total = sum(sp.fte_per_role.values())
            _cell(ws4, r, 1, arch.name, align=left_wrap)
            _cell(ws4, r, 2, sname)
            _cell(ws4, r, 3, sp.duration_months, fmt="#,##0")
            _cell(ws4, r, 4, sp.cost_min, fmt="#,##0.0")
            _cell(ws4, r, 5, sp.cost_max, fmt="#,##0.0")
            _cell(ws4, r, 6, fte_total, fmt="#,##0.0")

    out_path = Path(__file__).parent / f"{PEER_NAME}_Normalization.xlsx"
    wb.save(out_path)
    print(f"Excel saved to: {out_path}")
    return out_path


# ──────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────

def main():
    print(f"Running {PEER_NAME} FTE normalization (pipeline model)...")
    print("  Norm metric: FTE / MYR M")
    sm = cfg_no_lab.stage_mix
    print(f"  TRL intake split: {sm.get('TRL 1-4',0)*100:.0f}% TRL 1-4, {sm.get('TRL 5-7',0)*100:.0f}% TRL 5-7")
    print(f"  Norm source: {PEER_NAME} only")
    print()

    cfg, r_no, cfg_with, r_with = run_both()

    yearly_df, bucket_df = print_results(cfg, r_no, r_with)

    write_excel(yearly_df, bucket_df, r_no, r_with, cfg)


if __name__ == "__main__":
    main()

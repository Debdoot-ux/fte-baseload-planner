"""
Build a standalone, formula-driven FTE Baseload Excel Model.
Run this script to generate FTE_Baseload_Model_Live.xlsx.

Single-scenario model: user enters one value per parameter.
The yearly range comes from within-year variation (min/max monthly FTE).
Includes per-role contingency % for buffered headcount.
"""

import datetime
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

# ── Palette ──────────────────────────────────────────────────────────────
NAVY = "051C2C"
BLUE = "2251FF"
TEAL = "00A9F4"
GREY = "7F8C8D"
LIGHT = "F5F6F7"
WHITE = "FFFFFF"
DARK = "1A1A2E"
YELLOW_INPUT = "FFF9E6"
GREEN_OK = "E8F5E9"

# ── Styles ───────────────────────────────────────────────────────────────
navy_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
light_fill = PatternFill(start_color=LIGHT, end_color=LIGHT, fill_type="solid")
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
input_fill = PatternFill(start_color=YELLOW_INPUT, end_color=YELLOW_INPUT, fill_type="solid")
green_fill = PatternFill(start_color=GREEN_OK, end_color=GREEN_OK, fill_type="solid")

title_font = Font(name="Calibri", size=16, bold=True, color=NAVY)
section_font = Font(name="Calibri", size=12, bold=True, color=NAVY)
label_font = Font(name="Calibri", size=10, color=DARK)
input_font = Font(name="Calibri", size=10, color=BLUE, bold=True)
formula_font = Font(name="Calibri", size=10, color=GREY, italic=True)
hdr_font = Font(name="Calibri", size=10, bold=True, color=WHITE)
body_font = Font(name="Calibri", size=10, color=DARK)
note_font = Font(name="Calibri", size=9, color=GREY, italic=True)
bold_font = Font(name="Calibri", size=10, bold=True, color=DARK)

thin_border = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="thin", color="D0D5DD"),
)

align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

COMMENT_AUTHOR = "FTE Model"

def _cmt(text):
    return Comment(text, COMMENT_AUTHOR)

# ── Archetype definitions (defaults) ─────────────────────────────────────
ARCHETYPES = [
    {
        "name": "Chemistry",
        "share": 0.15,
        "stages": {
            "TRL 1-4": {"dur": 7, "cost": 6.5, "res": 3.5, "dev": 1.5},
            "TRL 5-7": {"dur": 12, "cost": 12.5, "res": 1.5, "dev": 3.5},
        },
    },
    {
        "name": "Process (Hardware)",
        "share": 0.70,
        "stages": {
            "TRL 1-4": {"dur": 9, "cost": 12.5, "res": 6.5, "dev": 1.5},
            "TRL 5-7": {"dur": 15, "cost": 15.0, "res": 1.5, "dev": 6.5},
        },
    },
    {
        "name": "Algorithm (Software)",
        "share": 0.15,
        "stages": {
            "TRL 1-4": {"dur": 6, "cost": 4.25, "res": 0.5, "dev": 0.5},
            "TRL 5-7": {"dur": 6, "cost": 4.25, "res": 0.5, "dev": 0.5},
        },
    },
]


def _style_header_row(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = navy_fill
        cell.font = hdr_font
        cell.alignment = align_center
        cell.border = thin_border


def _style_data_cell(ws, row, col, is_input=False, is_formula=False):
    cell = ws.cell(row=row, column=col)
    cell.border = thin_border
    cell.alignment = Alignment(vertical="center")
    if is_input:
        cell.fill = input_fill
        cell.font = input_font
    elif is_formula:
        cell.fill = green_fill
        cell.font = formula_font
    else:
        cell.font = body_font


# ═════════════════════════════════════════════════════════════════════════
# SHEET: INPUTS
# ═════════════════════════════════════════════════════════════════════════
def build_inputs_sheet(wb):
    ws = wb.active
    ws.title = "Inputs"
    ws.sheet_properties.tabColor = BLUE

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 5
    ws.column_dimensions["E"].width = 38
    ws.column_dimensions["F"].width = 16

    ws.merge_cells("B2:F2")
    ws["B2"] = "FTE Baseload Model \u2014 Inputs"
    ws["B2"].font = title_font

    ws["B3"] = "Yellow cells = you can change these.  Green cells = calculated by formulas.  All other sheets update automatically."
    ws["B3"].font = note_font

    ws["B2"].comment = _cmt(
        "This is the only sheet you need to edit.\n"
        "Yellow cells = your inputs. Green cells = calculated automatically.\n"
        "Change any yellow cell and every other sheet updates instantly."
    )

    row = 5

    # ── BUDGET & TIMELINE ──
    ws.cell(row=row, column=2, value="BUDGET & TIMELINE").font = section_font
    ws.cell(row=row, column=2).comment = _cmt(
        "BUDGET & TIMELINE\n"
        "Start here. Set your total R&D spend, how much goes to overhead, "
        "and the planning horizon."
    )
    row += 1

    labels_budget = [
        ("Total R&D budget (USD millions)", 400, "Budget", True, None,
         "Your total annual R&D budget in USD millions.\nThis is the gross figure before any deductions."),
        ("Overhead deduction (%)", 0.30, "Overhead", True, "0%",
         "What percentage of the budget goes to overhead (admin, facilities, management).\n"
         "Enter as a decimal: 0.30 = 30%.\nThe remainder is what actually funds projects."),
        ("Net project budget (M)", None, "NetBudget", False, "#,##0",
         "Auto-calculated: Budget \u00d7 (1 \u2013 Overhead).\n"
         "This is the money available to fund projects.\nDo not edit \u2014 this is a formula."),
        None,
        ("First year of new projects", 2026, "StartYear", False, "0",
         "The first calendar year when new projects begin.\nProjects from this year may still be running in later years."),
        ("Last year of new projects", 2029, "EndYear", False, "0",
         "The last calendar year when new projects are started.\n"
         "Projects already in progress continue beyond this year until they finish."),
        ("Intake window (months per year)", 6, "IntakeMonths", True, "0",
         "New projects are spread evenly across the first N months of each year.\n"
         "E.g. 6 means projects start in Jan\u2013Jun, not all in January.\nThis smooths out headcount peaks."),
        ("Utilization rate", 1.0, "Utilization", True, "0%",
         "What fraction of an FTE's time is spent on project work.\n"
         "1.0 = 100% (all time on projects).\n"
         "0.80 = 80% (the model inflates headcount by 1/0.80 = 25% to cover non-project time).\n"
         "Accounts for admin, training, leave, etc."),
    ]

    for item in labels_budget:
        if item is None:
            row += 1
            continue
        lbl, val, named, is_input, fmt, comment = item
        ws.cell(row=row, column=2, value=lbl).font = label_font
        cell = ws.cell(row=row, column=3)

        if named == "NetBudget":
            cell.value = "=Budget*(1-Overhead)"
            _style_data_cell(ws, row, 3, is_formula=True)
        else:
            cell.value = val
            _style_data_cell(ws, row, 3, is_input=True)

        if fmt:
            cell.number_format = fmt

        if named:
            ref = f"Inputs!$C${row}"
            wb.defined_names.add(DefinedName(named, attr_text=ref))

        cell.comment = _cmt(comment)
        row += 1

    row += 1

    # ── PIPELINE STAGES ──
    ws.cell(row=row, column=2, value="PIPELINE STAGES").font = section_font
    ws.cell(row=row, column=2).comment = _cmt(
        "PIPELINE STAGES\n"
        "Projects flow through two stages: TRL 1\u20134 (early research) and TRL 5\u20137 (late development).\n"
        "Set what share of new projects enters at each stage, "
        "and what share of early-stage completers advance to the late stage."
    )
    row += 1

    pipe_items = [
        ("TRL 1-4: % of new projects that start here", 0.20, "Alloc_Early", "0%",
         "Of all new projects started each year, what percentage enters the pipeline at the early stage (TRL 1\u20134)?\n"
         "Enter as a decimal: 0.20 = 20%.\nThis + the TRL 5\u20137 allocation below should add to 100%."),
        ("TRL 1-4: % of completers that advance to TRL 5-7", 0.50, "Conv_Early", "0%",
         "Of projects that finish TRL 1\u20134, what percentage advances to TRL 5\u20137?\n"
         "Enter as a decimal: 0.50 = 50%.\nThe rest are considered complete (or shelved) after the early stage."),
        ("TRL 5-7: % of new projects that start here directly", 0.80, "Alloc_Late", "0%",
         "Of all new projects started each year, what percentage enters directly at TRL 5\u20137?\n"
         "These skip the early stage entirely.\nEnter as a decimal: 0.80 = 80%."),
    ]
    for lbl, val, named, fmt, comment in pipe_items:
        ws.cell(row=row, column=2, value=lbl).font = label_font
        cell = ws.cell(row=row, column=3, value=val)
        cell.number_format = fmt
        _style_data_cell(ws, row, 3, is_input=True)
        wb.defined_names.add(DefinedName(named, attr_text=f"Inputs!$C${row}"))
        cell.comment = _cmt(comment)
        row += 1

    ws.cell(row=row, column=2, value="Total direct allocation (should = 100%)").font = label_font
    ws.cell(row=row, column=3, value="=Alloc_Early+Alloc_Late")
    ws.cell(row=row, column=3).number_format = "0%"
    _style_data_cell(ws, row, 3, is_formula=True)
    ws.cell(row=row, column=3).comment = _cmt(
        "Sanity check: TRL 1\u20134 allocation + TRL 5\u20137 allocation.\n"
        "Should equal 100%. If not, you're either missing projects or double-counting.\n"
        "Do not edit \u2014 this is a formula."
    )
    row += 2

    # ── PORTFOLIO MIX ──
    ws.cell(row=row, column=2, value="PORTFOLIO MIX").font = section_font
    ws.cell(row=row, column=2).comment = _cmt(
        "PORTFOLIO MIX\n"
        "Your R&D portfolio is a blend of different project types.\n"
        "Set the share for each type below. They must add to 100%.\n"
        "These shares determine the weighted average cost per project."
    )
    row += 1
    ws.cell(row=row, column=2, value="What share of your projects fall into each type? Must add to 100%.").font = note_font
    row += 1

    arch_share_names = ["Share_Chem", "Share_HW", "Share_SW"]
    share_comments = [
        "Percentage of your projects that are Chemistry-type.\nEnter as a decimal: 0.15 = 15%.",
        "Percentage of your projects that are Process (Hardware)-type.\nEnter as a decimal: 0.70 = 70%.",
        "Percentage of your projects that are Algorithm (Software)-type.\nEnter as a decimal: 0.15 = 15%.",
    ]
    for ai, arch in enumerate(ARCHETYPES):
        ws.cell(row=row, column=2, value=f"{arch['name']} (%)").font = label_font
        cell = ws.cell(row=row, column=3, value=arch["share"])
        cell.number_format = "0%"
        _style_data_cell(ws, row, 3, is_input=True)
        wb.defined_names.add(DefinedName(arch_share_names[ai], attr_text=f"Inputs!$C${row}"))
        cell.comment = _cmt(share_comments[ai])
        row += 1

    ws.cell(row=row, column=2, value="Total portfolio share").font = label_font
    ws.cell(row=row, column=3, value="=Share_Chem+Share_HW+Share_SW")
    ws.cell(row=row, column=3).number_format = "0%"
    _style_data_cell(ws, row, 3, is_formula=True)
    ws.cell(row=row, column=3).comment = _cmt("Sanity check: should equal 100%.\nDo not edit \u2014 this is a formula.")
    row += 2

    # ── ARCHETYPE PARAMETERS ──
    ws.cell(row=row, column=2, value="PROJECT TYPE PARAMETERS").font = section_font
    ws.cell(row=row, column=2).comment = _cmt(
        "PROJECT TYPE PARAMETERS\n"
        "For each project type and pipeline stage, set:\n"
        "\u2022 Duration \u2014 how many months a project takes in this stage\n"
        "\u2022 Cost \u2014 total cost (USD millions) for one project in this stage\n"
        "\u2022 Research FTE \u2014 number of researchers working on one project at any time\n"
        "\u2022 Developer FTE \u2014 number of developers working on one project at any time\n\n"
        "These are your best estimates. The model uses them to calculate total headcount."
    )
    row += 1
    ws.cell(row=row, column=2, value="For each project type and stage, set your best estimate for duration, cost, and team size.").font = note_font
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 1

    arch_short = ["Chem", "HW", "SW"]
    stage_short = ["E", "L"]
    stage_names = ["TRL 1-4", "TRL 5-7"]

    first_archetype = True
    for ai, arch in enumerate(ARCHETYPES):
        ws.cell(row=row, column=2, value=arch["name"]).font = Font(name="Calibri", size=11, bold=True, color=NAVY)
        row += 1

        for si, sname in enumerate(stage_names):
            prefix = f"{arch_short[ai]}_{stage_short[si]}"
            sp = arch["stages"][sname]

            ws.cell(row=row, column=2, value=f"  {sname}").font = Font(name="Calibri", size=10, bold=True, color=GREY)
            row += 1

            params = [
                ("Duration (months)", "Dur", sp["dur"], "0"),
                ("Cost per project (M)", "Cost", sp["cost"], "#,##0.0"),
                ("Research FTE per project", "Res", sp["res"], "0.0"),
                ("Developer FTE per project", "Dev", sp["dev"], "0.0"),
            ]

            _style_header_row(ws, row, 2, 3)
            ws.cell(row=row, column=2, value="Parameter")
            ws.cell(row=row, column=3, value="Value")
            row += 1

            for plbl, pshort, pval, pfmt in params:
                ws.cell(row=row, column=2, value=f"    {plbl}").font = label_font
                ws.cell(row=row, column=2).border = thin_border

                name = f"{prefix}_{pshort}"

                cell = ws.cell(row=row, column=3, value=pval)
                cell.number_format = pfmt
                _style_data_cell(ws, row, 3, is_input=True)
                wb.defined_names.add(DefinedName(name, attr_text=f"Inputs!$C${row}"))

                if first_archetype and si == 0:
                    hints = {
                        "Dur": f"How many months one {arch['name']} project spends in {sname}.",
                        "Cost": f"Total cost (USD M) of one {arch['name']} project during {sname}.",
                        "Res": f"Number of researchers assigned to one {arch['name']} {sname} project at any given time.",
                        "Dev": f"Number of developers assigned to one {arch['name']} {sname} project at any given time.",
                    }
                    cell.comment = _cmt(hints[pshort])

                row += 1

            row += 1

        first_archetype = False

    row += 1

    # ── DERIVED VALUES ──
    ws.cell(row=row, column=2, value="DERIVED VALUES (formulas \u2014 do not edit)").font = section_font
    ws.cell(row=row, column=2).comment = _cmt(
        "DERIVED VALUES\n"
        "Everything below is auto-calculated from your inputs above.\n"
        "Do not edit these cells \u2014 they contain formulas."
    )
    row += 1
    ws.cell(row=row, column=2, value="These are calculated from the inputs above. They show how the model converts budget into project count.").font = note_font
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    row += 1

    _style_header_row(ws, row, 2, 4)
    ws.cell(row=row, column=2, value="Metric")
    ws.cell(row=row, column=3, value="Value")
    ws.cell(row=row, column=4, value="How calculated")
    ws.column_dimensions["D"].width = 55
    row += 1

    wc_formula = (
        "=Share_Chem*(Alloc_Early*(Chem_E_Cost + Conv_Early*Chem_L_Cost) + Alloc_Late*Chem_L_Cost)"
        "+Share_HW*(Alloc_Early*(HW_E_Cost + Conv_Early*HW_L_Cost) + Alloc_Late*HW_L_Cost)"
        "+Share_SW*(Alloc_Early*(SW_E_Cost + Conv_Early*SW_L_Cost) + Alloc_Late*SW_L_Cost)"
    )

    ws.cell(row=row, column=2, value="Weighted cost per project (M)").font = label_font
    ws.cell(row=row, column=2).border = thin_border
    c_wc = ws.cell(row=row, column=3, value=wc_formula)
    c_wc.number_format = "#,##0.0"
    _style_data_cell(ws, row, 3, is_formula=True)
    wb.defined_names.add(DefinedName("WtdCost", attr_text=f"Inputs!$C${row}"))
    c_wc.comment = _cmt(
        "The portfolio-weighted average cost of one project, accounting for:\n"
        "\u2022 Which stages projects enter (stage mix)\n"
        "\u2022 What share advance from early to late (conversion)\n"
        "\u2022 How expensive each stage is per archetype (cost parameters)\n"
        "\u2022 What share each archetype makes up (portfolio mix)"
    )

    ws.cell(row=row, column=4, value="Portfolio-weighted average cost, accounting for stage mix and conversion").font = note_font
    row += 1

    ws.cell(row=row, column=2, value="Projects per year").font = label_font
    ws.cell(row=row, column=2).border = thin_border
    c_pp = ws.cell(row=row, column=3, value="=IF(WtdCost>0, NetBudget/WtdCost, 0)")
    c_pp.number_format = "#,##0.0"
    _style_data_cell(ws, row, 3, is_formula=True)
    wb.defined_names.add(DefinedName("ProjPerYr", attr_text=f"Inputs!$C${row}"))
    c_pp.comment = _cmt(
        "How many new projects the net budget can fund each year.\n"
        "= Net project budget \u00f7 Weighted cost per project."
    )

    ws.cell(row=row, column=4, value="Net budget \u00f7 weighted cost per project").font = note_font
    row += 1

    # ── CONTINGENCY ──
    row += 1
    ws.cell(row=row, column=2, value="CONTINGENCY").font = section_font
    ws.cell(row=row, column=2).comment = _cmt(
        "CONTINGENCY\n"
        "An optional buffer on top of the model's calculated FTE.\n"
        "Use this to account for uncertainty, attrition, leave, or estimation error.\n"
        "Set to 0% if you don't want any buffer \u2014 the adjusted columns will equal the base columns."
    )
    row += 1
    ws.cell(row=row, column=2,
            value="Buffer on top of calculated FTE. Set separately for Research and Developer roles. "
                  "Adjusted FTE = Base FTE \u00d7 (1 + Contingency %).").font = note_font
    row += 1

    ws.cell(row=row, column=2, value="Research contingency (%)").font = label_font
    ws.cell(row=row, column=2).border = thin_border
    cell_cr = ws.cell(row=row, column=3, value=0.0)
    cell_cr.number_format = "0%"
    _style_data_cell(ws, row, 3, is_input=True)
    wb.defined_names.add(DefinedName("Cont_Res", attr_text=f"Inputs!$C${row}"))
    cell_cr.comment = _cmt(
        "Extra buffer on Research headcount.\n"
        "E.g. 0.10 = 10% means 10% more researchers than the base model calculates.\n"
        "Adjusted Research FTE = Base Research FTE \u00d7 (1 + this %).\nSet to 0 for no buffer."
    )
    row += 1

    ws.cell(row=row, column=2, value="Developer contingency (%)").font = label_font
    ws.cell(row=row, column=2).border = thin_border
    cell_cd = ws.cell(row=row, column=3, value=0.0)
    cell_cd.number_format = "0%"
    _style_data_cell(ws, row, 3, is_input=True)
    wb.defined_names.add(DefinedName("Cont_Dev", attr_text=f"Inputs!$C${row}"))
    cell_cd.comment = _cmt(
        "Extra buffer on Developer headcount.\n"
        "E.g. 0.15 = 15% means 15% more developers than the base model calculates.\n"
        "Adjusted Developer FTE = Base Developer FTE \u00d7 (1 + this %).\nSet to 0 for no buffer."
    )

    return ws


# ═════════════════════════════════════════════════════════════════════════
# SHEET: GLOSSARY (How This Model Works)
# ═════════════════════════════════════════════════════════════════════════
def build_glossary_sheet(wb):
    ws = wb.create_sheet("Glossary")
    ws.sheet_properties.tabColor = NAVY

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 100

    ws["B2"] = "FTE Baseload Model \u2014 How It Works"
    ws["B2"].font = title_font
    ws["B2"].comment = _cmt(
        "Read this sheet first if you're new to the model.\n"
        "It explains what the model does, how it calculates headcount, "
        "and what all the key terms mean."
    )

    content = [
        ("WHAT THIS MODEL DOES", section_font),
        ("It answers: 'Given our R&D budget and project portfolio, how many researchers and developers do we need?'", label_font),
        ("You provide your best estimates for cost, duration, and team size. The model calculates a single headcount figure.", label_font),
        ("The yearly range (min\u2013max) reflects natural within-year variation as projects start, overlap, and complete.", label_font),
        ("", None),
        ("HOW IT CALCULATES HEADCOUNT \u2014 5 STEPS", section_font),
        ("", None),
        ("Step 1: Start with the money", Font(name="Calibri", size=10, bold=True, color=NAVY)),
        ("Total R&D budget minus overhead = net project budget.", label_font),
        ("", None),
        ("Step 2: Figure out how many projects you can afford", Font(name="Calibri", size=10, bold=True, color=NAVY)),
        ("Net budget \u00f7 weighted average cost per project = projects per year.", label_font),
        ("The cost is weighted by your portfolio mix and which stages projects go through.", label_font),
        ("", None),
        ("Step 3: Distribute projects across types and stages", Font(name="Calibri", size=10, bold=True, color=NAVY)),
        ("Projects are split across archetypes (Chemistry, Hardware, Software) by portfolio share.", label_font),
        ("Within each type, some start at TRL 1-4 (early) and some start directly at TRL 5-7 (late).", label_font),
        ("When early-stage projects finish, a percentage advance to TRL 5-7 as additional projects.", label_font),
        ("", None),
        ("Step 4: Simulate the pipeline month by month", Font(name="Calibri", size=10, bold=True, color=NAVY)),
        ("Each year's new projects are spread across the first N months (the intake window).", label_font),
        ("A project stays 'active' from its start month through start + duration months.", label_font),
        ("The Engine sheet tracks how many projects are running in each stage, every month.", label_font),
        ("", None),
        ("Step 5: Convert active projects into headcount", Font(name="Calibri", size=10, bold=True, color=NAVY)),
        ("Active projects \u00d7 staff per project = FTE needed that month.", label_font),
        ("If utilization < 100%, the model inflates by 1 \u00f7 utilization to get gross headcount.", label_font),
        ("", None),
        ("UNDERSTANDING STEADY STATE AND THE YEARLY RANGE", section_font),
        ("", None),
        ("Why does headcount grow at first?", Font(name="Calibri", size=10, bold=True, color=NAVY)),
        ("In Year 1, projects start but none have finished yet \u2014 the pipeline only fills up.", label_font),
        ("In Year 2, new projects start while Year 1 projects are still running. Headcount keeps climbing", label_font),
        ("until the rate of new starts roughly equals the rate of completions.", label_font),
        ("Once that happens, headcount stabilizes \u2014 this is the STEADY STATE.", label_font),
        ("The steady-state headcount is the long-run staffing level your hiring plan should target.", label_font),
        ("", None),
        ("Why is there a range within each year?", Font(name="Calibri", size=10, bold=True, color=NAVY)),
        ("Because new projects start during an intake window (not all at once), FTE demand varies month to month.", label_font),
        ("The Summary sheet shows:", label_font),
        ("  \u2022 Avg monthly FTE \u2014 the steady-state staffing level for that year", label_font),
        ("  \u2022 Min monthly FTE \u2014 the quietest month (e.g. just before a new cohort starts)", label_font),
        ("  \u2022 Max monthly FTE \u2014 the busiest month (e.g. when old and new cohorts overlap most)", label_font),
        ("Early years have a wide range (pipeline is still filling). Later years converge (pipeline has stabilized).", label_font),
        ("", None),
        ("KEY TERMS", section_font),
        ("FTE = Full-Time Equivalent. 1 FTE = one person working full time.", label_font),
        ("TRL = Technology Readiness Level. A 1-to-9 scale of technology maturity.", label_font),
        ("Pipeline = The sequence of stages a project goes through.", label_font),
        ("Archetype = A type of R&D project (e.g. Chemistry, Hardware, Software).", label_font),
        ("Conversion rate = % of projects finishing one stage that advance to the next.", label_font),
        ("", None),
        ("THINGS THIS EXCEL ASSUMES THAT CANNOT BE CHANGED", section_font),
        ("(Changing these would require rebuilding the sheet structure)", note_font),
        ("", None),
        ("\u2022 2 pipeline stages (TRL 1-4 and TRL 5-7)", label_font),
        ("\u2022 3 project types (Chemistry, Process Hardware, Algorithm Software)", label_font),
        ("\u2022 2 FTE roles (Research and Developer)", label_font),
        ("\u2022 Monthly granularity \u2014 projects tracked month by month", label_font),
        ("\u2022 No ramp-up \u2014 projects start at full staffing immediately", label_font),
        ("\u2022 No mid-stage cancellation \u2014 projects run to completion", label_font),
        ("\u2022 Same budget every year across the planning horizon", label_font),
        ("\u2022 Projects spread evenly across the intake window", label_font),
        ("\u2022 Linear pipeline \u2014 no branching or looping between stages", label_font),
        ("\u2022 No economies of scale \u2014 cost per project is constant regardless of volume", label_font),
        ("", None),
        ("SHEET GUIDE", section_font),
        ("Inputs \u2014 the only sheet you need to edit. All yellow cells are changeable.", label_font),
        ("Engine \u2014 monthly calculations for all archetypes and stages. All formulas.", label_font),
        ("Output \u2014 annual averages and within-year range from the engine. All formulas.", label_font),
        ("", None),
        ("CONTINGENCY", section_font),
        ("Contingency % is a buffer added on top of calculated FTE to account for uncertainty,", label_font),
        ("attrition, leave, or estimation error. It is set separately for Research and Developer roles.", label_font),
        ("Adjusted FTE = Base FTE \u00d7 (1 + Contingency %). Set on the Inputs sheet; defaults to 0%.", label_font),
    ]

    ws.cell(row=70, column=2).comment = _cmt(
        "SHEET GUIDE\nQuick reference for what each sheet does:\n"
        "\u2022 Inputs \u2014 the only sheet you edit\n"
        "\u2022 Engine \u2014 monthly formulas (don't touch)\n"
        "\u2022 Output \u2014 annual summary (don't touch)\n"
        "\u2022 Glossary \u2014 this sheet (explanation only)"
    )

    row = 4
    for text, font in content:
        if text or font:
            cell = ws.cell(row=row, column=2, value=text)
            if font:
                cell.font = font
        row += 1

    return ws


# ═════════════════════════════════════════════════════════════════════════
# SHEET: ENGINE
# Layout: Row 2 = group headers, Row 3 = sub-headers, Rows 4-63 = data
# Columns: B=Date, C=Year, D=Month, E+ = archetype blocks (9 cols each)
# ═════════════════════════════════════════════════════════════════════════
def build_engine_sheet(wb):
    ws = wb.create_sheet("Engine")
    ws.sheet_properties.tabColor = BLUE

    arch_short = ["Chem", "HW", "SW"]
    share_names = ["Share_Chem", "Share_HW", "Share_SW"]

    n_months = 60
    data_start_row = 4
    ds = 3  # used in offset arithmetic within formulas

    cols_per_arch = 9
    arch_start_col = 5  # column E

    # Row 2: group headers
    for ai, arch in enumerate(ARCHETYPES):
        start_c = arch_start_col + ai * cols_per_arch
        ws.merge_cells(start_row=2, start_column=start_c, end_row=2, end_column=start_c + cols_per_arch - 1)
        ws.cell(row=2, column=start_c, value=arch["name"]).font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        ws.cell(row=2, column=start_c).fill = navy_fill
        ws.cell(row=2, column=start_c).alignment = align_center
        for c in range(start_c, start_c + cols_per_arch):
            ws.cell(row=2, column=c).fill = navy_fill

    totals_col = arch_start_col + 3 * cols_per_arch  # column AF (col 32)
    ws.merge_cells(start_row=2, start_column=totals_col, end_row=2, end_column=totals_col + 2)
    ws.cell(row=2, column=totals_col, value="TOTALS").font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    ws.cell(row=2, column=totals_col).fill = navy_fill
    ws.cell(row=2, column=totals_col).alignment = align_center
    for c in range(totals_col, totals_col + 3):
        ws.cell(row=2, column=c).fill = navy_fill

    adj_col = totals_col + 3  # column AI (col 35)
    ws.merge_cells(start_row=2, start_column=adj_col, end_row=2, end_column=adj_col + 2)
    ws.cell(row=2, column=adj_col, value="ADJUSTED TOTALS").font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    ws.cell(row=2, column=adj_col).fill = navy_fill
    ws.cell(row=2, column=adj_col).alignment = align_center
    for c in range(adj_col, adj_col + 3):
        ws.cell(row=2, column=c).fill = navy_fill

    # Row 3: sub-headers
    _style_header_row(ws, 3, 2, adj_col + 2)
    ws.cell(row=3, column=2, value="Date")
    ws.cell(row=3, column=3, value="Year")
    ws.cell(row=3, column=4, value="Month")

    sub_headers = [
        "Early Starts/mo", "Early Active",
        "Late Conv Starts", "Late Direct Starts", "Late Total Starts", "Late Active",
        "Research FTE", "Developer FTE", "Total FTE",
    ]
    for ai in range(3):
        sc = arch_start_col + ai * cols_per_arch
        for si, sh in enumerate(sub_headers):
            ws.cell(row=3, column=sc + si, value=sh)

    ws.cell(row=3, column=totals_col, value="Total Research")
    ws.cell(row=3, column=totals_col + 1, value="Total Developer")
    ws.cell(row=3, column=totals_col + 2, value="Total FTE")
    ws.cell(row=3, column=adj_col, value="Adj Research FTE")
    ws.cell(row=3, column=adj_col + 1, value="Adj Developer FTE")
    ws.cell(row=3, column=adj_col + 2, value="Adj Total FTE")

    # Column widths
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 5
    for c in range(arch_start_col, adj_col + 3):
        ws.column_dimensions[get_column_letter(c)].width = 14
    for c in range(adj_col, adj_col + 3):
        ws.column_dimensions[get_column_letter(c)].width = 16

    # Comments on headers
    ws.cell(row=2, column=2).comment = _cmt(
        "ENGINE SHEET \u2014 ALL FORMULAS, DO NOT EDIT\n\n"
        "Each row = one month. 60 rows = 5 years of monthly tracking.\n"
        "Each archetype (Chemistry, Hardware, Software) has its own block of columns.\n\n"
        "Column groups per archetype:\n"
        "\u2022 Early Starts/mo \u2014 how many projects start in TRL 1\u20134 this month\n"
        "\u2022 Early Active \u2014 total TRL 1\u20134 projects running this month (sliding window)\n"
        "\u2022 Late Conv Starts \u2014 projects entering TRL 5\u20137 as graduates from TRL 1\u20134\n"
        "\u2022 Late Direct Starts \u2014 projects entering TRL 5\u20137 directly\n"
        "\u2022 Late Total Starts \u2014 sum of conversion + direct starts\n"
        "\u2022 Late Active \u2014 total TRL 5\u20137 projects running this month\n"
        "\u2022 Research FTE / Developer FTE / Total FTE \u2014 headcount this month"
    )
    ws.cell(row=2, column=totals_col).comment = _cmt(
        "TOTALS \u2014 base headcount (before contingency)\n"
        "Sum of Research/Developer/Total FTE across all three archetypes."
    )
    ws.cell(row=2, column=adj_col).comment = _cmt(
        "ADJUSTED TOTALS \u2014 headcount with contingency buffer\n"
        "= Base FTE \u00d7 (1 + Contingency %).\n"
        "At 0% contingency, these equal the base totals exactly."
    )

    ws.cell(row=3, column=2).comment = _cmt("Date of the month (1st of each month).")
    ws.cell(row=3, column=3).comment = _cmt("Calendar year extracted from the date.")
    ws.cell(row=3, column=4).comment = _cmt("Calendar month (1\u201312) extracted from the date.")

    # Comments on first archetype FTE headers
    first_res_col = arch_start_col + 6
    first_dev_col = arch_start_col + 7
    first_tot_col = arch_start_col + 8
    ws.cell(row=3, column=first_res_col).comment = _cmt(
        "Research FTE for this archetype this month.\n"
        "= (Early Active \u00d7 Research per early project + Late Active \u00d7 Research per late project) \u00f7 Utilization"
    )
    ws.cell(row=3, column=first_dev_col).comment = _cmt(
        "Developer FTE for this archetype this month.\n"
        "Same formula pattern as Research FTE but with Developer parameters."
    )
    ws.cell(row=3, column=first_tot_col).comment = _cmt("Total FTE = Research + Developer.")

    ws.cell(row=3, column=totals_col).comment = _cmt("Sum of Research FTE across all three archetypes.")
    ws.cell(row=3, column=totals_col + 1).comment = _cmt("Sum of Developer FTE across all three archetypes.")
    ws.cell(row=3, column=totals_col + 2).comment = _cmt("Grand total FTE (base) = Total Research + Total Developer.")
    ws.cell(row=3, column=adj_col).comment = _cmt("= Total Research \u00d7 (1 + Research Contingency %).")
    ws.cell(row=3, column=adj_col + 1).comment = _cmt("= Total Developer \u00d7 (1 + Developer Contingency %).")
    ws.cell(row=3, column=adj_col + 2).comment = _cmt("= Adj Research + Adj Developer. Final adjusted headcount.")

    ai_letter = get_column_letter(adj_col)
    aj_letter = get_column_letter(adj_col + 1)
    ak_letter = get_column_letter(adj_col + 2)

    # ── Data rows ──
    for r_offset in range(n_months):
        r = data_start_row + r_offset

        ws.cell(row=r, column=2, value=f"=DATE(StartYear + INT({r_offset}/12), MOD({r_offset},12)+1, 1)")
        ws.cell(row=r, column=2).number_format = "YYYY-MM"
        ws.cell(row=r, column=3, value=f"=YEAR(B{r})")
        ws.cell(row=r, column=4, value=f"=MONTH(B{r})")

        for ai in range(3):
            sc = arch_start_col + ai * cols_per_arch
            ashort = arch_short[ai]
            share = share_names[ai]

            early_starts_col = get_column_letter(sc)
            early_active_col = get_column_letter(sc + 1)
            late_conv_col = get_column_letter(sc + 2)
            late_direct_col = get_column_letter(sc + 3)
            late_total_col = get_column_letter(sc + 4)
            late_active_col = get_column_letter(sc + 5)
            res_col = get_column_letter(sc + 6)
            dev_col = get_column_letter(sc + 7)
            tot_col = get_column_letter(sc + 8)

            dur_early = f"{ashort}_E_Dur"
            dur_late = f"{ashort}_L_Dur"
            res_early = f"{ashort}_E_Res"
            dev_early = f"{ashort}_E_Dev"
            res_late = f"{ashort}_L_Res"
            dev_late = f"{ashort}_L_Dev"

            # Early Direct Starts per month
            ws.cell(row=r, column=sc).value = (
                f"=IF(AND(C{r}>=StartYear, C{r}<=EndYear, D{r}<=IntakeMonths),"
                f" ProjPerYr*{share}*Alloc_Early/IntakeMonths, 0)"
            )
            ws.cell(row=r, column=sc).number_format = "0.00"

            # Early Active Stock (sliding window)
            if r_offset == 0:
                ws.cell(row=r, column=sc + 1).value = f"={early_starts_col}{r}"
            else:
                ws.cell(row=r, column=sc + 1).value = (
                    f"=IF({r - 1}-{ds} < {dur_early},"
                    f" SUM({early_starts_col}${data_start_row}:{early_starts_col}{r}),"
                    f" SUM(OFFSET({early_starts_col}{r},-{dur_early}+1,0,{dur_early},1)))"
                )
            ws.cell(row=r, column=sc + 1).number_format = "0.00"

            # Late Conversion Starts
            ws.cell(row=r, column=sc + 2).value = (
                f"=IF({r - 1}-{ds} >= {dur_early},"
                f" OFFSET({early_starts_col}{r}, -{dur_early}, 0) * Conv_Early, 0)"
            )
            ws.cell(row=r, column=sc + 2).number_format = "0.00"

            # Late Direct Starts
            ws.cell(row=r, column=sc + 3).value = (
                f"=IF(AND(C{r}>=StartYear, C{r}<=EndYear, D{r}<=IntakeMonths),"
                f" ProjPerYr*{share}*Alloc_Late/IntakeMonths, 0)"
            )
            ws.cell(row=r, column=sc + 3).number_format = "0.00"

            # Late Total Starts
            ws.cell(row=r, column=sc + 4).value = f"={late_conv_col}{r}+{late_direct_col}{r}"
            ws.cell(row=r, column=sc + 4).number_format = "0.00"

            # Late Active Stock (sliding window)
            if r_offset == 0:
                ws.cell(row=r, column=sc + 5).value = f"={late_total_col}{r}"
            else:
                ws.cell(row=r, column=sc + 5).value = (
                    f"=IF({r - 1}-{ds} < {dur_late},"
                    f" SUM({late_total_col}${data_start_row}:{late_total_col}{r}),"
                    f" SUM(OFFSET({late_total_col}{r},-{dur_late}+1,0,{dur_late},1)))"
                )
            ws.cell(row=r, column=sc + 5).number_format = "0.00"

            # FTE
            util = "Utilization"
            ws.cell(row=r, column=sc + 6).value = (
                f"=({early_active_col}{r}*{res_early} + {late_active_col}{r}*{res_late}) / {util}"
            )
            ws.cell(row=r, column=sc + 6).number_format = "0.0"

            ws.cell(row=r, column=sc + 7).value = (
                f"=({early_active_col}{r}*{dev_early} + {late_active_col}{r}*{dev_late}) / {util}"
            )
            ws.cell(row=r, column=sc + 7).number_format = "0.0"

            ws.cell(row=r, column=sc + 8).value = f"={res_col}{r}+{dev_col}{r}"
            ws.cell(row=r, column=sc + 8).number_format = "0.0"

        # Grand Totals (base)
        res_cols = [get_column_letter(arch_start_col + ai * cols_per_arch + 6) for ai in range(3)]
        dev_cols = [get_column_letter(arch_start_col + ai * cols_per_arch + 7) for ai in range(3)]

        ws.cell(row=r, column=totals_col).value = f"={'+'.join(f'{c}{r}' for c in res_cols)}"
        ws.cell(row=r, column=totals_col).number_format = "0.0"
        ws.cell(row=r, column=totals_col + 1).value = f"={'+'.join(f'{c}{r}' for c in dev_cols)}"
        ws.cell(row=r, column=totals_col + 1).number_format = "0.0"
        tc = get_column_letter(totals_col)
        tc1 = get_column_letter(totals_col + 1)
        ws.cell(row=r, column=totals_col + 2).value = f"={tc}{r}+{tc1}{r}"
        ws.cell(row=r, column=totals_col + 2).number_format = "0.0"

        # Adjusted Totals (with contingency)
        ws.cell(row=r, column=adj_col).value = f"=AF{r}*(1+Cont_Res)"
        ws.cell(row=r, column=adj_col).number_format = "0.0"
        ws.cell(row=r, column=adj_col).font = body_font
        ws.cell(row=r, column=adj_col).border = thin_border

        ws.cell(row=r, column=adj_col + 1).value = f"=AG{r}*(1+Cont_Dev)"
        ws.cell(row=r, column=adj_col + 1).number_format = "0.0"
        ws.cell(row=r, column=adj_col + 1).font = body_font
        ws.cell(row=r, column=adj_col + 1).border = thin_border

        ws.cell(row=r, column=adj_col + 2).value = f"={ai_letter}{r}+{aj_letter}{r}"
        ws.cell(row=r, column=adj_col + 2).number_format = "0.0"
        ws.cell(row=r, column=adj_col + 2).font = body_font
        ws.cell(row=r, column=adj_col + 2).border = thin_border

    # Hide only the Early Starts column per archetype (matching desktop file)
    for ai in range(3):
        sc = arch_start_col + ai * cols_per_arch
        col_letter = get_column_letter(sc)
        ws.column_dimensions[col_letter].hidden = True

    return ws, totals_col, adj_col


# ═════════════════════════════════════════════════════════════════════════
# SHEET: OUTPUT (Summary)
# Column order: B=Year, C=Min, D=Max, E=Avg, F=Avg Research, G=Avg Dev,
#               H=Min(adj), I=Max(adj), J=Avg(adj)
# ═════════════════════════════════════════════════════════════════════════
def build_output_sheet(wb, totals_col, adj_col):
    ws = wb.create_sheet("Output")
    ws.sheet_properties.tabColor = NAVY

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 10
    for c_letter in ["C", "D", "E", "F", "G", "H", "I", "J"]:
        ws.column_dimensions[c_letter].width = 18

    ws["B2"] = "Annual FTE Summary"
    ws["B2"].font = title_font
    ws["B3"] = "Avg = average across all months. Min/Max = lowest and highest single-month FTE that year."
    ws["B3"].font = note_font

    ws["B2"].comment = _cmt(
        "ANNUAL FTE SUMMARY\n\n"
        "This sheet summarises the Engine's 60 months of data into yearly figures.\n"
        "All values are formulas \u2014 they update automatically when you change inputs.\n\n"
        "Columns B\u2013G: base headcount (before contingency).\n"
        "Columns H\u2013J: adjusted headcount (after contingency buffer).\n\n"
        "At 0% contingency, the adjusted columns equal the base columns."
    )

    # Headers
    headers = [
        ("Year", "Calendar year (derived from StartYear on the Inputs sheet)."),
        ("Min monthly FTE", "The quietest month that year (base).\nUseful for planning minimum staffing levels."),
        ("Max monthly FTE", "The busiest month that year (base).\nUseful for planning peak staffing capacity."),
        ("Avg monthly FTE", "Average monthly FTE that year (base).\nThis is the steady-state staffing level you should target."),
        ("Avg Research FTE", "Average monthly Research FTE that year (base)."),
        ("Avg Developer FTE", "Average monthly Developer FTE that year (base)."),
        ("Min FTE (adj)", "Minimum monthly FTE that year, AFTER applying contingency."),
        ("Max FTE (adj)", "Maximum monthly FTE that year, AFTER applying contingency."),
        ("Avg FTE (adj)", "Average monthly FTE that year, AFTER applying contingency.\nThis is the adjusted planning target."),
    ]

    _style_header_row(ws, 5, 2, 10)
    for ci, (h, comment) in enumerate(headers):
        cell = ws.cell(row=5, column=2 + ci, value=h)
        cell.comment = _cmt(comment)

    tc_total = get_column_letter(totals_col + 2)  # AH = Total FTE
    tc_res = get_column_letter(totals_col)          # AF = Total Research
    tc_dev = get_column_letter(totals_col + 1)      # AG = Total Developer
    tc_adj = get_column_letter(adj_col + 2)          # AK = Adj Total FTE

    data_start = 4
    data_end = 63

    n_years = 2029 - 2026 + 1
    for yi in range(n_years):
        r = 6 + yi
        ws.cell(row=r, column=2, value=f"=StartYear+{yi}")
        ws.cell(row=r, column=2).number_format = "0"
        ws.cell(row=r, column=2).font = bold_font
        ws.cell(row=r, column=2).border = thin_border

        year_ref = f"B{r}"

        # C: Min monthly FTE (base)
        ws.cell(row=r, column=3).value = (
            f"=IFERROR(_xlfn.MINIFS(Engine!{tc_total}${data_start}:{tc_total}${data_end},"
            f"Engine!$C${data_start}:$C${data_end},{year_ref},"
            f"Engine!{tc_total}${data_start}:{tc_total}${data_end},\">0\"), 0)"
        )
        ws.cell(row=r, column=3).number_format = "0.0"
        _style_data_cell(ws, r, 3, is_formula=True)

        # D: Max monthly FTE (base)
        ws.cell(row=r, column=4).value = (
            f"=IFERROR(_xlfn.MAXIFS(Engine!{tc_total}${data_start}:{tc_total}${data_end},"
            f"Engine!$C${data_start}:$C${data_end},{year_ref},"
            f"Engine!{tc_total}${data_start}:{tc_total}${data_end},\">0\"), 0)"
        )
        ws.cell(row=r, column=4).number_format = "0.0"
        _style_data_cell(ws, r, 4, is_formula=True)

        # E: Avg monthly FTE (base)
        ws.cell(row=r, column=5).value = (
            f"=IFERROR(AVERAGEIFS(Engine!{tc_total}${data_start}:{tc_total}${data_end},"
            f"Engine!$C${data_start}:$C${data_end},{year_ref},"
            f"Engine!{tc_total}${data_start}:{tc_total}${data_end},\">0\"), 0)"
        )
        ws.cell(row=r, column=5).number_format = "0.0"
        _style_data_cell(ws, r, 5, is_formula=True)

        # F: Avg Research FTE (base)
        ws.cell(row=r, column=6).value = (
            f"=IFERROR(AVERAGEIFS(Engine!{tc_res}${data_start}:{tc_res}${data_end},"
            f"Engine!$C${data_start}:$C${data_end},{year_ref},"
            f"Engine!{tc_total}${data_start}:{tc_total}${data_end},\">0\"), 0)"
        )
        ws.cell(row=r, column=6).number_format = "0.0"
        _style_data_cell(ws, r, 6, is_formula=True)

        # G: Avg Developer FTE (base)
        ws.cell(row=r, column=7).value = (
            f"=IFERROR(AVERAGEIFS(Engine!{tc_dev}${data_start}:{tc_dev}${data_end},"
            f"Engine!$C${data_start}:$C${data_end},{year_ref},"
            f"Engine!{tc_total}${data_start}:{tc_total}${data_end},\">0\"), 0)"
        )
        ws.cell(row=r, column=7).number_format = "0.0"
        _style_data_cell(ws, r, 7, is_formula=True)

        # H: Min FTE (adjusted)
        ws.cell(row=r, column=8).value = (
            f"=IFERROR(_xlfn.MINIFS(Engine!{tc_adj}${data_start}:{tc_adj}${data_end},"
            f"Engine!$C${data_start}:$C${data_end},{year_ref},"
            f"Engine!{tc_adj}${data_start}:{tc_adj}${data_end},\">0\"), 0)"
        )
        ws.cell(row=r, column=8).number_format = "0.0"
        _style_data_cell(ws, r, 8, is_formula=True)

        # I: Max FTE (adjusted)
        ws.cell(row=r, column=9).value = (
            f"=IFERROR(_xlfn.MAXIFS(Engine!{tc_adj}${data_start}:{tc_adj}${data_end},"
            f"Engine!$C${data_start}:$C${data_end},{year_ref},"
            f"Engine!{tc_adj}${data_start}:{tc_adj}${data_end},\">0\"), 0)"
        )
        ws.cell(row=r, column=9).number_format = "0.0"
        _style_data_cell(ws, r, 9, is_formula=True)

        # J: Avg FTE (adjusted)
        ws.cell(row=r, column=10).value = (
            f"=IFERROR(AVERAGEIFS(Engine!{tc_adj}${data_start}:{tc_adj}${data_end},"
            f"Engine!$C${data_start}:$C${data_end},{year_ref},"
            f"Engine!{tc_adj}${data_start}:{tc_adj}${data_end},\">0\"), 0)"
        )
        ws.cell(row=r, column=10).number_format = "0.0"
        _style_data_cell(ws, r, 10, is_formula=True)

    return ws


# ═════════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════════
def main():
    wb = Workbook()

    print("Building Inputs sheet...")
    build_inputs_sheet(wb)

    print("Building Glossary sheet...")
    build_glossary_sheet(wb)

    print("Building Engine sheet...")
    _, totals_col, adj_col = build_engine_sheet(wb)

    print("Building Output sheet...")
    build_output_sheet(wb, totals_col, adj_col)

    desired_order = [
        "Output",
        "Inputs",
        "Engine",
        "Glossary",
    ]
    sheet_indices = {ws.title: i for i, ws in enumerate(wb.worksheets)}
    new_order = [sheet_indices[name] for name in desired_order]
    wb._sheets = [wb.worksheets[i] for i in new_order]

    out_path = r"c:\Users\Debdoot Ray\genAI training\fte_model\FTE_Baseload_Model_Live.xlsx"
    wb.save(out_path)
    print(f"\nSaved to: {out_path}")
    print("Open in Excel and try changing a yellow cell on the Inputs sheet!")


if __name__ == "__main__":
    main()

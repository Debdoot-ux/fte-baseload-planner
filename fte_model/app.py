"""
FTE Baseload Planning Tool — Streamlit UI
Multi-scenario flow: Configure one or more scenarios → View results → Compare.
"""

import io
import sys
from pathlib import Path

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))

from config import Archetype, ModelConfig, StageParams
from defaults import default_baseline
from model import run_model, _weighted_cost_per_project
from scenario_engine import run_all, comparison_summary, generate_comparison_excel

# ---------------------------------------------------------------------------
# Palette  (McKinsey Design System)
# ---------------------------------------------------------------------------
MCK_NAVY = "#051C2C"
MCK_BLUE = "#00A3E0"
MCK_TEAL = "#0067A0"
MCK_GREEN = "#2E7D32"
MCK_GREY = "#63666A"
MCK_LIGHT = "#F5F5F5"
MCK_WHITE = "#FFFFFF"
MCK_DARK = "#333333"
MCK_BORDER = "#E0E0E0"

ARCH_COLORS = [
    "#051C2C", "#00A3E0", "#0067A0", "#ED6C02", "#7B1FA2",
    "#2E7D32", "#F9A825", "#C62828", "#0088c2", "#4CAF50",
]

SCENARIO_COLORS = [
    "#051C2C", "#00A3E0", "#0067A0", "#ED6C02", "#7B1FA2",
    "#2E7D32", "#F9A825", "#C62828", "#0088c2", "#4CAF50",
]


def _rgba(hex_color: str, opacity: float) -> str:
    h = hex_color.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return f"rgba({r},{g},{b},{opacity})"


# ---------------------------------------------------------------------------
# Page config & CSS
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="FTE Baseload Planner",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown(f"""
<style>
    /* ── Force light mode ── */
    html, .stApp, .main, section[data-testid="stSidebar"] {{
        color-scheme: light !important;
        background-color: {MCK_WHITE} !important;
        color: {MCK_DARK} !important;
    }}
    @media (prefers-color-scheme: dark) {{
        html, .stApp, .main {{ color-scheme: light !important; background-color: {MCK_WHITE} !important; color: {MCK_DARK} !important; }}
    }}

    /* ── Base typography ── */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    :root {{ --primary-color: {MCK_NAVY}; }}
    .stApp {{
        font-family: 'Inter', 'Helvetica Neue', Arial, sans-serif;
        -webkit-font-smoothing: antialiased;
    }}
    .main .block-container {{ padding-top: 1.5rem; max-width: 1200px; }}

    /* ── Hide Streamlit chrome ── */
    [data-testid="collapsedControl"] {{ display: none; }}
    [data-testid="stStatusWidget"] {{ display: none; }}
    [data-testid="stToolbar"] {{ display: none; }}
    header[data-testid="stHeader"] {{ display: none; }}

    /* ── Tabs (general) ── */
    .stTabs [data-baseweb="tab-highlight"] {{ background-color: {MCK_BLUE} !important; height: 3px !important; }}
    .stTabs [data-baseweb="tab"] {{
        color: {MCK_GREY}; transition: color 0.25s cubic-bezier(.4,0,.2,1);
        font-size: 1rem; padding: 0.6rem 1.2rem;
    }}
    .stTabs [aria-selected="true"] {{ color: {MCK_NAVY} !important; font-weight: 700; }}
    .stTabs [data-baseweb="tab-list"] {{ gap: 0.5rem; }}


    /* ── Buttons ── */
    button[kind="primary"], .stDownloadButton button {{
        background: linear-gradient(135deg, #0a1628, #163042) !important;
        border: none !important; color: {MCK_WHITE} !important;
        border-radius: 12px !important; font-weight: 600 !important;
        box-shadow: 0 2px 10px rgba(5,28,44,0.15) !important;
        transition: all 0.25s cubic-bezier(.4,0,.2,1) !important;
    }}
    button[kind="primary"]:hover, .stDownloadButton button:hover {{
        background: linear-gradient(135deg, {MCK_BLUE}, #0085b8) !important;
        box-shadow: 0 6px 20px rgba(0,163,224,0.25) !important;
        transform: translateY(-1px) !important;
    }}
    button[kind="primary"] *, .stDownloadButton button * {{ color: {MCK_WHITE} !important; }}

    .stSlider [data-baseweb="slider"] div[role="slider"] {{ background: {MCK_NAVY} !important; }}
    a {{ color: {MCK_BLUE}; }}

    /* ── Hero header (gradient banner) ── */
    .mck-header {{
        background: linear-gradient(135deg, #0a1628 0%, #122240 55%, #163042 100%);
        color: white; padding: 1.8rem 2rem; border-radius: 12px;
        margin-bottom: 1.2rem; position: relative; overflow: hidden;
        box-shadow: 0 8px 32px rgba(5,28,44,0.25);
    }}
    .mck-header::before {{
        content: ''; position: absolute; inset: 0;
        background: radial-gradient(ellipse at center, rgba(0,163,224,0.08) 0%, transparent 70%);
        pointer-events: none;
    }}
    .mck-header h1 {{
        margin: 0; font-size: 1.5rem; font-weight: 700;
        letter-spacing: -0.5px; position: relative;
    }}
    .mck-header .accent-line {{
        width: 60px; height: 3px; margin: 0.5rem 0;
        background: linear-gradient(90deg, {MCK_BLUE}, rgba(0,163,224,0.3));
        border-radius: 3px; position: relative;
    }}
    .mck-header p {{
        margin: 0; font-size: 0.88rem; opacity: 0.85; position: relative;
    }}

    /* ── Config header ── */
    .config-header {{
        background: rgba(255,255,255,0.97); border: 1px solid rgba(5,28,44,0.08);
        border-radius: 12px; padding: 1.2rem 1.6rem; margin-bottom: 1.2rem;
        box-shadow: 0 4px 24px rgba(5,28,44,0.06);
        position: relative; overflow: hidden;
    }}
    .config-header::before {{
        content: ''; display: none;
    }}
    .config-header h2 {{
        margin: 0 0 0 0.5rem; font-size: 1.05rem; font-weight: 700;
        color: {MCK_NAVY}; letter-spacing: -0.3px;
    }}
    .config-header p {{
        margin: 0.25rem 0 0 0.5rem; font-size: 0.8rem; color: {MCK_GREY}; line-height: 1.5;
    }}

    /* ── KPI cards ── */
    .kpi-row {{ display: flex; gap: 1rem; margin-bottom: 1.5rem; }}
    .kpi-card {{
        flex: 1; background: rgba(255,255,255,0.97);
        border: 1px solid rgba(5,28,44,0.08); border-radius: 12px;
        padding: 1.2rem 1.4rem; box-shadow: 0 4px 24px rgba(5,28,44,0.06);
        position: relative; overflow: hidden;
    }}
    .kpi-card::before {{
        content: ''; display: none;
    }}
    .kpi-card .kpi-label {{
        font-size: 10px; font-weight: 600; color: {MCK_GREY};
        text-transform: uppercase; letter-spacing: 1.2px; margin-bottom: 0.35rem;
    }}
    .kpi-card .kpi-value {{ font-size: 1.5rem; font-weight: 700; color: {MCK_NAVY}; }}
    .kpi-card .kpi-sub {{ font-size: 0.75rem; color: #737d8c; margin-top: 0.15rem; }}

    /* ── Section cards ── */
    .card {{
        background: transparent; border: none;
        border-radius: 0; padding: 0 0 0.8rem 0; margin-bottom: 1.2rem;
        box-shadow: none;
        position: relative; overflow: visible;
    }}
    .card::before {{
        content: ''; display: none;
    }}
    .card h5 {{
        background: #E3F0FA; border-radius: 6px;
        padding: 0.85rem 1.4rem; margin: 0 0 1.2rem -0.5rem;
        color: {MCK_NAVY}; font-size: 12px; font-weight: 700;
        letter-spacing: 0.08em; line-height: 1.4;
        text-transform: uppercase;
    }}
    .card h5 .card-sub {{
        font-size: 12px; font-weight: 400; color: {MCK_NAVY};
        margin-left: 0; letter-spacing: 0.04em;
        opacity: 0.7; text-transform: none;
    }}

    .help-text {{
        font-size: 13px; color: #737d8c; line-height: 1.5;
        margin-top: -0.2rem; margin-bottom: 1rem;
    }}

    .section-intro {{
        font-size: 14px; color: #4a5568; margin-bottom: 1rem; line-height: 1.6;
    }}

    /* ── How-It-Works visual blocks ── */
    .hiw-step {{
        display: flex; gap: 1rem; align-items: flex-start;
        margin-bottom: 1.5rem; padding: 1.2rem; background: {MCK_WHITE};
        border: 1px solid #E0E4E8; border-radius: 10px;
    }}
    .hiw-step-num {{
        background: {MCK_NAVY}; color: {MCK_WHITE};
        font-size: 1rem; font-weight: 700; min-width: 36px; height: 36px;
        line-height: 36px; border-radius: 50%; text-align: center; flex-shrink: 0;
    }}
    .hiw-step-body {{ flex: 1; }}
    .hiw-step-body h5 {{
        margin: 0 0 0.4rem 0; font-size: 0.95rem; font-weight: 600; color: {MCK_NAVY};
        background: none; padding: 0; border-radius: 0; text-transform: none; letter-spacing: 0;
    }}
    .hiw-step-body p {{ margin: 0 0 0.4rem 0; font-size: 0.85rem; color: #374151; line-height: 1.55; }}
    .hiw-step-body .hiw-formula {{
        display: inline-block; background: #EBF4FA; padding: 0.35rem 0.7rem;
        border-radius: 6px; font-family: 'Courier New', monospace;
        font-size: 0.82rem; color: {MCK_NAVY}; font-weight: 600; margin: 0.3rem 0;
    }}
    .hiw-step-body ul {{ margin: 0.3rem 0 0 1.1rem; padding: 0; }}
    .hiw-step-body li {{ font-size: 0.85rem; color: #374151; margin-bottom: 0.2rem; line-height: 1.5; }}
    .hiw-concept-row {{
        display: flex; gap: 1rem; margin: 1rem 0 1.5rem 0; flex-wrap: wrap;
    }}
    .hiw-concept {{
        flex: 1; min-width: 170px; background: #EBF4FA; border-radius: 10px;
        padding: 1rem 1.1rem; text-align: left;
    }}
    .hiw-concept h6 {{
        margin: 0 0 0.3rem 0; font-size: 0.82rem; font-weight: 700;
        color: {MCK_NAVY}; text-transform: uppercase; letter-spacing: 0.02em;
    }}
    .hiw-concept p {{
        margin: 0; font-size: 0.8rem; color: #374151; line-height: 1.5;
    }}

    .big-btn {{
        display: flex; justify-content: center; margin-top: 1.5rem; margin-bottom: 1rem;
    }}

    /* ── Tables ── */
    .stDataFrame table {{ font-size: 0.82rem; }}
    .stDataFrame thead tr th {{ border-bottom: none !important; }}
    .stDataFrame [data-testid="stDataFrameResizableContainer"] thead tr th {{
        border-bottom: none !important;
    }}

    /* ── Flow diagram ── */
    .flow-row {{
        display: flex; align-items: center; justify-content: center;
        gap: 0; margin: 1rem 0 1.8rem 0; flex-wrap: nowrap;
    }}
    .flow-box {{
        background: rgba(255,255,255,0.97); border: 1px solid rgba(5,28,44,0.08);
        border-radius: 12px; padding: 0.9rem 1.1rem; text-align: center;
        min-width: 135px; max-width: 190px; flex-shrink: 0;
        box-shadow: 0 2px 10px rgba(5,28,44,0.05);
        transition: all 0.25s cubic-bezier(.4,0,.2,1);
    }}
    .flow-box:hover {{
        box-shadow: 0 6px 20px rgba(5,28,44,0.1);
        transform: translateY(-2px);
    }}
    .flow-box .flow-num {{
        display: inline-block;
        background: linear-gradient(135deg, {MCK_NAVY}, #163042);
        color: {MCK_WHITE}; font-size: 0.7rem; font-weight: 700;
        width: 22px; height: 22px; line-height: 22px;
        border-radius: 50%; text-align: center; margin-bottom: 0.3rem;
    }}
    .flow-box .flow-title {{
        font-size: 0.82rem; font-weight: 700; color: {MCK_NAVY};
        margin-bottom: 0.25rem; letter-spacing: -0.3px;
    }}
    .flow-box .flow-formula {{
        font-size: 0.67rem; color: #737d8c; font-weight: 500;
        font-family: 'Consolas', 'SF Mono', monospace;
        margin-bottom: 0.15rem;
    }}
    .flow-box .flow-desc {{
        font-size: 0.67rem; color: #737d8c; line-height: 1.35;
    }}
    .flow-arrow {{
        font-size: 1.4rem; color: {MCK_BLUE}; padding: 0 0.4rem; flex-shrink: 0;
        font-weight: 600;
    }}

    /* ── Expanders ── */
    .streamlit-expanderHeader {{
        font-weight: 600 !important; color: {MCK_NAVY} !important;
        background: {MCK_LIGHT} !important; border-radius: 12px !important;
    }}

    /* ── Global transition ── */
    @media (prefers-reduced-motion: no-preference) {{
        button, .flow-box, .kpi-card {{
            transition: all 0.25s cubic-bezier(.4,0,.2,1);
        }}
    }}
    @media (prefers-reduced-motion: reduce) {{
        *, *::before, *::after {{ transition-duration: 0.01ms !important; animation-duration: 0.01ms !important; }}
    }}

    /* ── Subtle remove (✕) buttons ── */
    button[kind="secondary"] {{
        border: none !important;
        background: transparent !important;
        color: #C0C0C0 !important;
        font-size: 0.75rem !important;
        padding: 0.15rem 0.4rem !important;
        min-height: 0 !important;
        box-shadow: none !important;
    }}
    button[kind="secondary"]:hover {{
        color: #00A9F4 !important;
        background: rgba(0,169,244,0.08) !important;
    }}
    button[kind="secondary"] * {{
        color: inherit !important;
    }}
</style>
""", unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# Session state
# ---------------------------------------------------------------------------
if "scenarios" not in st.session_state:
    st.session_state.scenarios = [{"name": "Scenario 1", "cfg": default_baseline()}]
if "page" not in st.session_state:
    st.session_state.page = "configure"
if "scenario_results" not in st.session_state:
    st.session_state.scenario_results = []

# Migrate old dict contingency_pct → single float
for _scen in st.session_state.scenarios:
    _cfg = _scen.get("cfg")
    if _cfg and isinstance(_cfg.contingency_pct, dict):
        _cfg.contingency_pct = max(_cfg.contingency_pct.values()) if _cfg.contingency_pct else 0.0


def _clear_scenario_keys(from_idx: int):
    """Remove cached widget keys for scenarios at or above *from_idx*
    so that Streamlit picks up the real cfg values on re-render."""
    import re
    pattern = re.compile(r"^s(\d+)_")
    to_delete = [k for k in st.session_state
                 if (m := pattern.match(k)) and int(m.group(1)) >= from_idx]
    for k in to_delete:
        del st.session_state[k]


def _sync_archetypes(cfg: ModelConfig):
    for arch in cfg.archetypes:
        # Collect the union of roles defined across this archetype's stages
        arch_roles: set = set()
        for sp in arch.stages.values():
            arch_roles.update(sp.fte_per_role.keys())
        if not arch_roles:
            arch_roles = set(cfg.workforce_roles)

        for sname in cfg.pipeline_stages:
            if sname not in arch.stages:
                default_roles = {r: 1.0 for r in arch_roles}
                arch.stages[sname] = StageParams(9, 8.0, 8.0, default_roles)
        extra = [k for k in arch.stages if k not in cfg.pipeline_stages]
        for k in extra:
            del arch.stages[k]
        ordered = {s: arch.stages[s] for s in cfg.pipeline_stages if s in arch.stages}
        arch.stages = ordered
        # Ensure all stages within this archetype share the same role set
        for sp in arch.stages.values():
            for role in arch_roles:
                if role not in sp.fte_per_role:
                    sp.fte_per_role[role] = 0.0


# ---------------------------------------------------------------------------
# Header
# ---------------------------------------------------------------------------
def _render_header(subtitle: str):
    st.markdown(f"""
    <div class="mck-header">
        <h1>FTE Baseload Planning Tool</h1>
        <div class="accent-line"></div>
        <p>{subtitle}</p>
    </div>
    """, unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════
# REUSABLE: Scenario config form
# ═══════════════════════════════════════════════════════════════════════════
def _render_scenario_form(idx: int):
    """Render the full configuration form for scenario at index *idx*."""
    P = f"s{idx}_"
    scen = st.session_state.scenarios[idx]
    cfg = scen["cfg"]

    # 1. Scenario name
    new_name = st.text_input("Scenario name", value=scen["name"], key=f"{P}name")
    scen["name"] = new_name

    # 2. Budget & Timeline (left) | Project Stages (right) ────────────────
    st.markdown("")
    col_left, col_right = st.columns(2)

    with col_left:
        st.markdown('<div class="card"><h5>Budget &amp; Timeline &mdash; <span class="card-sub">Annual spend, overhead, and funding model</span></h5>', unsafe_allow_html=True)
        cfg.total_budget_m = st.number_input(
            "Annual R&D budget (M)", value=cfg.total_budget_m, min_value=1.0,
            step=10.0, format="%.0f", key=f"{P}budget",
            help="Total yearly R&D spend, before any deductions",
        )
        overhead_pct = st.slider(
            "Overhead (%)", 0, 60,
            int(cfg.overhead_pct * 100), 5, "%d%%", key=f"{P}overhead",
            help="Admin, facilities, and management costs — subtracted from budget before funding projects",
        )
        cfg.overhead_pct = overhead_pct / 100.0
        avail = cfg.total_budget_m * (1 - cfg.overhead_pct)
        st.markdown(
            f'<div class="help-text">Net budget: <strong>{avail:,.0f} M</strong></div>',
            unsafe_allow_html=True,
        )
        _mode_options = ["Cash-flow", "Commitment"]
        _mode_map = {"Cash-flow": "cashflow", "Commitment": "commitment"}
        _mode_map_inv = {v: k for k, v in _mode_map.items()}
        _mode_sel = st.radio(
            "Budget model", _mode_options,
            index=_mode_options.index(_mode_map_inv.get(cfg.budget_mode, "Cash-flow")),
            horizontal=True, key=f"{P}bmode",
            help=(
                "**Cash-flow**: annual budget covers ongoing project costs first; "
                "only the remainder funds new starts (project counts vary year to year).  \n\n"
                "**Commitment**: each year's budget funds the full lifecycle cost of new "
                "projects upfront (same count every year within a phase)."
            ),
        )
        cfg.budget_mode = _mode_map[_mode_sel]

        yc1, yc2 = st.columns(2)
        with yc1:
            cfg.start_year = int(st.number_input(
                "First year of new projects", value=cfg.start_year, step=1,
                key=f"{P}start_yr", help="Year new project intake begins",
            ))
        with yc2:
            cfg.end_year = int(st.number_input(
                "Last year of new projects", value=cfg.end_year, step=1,
                key=f"{P}end_yr",
                help="Last year new projects are started — existing projects continue beyond this",
            ))
            if cfg.end_year <= cfg.start_year:
                cfg.end_year = cfg.start_year + 1
        st.markdown('</div>', unsafe_allow_html=True)

    with col_right:
        st.markdown('<div class="card"><h5>Project Stages &mdash; <span class="card-sub">The Journey Every Project Takes</span></h5>', unsafe_allow_html=True)
        st.markdown("""<div class="help-text">
            Projects go through stages from early research to late development.
            <br><br>
            <strong>% start here</strong> = of all new projects each year, what share enters at this stage
            <br><br>
            <strong>% move to next</strong> = of projects that finish this stage, what share continues to the next one
        </div>""", unsafe_allow_html=True)

        stages_to_remove = None
        hc1, hc2, hc3, hc4 = st.columns([3, 2, 2, 1])
        with hc1:
            st.markdown("**Stage name**", help=None)
        with hc2:
            st.markdown("**% start here**")
        with hc3:
            st.markdown("**% move to next**")
        with hc4:
            st.markdown("")

        for si, sname in enumerate(cfg.pipeline_stages):
            sc1, sc2, sc3, sc4 = st.columns([3, 2, 2, 1])
            with sc1:
                new_sname = st.text_input("Name", value=sname, key=f"{P}sn_{si}",
                                          label_visibility="collapsed")
            with sc2:
                alloc = st.number_input(
                    "Start", value=int(cfg.stage_mix.get(sname, 0) * 100),
                    min_value=0, max_value=100, step=5, key=f"{P}sa_{si}",
                    label_visibility="collapsed",
                )
            with sc3:
                is_terminal = si == len(cfg.pipeline_stages) - 1
                if not is_terminal:
                    conv = st.number_input(
                        "Move", value=int(cfg.stage_conversion_rates.get(sname, 0) * 100),
                        min_value=0, max_value=100, step=5, key=f"{P}sc_{si}",
                        label_visibility="collapsed",
                    )
                else:
                    st.markdown("—")
                    conv = None
            with sc4:
                if len(cfg.pipeline_stages) > 1:
                    if st.button("✕", key=f"{P}sr_{si}"):
                        stages_to_remove = si

            if new_sname != sname and new_sname.strip():
                old = sname
                cfg.pipeline_stages[si] = new_sname
                if old in cfg.stage_mix:
                    cfg.stage_mix[new_sname] = cfg.stage_mix.pop(old)
                if old in cfg.stage_conversion_rates:
                    cfg.stage_conversion_rates[new_sname] = cfg.stage_conversion_rates.pop(old)
                for arch in cfg.archetypes:
                    if old in arch.stages:
                        arch.stages[new_sname] = arch.stages.pop(old)
                sname = new_sname

            cfg.stage_mix[sname] = alloc / 100.0
            if conv is not None:
                cfg.stage_conversion_rates[sname] = conv / 100.0

        if stages_to_remove is not None:
            removed = cfg.pipeline_stages.pop(stages_to_remove)
            cfg.stage_mix.pop(removed, None)
            cfg.stage_conversion_rates.pop(removed, None)
            for arch in cfg.archetypes:
                arch.stages.pop(removed, None)
            st.rerun()

        if st.button("＋ Add stage", key=f"{P}add_stage"):
            new_s = f"Stage {len(cfg.pipeline_stages) + 1}"
            cfg.pipeline_stages.append(new_s)
            cfg.stage_mix[new_s] = 0.0
            if len(cfg.pipeline_stages) >= 2:
                prev = cfg.pipeline_stages[-2]
                cfg.stage_conversion_rates.setdefault(prev, 0.50)
            _sync_archetypes(cfg)
            st.rerun()

        alloc_sum = sum(cfg.stage_mix.get(s, 0) for s in cfg.pipeline_stages)
        if abs(alloc_sum - 1.0) > 0.01:
            st.warning(f'"% start here" total: {alloc_sum*100:.0f}% — should be 100%')
        else:
            st.success("Stage percentages add up to 100%")

        # Phase 2 toggle (inside right column, under stages)
        p2_on = st.checkbox(
            "Change stage allocation in later years",
            value=cfg.phase2_start_year > 0, key=f"{P}p2_toggle",
            help="Use a different \"% start here\" split from a chosen year onward.",
        )
        if p2_on:
            p2_year = int(st.number_input(
                "Shift allocation from year",
                value=max(cfg.phase2_start_year, cfg.start_year + 1),
                min_value=cfg.start_year + 1, max_value=cfg.end_year,
                step=1, key=f"{P}p2_year",
            ))
            cfg.phase2_start_year = p2_year
            p1c, p2c = st.columns(2)
            with p1c:
                st.markdown(f"**Phase 1 — {cfg.start_year} to {p2_year - 1}**")
                for si, sname in enumerate(cfg.pipeline_stages):
                    st.markdown(f"**{sname}:** {cfg.stage_mix.get(sname, 0)*100:.0f}%")
            with p2c:
                st.markdown(f"**Phase 2 — {p2_year} onward**")
                for si, sname in enumerate(cfg.pipeline_stages):
                    default_p2 = cfg.stage_mix_phase2.get(sname, cfg.stage_mix.get(sname, 0))
                    p2_alloc = st.number_input(
                        sname, value=int(default_p2 * 100),
                        min_value=0, max_value=100, step=5,
                        key=f"{P}p2a_{si}", label_visibility="visible",
                    )
                    cfg.stage_mix_phase2[sname] = p2_alloc / 100.0
            p2_sum = sum(cfg.stage_mix_phase2.get(s, 0) for s in cfg.pipeline_stages)
            if abs(p2_sum - 1.0) > 0.01:
                st.warning(f"Phase 2 percentages add up to {p2_sum*100:.0f}% — should be 100%")
            else:
                st.success("Phase 2 percentages add up to 100%")
            st.caption("Conversion rates (% move to next) remain the same across both phases.")
        else:
            cfg.phase2_start_year = 0
            cfg.stage_mix_phase2 = {}
        st.markdown('</div>', unsafe_allow_html=True)

    # 3. Project Type Details ─────────────────────────────────────────────
    st.markdown("<div style='margin-top: 0.8rem'></div>", unsafe_allow_html=True)
    st.markdown(
        '<div class="card"><h5>Portfolio Mix &mdash; '
        '<span class="card-sub">What types of R&amp;D projects do you run?</span></h5>',
        unsafe_allow_html=True,
    )

    # Portfolio mix — name, share, remove ─────────────────────────────────
    st.markdown(
        '<div class="help-text">Your R&amp;D portfolio is a mix of different project types '
        '(e.g.&nbsp;chemistry, hardware, algorithm). Set what percentage of your projects belong '
        'to each type. These must add up to 100%.</div>',
        unsafe_allow_html=True,
    )
    arch_to_remove = None
    for ai, arch in enumerate(cfg.archetypes):
        nc, sc, xc = st.columns([3, 2, 0.5])
        with nc:
            new_name = st.text_input(
                "Type name", value=arch.name, key=f"{P}an_{ai}",
                label_visibility="collapsed",
            )
            arch.name = new_name
        with sc:
            share = st.number_input(
                f"Share (%)", value=round(arch.portfolio_share * 100, 1),
                min_value=0.0, max_value=100.0, step=1.0, format="%.1f",
                key=f"{P}ps_{ai}", label_visibility="collapsed",
            )
            arch.portfolio_share = share / 100.0
        with xc:
            if len(cfg.archetypes) > 1:
                if st.button("✕", key=f"{P}ra_{ai}"):
                    arch_to_remove = ai
    if arch_to_remove is not None:
        cfg.archetypes.pop(arch_to_remove)
        _clear_scenario_keys(idx)
        st.rerun()

    total_share = sum(a.portfolio_share for a in cfg.archetypes)
    if abs(total_share - 1.0) > 0.01 and cfg.archetypes:
        st.warning(f"Shares sum to {total_share*100:.0f}% — should be 100%")

    if st.button("＋ Add project type", key=f"{P}add_arch"):
        existing_roles = list(cfg.archetypes[0].stages.values())[0].fte_per_role.keys() \
            if cfg.archetypes and cfg.archetypes[0].stages else cfg.workforce_roles
        default_roles = {r: 1.0 for r in existing_roles}
        new_stages = {s: StageParams(9, 8.0, 8.0, dict(default_roles))
                      for s in cfg.pipeline_stages}
        new_name = f"Type {len(cfg.archetypes) + 1}"
        cfg.archetypes.append(Archetype(name=new_name, portfolio_share=0.0,
                                        stages=new_stages))
        st.rerun()

    st.markdown("<div style='margin-top: 0.5rem'></div>", unsafe_allow_html=True)
    st.markdown(
        '<div class="card"><h5>Project Type Details &mdash; <span class="card-sub">What does each project type look like stage by stage?</span></h5>',
        unsafe_allow_html=True,
    )
    st.markdown(
        '<div class="help-text">For each project type and stage, set your best estimate for '
        'duration, cost range, and team size (FTE per role).</div>',
        unsafe_allow_html=True,
    )
    # Archetype tabs ──────────────────────────────────────────────────────
    _sync_archetypes(cfg)
    if cfg.archetypes:
        arch_tabs = st.tabs([a.name for a in cfg.archetypes])

        for ai, arch in enumerate(cfg.archetypes):
            with arch_tabs[ai]:
                # Determine roles for this archetype
                arch_roles = list(dict.fromkeys(
                    role for sp in arch.stages.values()
                    for role in sp.fte_per_role
                ))
                if not arch_roles:
                    arch_roles = list(cfg.workforce_roles)

                # Role management for this archetype ──────────────────────
                st.markdown(
                    '<span style="color:#051C2C;font-weight:700;font-size:0.8rem;'
                    'letter-spacing:0.03em;">Roles</span>',
                    unsafe_allow_html=True,
                )
                role_to_remove = None
                n_rl = max(len(arch_roles), 1)
                col_spec = []
                for ri in range(n_rl):
                    col_spec += [3, 0.5]
                col_spec.append(2)
                rl_cols = st.columns(col_spec)
                for ri, role in enumerate(arch_roles):
                    with rl_cols[ri * 2]:
                        new_rn = st.text_input(
                            f"Role {ri+1}", value=role,
                            key=f"{P}rn_{ai}_{ri}",
                        )
                        if new_rn != role and new_rn.strip():
                            for sp in arch.stages.values():
                                if role in sp.fte_per_role:
                                    sp.fte_per_role[new_rn] = sp.fte_per_role.pop(role)
                            _clear_scenario_keys(idx)
                            st.rerun()
                    with rl_cols[ri * 2 + 1]:
                        st.markdown("<div style='height:1.65rem'></div>",
                                    unsafe_allow_html=True)
                        if len(arch_roles) > 1:
                            if st.button("✕", key=f"{P}rm_role_{ai}_{ri}",
                                         help="Remove this role"):
                                role_to_remove = ri
                with rl_cols[-1]:
                    st.markdown("<div style='height:1.65rem'></div>",
                                unsafe_allow_html=True)
                    if st.button("＋ Add role", key=f"{P}add_role_{ai}"):
                        n = len(arch_roles) + 1
                        new_role = f"Role {n}"
                        for sp in arch.stages.values():
                            sp.fte_per_role[new_role] = 0.0
                        _clear_scenario_keys(idx)
                        st.rerun()
                if role_to_remove is not None:
                    removed = arch_roles[role_to_remove]
                    for sp in arch.stages.values():
                        sp.fte_per_role.pop(removed, None)
                    _clear_scenario_keys(idx)
                    st.rerun()

                # Stage details ───────────────────────────────────────────
                st.markdown("<div style='margin-top: 1rem'></div>", unsafe_allow_html=True)
                st.markdown(
                    '<span style="color:#051C2C;font-weight:700;font-size:0.8rem;'
                    'letter-spacing:0.03em;">Stages</span>',
                    unsafe_allow_html=True,
                )
                for _si_stage, sname in enumerate(cfg.pipeline_stages):
                    if sname not in arch.stages:
                        continue
                    sp = arch.stages[sname]
                    if _si_stage > 0:
                        st.markdown("<div style='margin-top: 0.4rem'></div>", unsafe_allow_html=True)
                    st.markdown(f"**{sname}**")

                    n_fixed_cols = 3
                    n_role_cols = len(arch_roles)
                    col_widths = [2] * n_fixed_cols + [2] * n_role_cols
                    cols = st.columns(col_widths)

                    with cols[0]:
                        sp.duration_months = st.number_input(
                            "Duration (months)", value=sp.duration_months,
                            min_value=1, step=1, key=f"{P}dm_{ai}_{sname}",
                        )
                    with cols[1]:
                        sp.cost_min = st.number_input(
                            "Cost min (M)", value=sp.cost_min,
                            min_value=0.0, step=0.5, format="%.1f",
                            key=f"{P}cmin_{ai}_{sname}",
                            help="Minimum project cost for this stage",
                        )
                    with cols[2]:
                        sp.cost_max = st.number_input(
                            "Cost max (M)", value=sp.cost_max,
                            min_value=0.0, step=0.5, format="%.1f",
                            key=f"{P}cmax_{ai}_{sname}",
                            help="Maximum project cost. Model uses midpoint as expected cost.",
                        )
                        if sp.cost_max < sp.cost_min:
                            sp.cost_max = sp.cost_min

                    for rr, role in enumerate(arch_roles):
                        with cols[n_fixed_cols + rr]:
                            val = sp.fte_per_role.get(role, 0.0)
                            new_val = st.number_input(
                                f"{role} / project", value=val,
                                min_value=0.0, step=0.5, format="%.1f",
                                key=f"{P}fte_{ai}_{sname}_{rr}",
                            )
                            sp.fte_per_role[role] = new_val

                    if sp.cost_min != sp.cost_max:
                        st.caption(f"Expected cost: {sp.cost_millions:.1f} M (midpoint)")

    st.markdown('</div>', unsafe_allow_html=True)

    # 4. Advanced Settings ─────────────────────────────────────────────────
    st.markdown("<div style='margin-top: 0.8rem'></div>", unsafe_allow_html=True)
    st.markdown(
        '<div class="card"><h5>Advanced Settings &mdash; '
        '<span class="card-sub">Utilization, ramp-up, intake &amp; contingency</span></h5>',
        unsafe_allow_html=True,
    )
    ac1, ac2 = st.columns(2)
    with ac1:
        util_pct = st.slider(
            "Utilization rate", 50, 100, int(cfg.utilization_rate * 100), 5, "%d%%",
            key=f"{P}util",
            help="Fraction of time an FTE spends on project work. If 80%, gross headcount = model FTE ÷ 0.80.",
        )
        cfg.utilization_rate = util_pct / 100.0
    with ac2:
        cfg.ramp_months = st.slider(
            "Ramp-up period (months)", 0, 6, cfg.ramp_months, 1,
            key=f"{P}ramp",
            help="Projects ramp FTE linearly. Month 1 of a 3-month ramp = 33% staffing, month 2 = 67%, month 3+ = 100%.",
        )
    ac3, ac4 = st.columns(2)
    with ac3:
        cfg.intake_spread_months = st.slider(
            "Intake window (months/year)", 1, 12, cfg.intake_spread_months, 1,
            key=f"{P}intake",
            help="New projects start evenly across the first N months of each year.",
        )
    with ac4:
        cont_val = st.slider(
            "Contingency buffer", 0, 50, int(cfg.contingency_pct * 100), 5, "%d%%",
            key=f"{P}cont",
            help="Optional buffer on top of the model's FTE. Accounts for uncertainty, attrition, leave, or estimation error. At 0% the adjusted numbers equal the base numbers.",
        )
        cfg.contingency_pct = cont_val / 100.0
    st.markdown('</div>', unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════
# REUSABLE: Single scenario results
# ═══════════════════════════════════════════════════════════════════════════
def _render_single_result(name: str, cfg: ModelConfig, result, key_prefix: str = "r0"):
    """Render the full results view for one scenario."""
    P = f"{key_prefix}_"
    cont = cfg.contingency_pct
    has_contingency = cont > 0
    is_cashflow = cfg.budget_mode == "cashflow"
    has_phase2 = cfg.phase2_start_year > 0 and bool(cfg.stage_mix_phase2)
    monthly = result.monthly

    has_cost_range = (
        isinstance(result.cost_low_annual, pd.DataFrame)
        and not result.cost_low_annual.empty
    )

    adj_ss_avg = result.steady_state_avg * (1 + cont) if has_contingency else result.steady_state_avg

    _cont_range_sub = ""
    _cont_ss_sub = ""
    if has_contingency:
        adj_min = result.steady_state_min_month * (1 + cont)
        adj_max = result.steady_state_max_month * (1 + cont)
        _cont_range_sub = f'<div class="kpi-sub"><strong>{adj_min:,.0f} – {adj_max:,.0f}</strong> with contingency</div>'
        _cont_ss_sub = f'<div class="kpi-sub"><strong>{adj_ss_avg:,.0f}</strong> with contingency</div>'

    _cost_range_sub = ""
    if has_cost_range:
        _cost_range_sub = (
            f'<div class="kpi-sub">Cost range: '
            f'<strong>{result.cost_high_ss_avg:,.0f}</strong> (high cost) – '
            f'<strong>{result.cost_low_ss_avg:,.0f}</strong> (low cost)</div>'
        )

    yp = result.yearly_projects

    if is_cashflow and yp:
        total_proj = sum(yp.values())
        yr_parts = [f"{yp.get(y, 0):,.0f}" for y in range(cfg.start_year, cfg.end_year + 1)]
        _proj_kpi_value = f"{total_proj:,.0f}"
        _proj_kpi_label = "Total new projects"
        _proj_kpi_sub = "Per year: " + " → ".join(yr_parts)

        _ann = result.annual_summary
        if not _ann.empty:
            _peak_row = _ann.loc[_ann["Max monthly FTE"].idxmax()]
            _peak_fte = _peak_row["Max monthly FTE"]
            _peak_yr = int(_peak_row["Year"])
        else:
            _peak_fte = result.steady_state_max_month
            _peak_yr = cfg.end_year

        _kpi3_label = "Peak monthly FTE"
        _kpi3_value = f"{_peak_fte:,.0f}"
        _kpi3_sub = f"Highest single-month FTE across all years (in {_peak_yr})"
        _kpi4_label = f"Avg FTE in {cfg.end_year}"
        _kpi4_value = f"{result.steady_state_avg:,.0f}"
        _kpi4_sub = "Average monthly headcount in the final year"
    else:
        _proj_kpi_value = f"{yp.get(cfg.start_year, result.projects_per_year):,.0f}" if yp else f"{result.projects_per_year:,.0f}"
        _proj_kpi_label = "New projects per year"
        _proj_kpi_sub = "Same count every year"
        if has_phase2 and yp:
            p1_val = yp.get(cfg.start_year, 0)
            p2_val = yp.get(cfg.phase2_start_year, 0)
            _proj_kpi_sub = (
                f"{p1_val:,.0f}/yr in {cfg.start_year}\u2013{cfg.phase2_start_year - 1}, "
                f"{p2_val:,.0f}/yr from {cfg.phase2_start_year}"
            )
        _kpi3_label = f"FTE range in {cfg.end_year}"
        _kpi3_value = f"{result.steady_state_min_month:,.0f} \u2013 {result.steady_state_max_month:,.0f}"
        _kpi3_sub = "Min to max monthly FTE — narrows as pipeline stabilizes"
        _kpi4_label = "Steady-state headcount"
        _kpi4_value = f"{result.steady_state_avg:,.0f}"
        _kpi4_sub = f"Avg monthly FTE in {cfg.end_year} — the level the pipeline settles at"

    # KPI cards (headcount first, budget last)
    st.markdown(f"""<div class="kpi-row">
<div class="kpi-card">
<div class="kpi-label">{_kpi4_label}</div>
<div class="kpi-value">{_kpi4_value}</div>
<div class="kpi-sub">{_kpi4_sub}</div>
{_cont_ss_sub}
{_cost_range_sub}
</div>
<div class="kpi-card">
<div class="kpi-label">{_kpi3_label}</div>
<div class="kpi-value">{_kpi3_value}</div>
<div class="kpi-sub">{_kpi3_sub}</div>
{_cont_range_sub}
</div>
<div class="kpi-card">
<div class="kpi-label">{_proj_kpi_label}</div>
<div class="kpi-value">{_proj_kpi_value}</div>
<div class="kpi-sub">{_proj_kpi_sub}</div>
</div>
<div class="kpi-card">
<div class="kpi-label">Budget available for projects</div>
<div class="kpi-value">{cfg.total_budget_m*(1-cfg.overhead_pct):,.0f} M</div>
<div class="kpi-sub">{cfg.total_budget_m:,.0f} M total minus {cfg.overhead_pct*100:.0f}% overhead</div>
</div>
</div>""", unsafe_allow_html=True)

    # Result tabs
    tab_dash, tab_annual, tab_monthly, tab_how, tab_assumptions = st.tabs([
        "Dashboard", "Annual Summary", "Monthly Detail", "How It Works", "Assumptions",
    ])

    # ── Dashboard ──
    with tab_dash:
        if monthly.empty:
            st.info("No data. Check that archetypes and shares are configured.")
        else:
            ann = result.annual_summary
            if not ann.empty:
                _last_avg = ann["Avg monthly FTE"].iloc[-1]
                if is_cashflow:
                    _c1_title = f"Peak demand hits {_peak_fte:,.0f} FTE in {_peak_yr}, settling to ~{_last_avg:,.0f}"
                else:
                    _c1_title = f"Headcount builds to ~{_last_avg:,.0f} FTE by {cfg.end_year}"
                st.markdown(f"#### {_c1_title}")
                st.markdown(
                    '<div class="section-intro">Bars show average monthly FTE per year; '
                    'triangles mark the min and max months.</div>',
                    unsafe_allow_html=True,
                )

                fig_main = go.Figure()
                fig_main.add_trace(go.Bar(
                    x=ann["Year"], y=ann["Avg monthly FTE"], name="Avg monthly FTE",
                    marker_color=MCK_NAVY, opacity=0.85,
                ))
                if has_contingency:
                    fig_main.add_trace(go.Bar(
                        x=ann["Year"],
                        y=ann["Avg monthly FTE"] * (1 + cont),
                        name="Avg FTE (with contingency)",
                        marker_color=MCK_NAVY, opacity=0.3,
                    ))
                if has_cost_range:
                    low_ann = result.cost_low_annual
                    high_ann = result.cost_high_annual
                    if not low_ann.empty:
                        fig_main.add_trace(go.Scatter(
                            x=low_ann["Year"], y=low_ann["Avg monthly FTE"],
                            mode="lines+markers", name="Low cost scenario",
                            line=dict(dash="dash", width=1.5, color=MCK_GREEN),
                            marker=dict(size=5),
                        ))
                    if not high_ann.empty:
                        fig_main.add_trace(go.Scatter(
                            x=high_ann["Year"], y=high_ann["Avg monthly FTE"],
                            mode="lines+markers", name="High cost scenario",
                            line=dict(dash="dash", width=1.5, color="#C62828"),
                            marker=dict(size=5),
                        ))
                fig_main.add_trace(go.Scatter(
                    x=ann["Year"], y=ann["Min monthly FTE"], name="Min month",
                    mode="markers", marker=dict(color=MCK_TEAL, size=8, symbol="triangle-down"),
                ))
                fig_main.add_trace(go.Scatter(
                    x=ann["Year"], y=ann["Max monthly FTE"], name="Max month",
                    mode="markers", marker=dict(color=MCK_BLUE, size=8, symbol="triangle-up"),
                ))
                _final_yr = int(ann["Year"].iloc[-1])
                fig_main.add_annotation(
                    x=_final_yr, y=_last_avg,
                    text=f"<b>{_last_avg:,.0f}</b>",
                    showarrow=True, arrowhead=0, arrowcolor=MCK_NAVY,
                    ax=30, ay=-25,
                    font=dict(size=13, color=MCK_NAVY),
                    bgcolor="white", bordercolor=MCK_NAVY, borderwidth=1, borderpad=3,
                )
                fig_main.update_layout(
                    barmode="overlay" if has_contingency else "relative",
                    height=400,
                    margin=dict(l=20, r=20, t=30, b=20),
                    plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                    font=dict(family="Inter, Helvetica Neue, Arial, sans-serif", size=12, color=MCK_DARK),
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0),
                    xaxis=dict(gridcolor="#E0E0E0", title="", dtick=1),
                    yaxis=dict(gridcolor="#E0E0E0", title="Monthly FTE"),
                    hovermode="x unified",
                )
                st.plotly_chart(fig_main, width="stretch", key=f"{P}chart_main")

            st.divider()

            if not ann.empty:
                _role_final = {}
                for role in cfg.all_roles:
                    col = f"Avg {role} FTE"
                    if col in ann.columns:
                        _role_final[role] = ann[col].iloc[-1]
                _total_final = sum(_role_final.values()) if _role_final else 1
                _dominant_role = max(_role_final, key=_role_final.get) if _role_final else ""
                _dominant_pct = (_role_final.get(_dominant_role, 0) / _total_final * 100) if _total_final > 0 else 0
                _c2_title = f"{_dominant_role} roles drive {_dominant_pct:.0f}% of staffing demand" if _dominant_role else "Workforce split by year"
                st.markdown(f"#### {_c2_title}")
                st.markdown(
                    '<div class="section-intro">Stacked bars show how total FTE breaks down '
                    'across roles each year.</div>',
                    unsafe_allow_html=True,
                )

                fig2 = go.Figure()
                for ri, role in enumerate(cfg.all_roles):
                    col = f"Avg {role} FTE"
                    if col in ann.columns:
                        color = ARCH_COLORS[ri % len(ARCH_COLORS)]
                        fig2.add_trace(go.Bar(
                            x=ann["Year"], y=ann[col], name=role, marker_color=color,
                        ))
                if has_contingency:
                    _cont_total = ann["Avg monthly FTE"] * cont
                    fig2.add_trace(go.Bar(
                        x=ann["Year"], y=_cont_total,
                        name=f"Contingency (+{cont*100:.0f}%)",
                        marker_color="#B0BEC5", opacity=0.5,
                    ))
                fig2.update_layout(
                    barmode="stack", height=360,
                    margin=dict(l=20, r=20, t=30, b=20),
                    plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                    font=dict(family="Inter, Helvetica Neue, Arial, sans-serif", size=12, color=MCK_DARK),
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0),
                    xaxis=dict(gridcolor="#E0E0E0", title="", dtick=1),
                    yaxis=dict(gridcolor="#E0E0E0", title="Avg monthly FTE"),
                    hovermode="x unified",
                )
                st.plotly_chart(fig2, width="stretch", key=f"{P}chart_split")

            # Auto-generated narrative insight
            st.markdown("#### Key drivers")
            st.markdown(
                '<div class="section-intro">Based on your inputs, here\'s what shapes the headcount number:</div>',
                unsafe_allow_html=True,
            )
            _insights = []
            if cfg.archetypes:
                _sorted_arch = sorted(cfg.archetypes, key=lambda a: a.portfolio_share, reverse=True)
                _top = _sorted_arch[0]
                _insights.append(
                    f"**{_top.name}** projects dominate the portfolio at "
                    f"{_top.portfolio_share*100:.0f}% share, making them the largest driver of headcount."
                )
            if not ann.empty and len(ann) >= 2:
                _yr1_avg = ann["Avg monthly FTE"].iloc[0]
                _yr3_avg = ann["Avg monthly FTE"].iloc[min(2, len(ann) - 1)]
                _yr3_label = int(ann["Year"].iloc[min(2, len(ann) - 1)])
                if _yr1_avg > 0:
                    _growth = ((_yr3_avg / _yr1_avg) - 1) * 100
                    _insights.append(
                        f"FTE grows **{_growth:+.0f}%** from {int(ann['Year'].iloc[0])} to {_yr3_label} "
                        f"as the pipeline fills — plan hiring ramp accordingly."
                    )
            if not ann.empty and cfg.all_roles:
                _role_totals = {}
                for role in cfg.all_roles:
                    col = f"Avg {role} FTE"
                    if col in ann.columns:
                        _role_totals[role] = ann[col].iloc[-1]
                if _role_totals:
                    _total_fte = sum(_role_totals.values())
                    _top_role = max(_role_totals, key=_role_totals.get)
                    _top_role_pct = (_role_totals[_top_role] / _total_fte * 100) if _total_fte > 0 else 0
                    _insights.append(
                        f"The **{_top_role}** role makes up {_top_role_pct:.0f}% of total FTE demand in {cfg.end_year}."
                    )
            if has_cost_range:
                _insights.append(
                    f"If project costs hit the high end, headcount drops to "
                    f"**{result.cost_high_ss_avg:,.0f}**; at the low end it rises to "
                    f"**{result.cost_low_ss_avg:,.0f}**."
                )
            if has_contingency:
                _insights.append(
                    f"With the {cont*100:.0f}% contingency buffer, plan for **{adj_ss_avg:,.0f} FTE** "
                    f"rather than the base {result.steady_state_avg:,.0f}."
                )
            if _insights:
                st.markdown("\n".join(f"- {i}" for i in _insights))

            with st.expander("What goes into each project — cost, duration, and staffing by type"):
                bk_rows = []
                for arch in cfg.archetypes:
                    for sn, sp in arch.stages.items():
                        row_data = {
                            "Archetype": arch.name, "Stage": sn,
                            "Share": f"{arch.portfolio_share*100:.0f}%",
                            "Duration": f"{sp.duration_months} mo",
                            "Cost (M)": f"{sp.cost_millions:.1f}",
                        }
                        for role in sp.fte_per_role:
                            row_data[f"{role} FTE"] = f"{sp.fte_per_role[role]:.1f}"
                        bk_rows.append(row_data)
                st.dataframe(pd.DataFrame(bk_rows), use_container_width=True, hide_index=True)

    # ── How It Works ──
    with tab_how:
        _role_names = cfg.all_roles
        _role_list_str = ", ".join(_role_names) if _role_names else "Researcher, Developer"
        _net = cfg.total_budget_m * (1 - cfg.overhead_pct)
        _cont_pct_display = f"{cont*100:.0f}%" if has_contingency else "0%"

        # ── 1. The big question ──────────────────────────────────────────
        st.markdown("""
<div class="context-block">
<h4>What does this tool do?</h4>
<p style="font-size:1rem; line-height:1.7;">
It answers one question: <strong>"Given our R&D budget, how many people do we need to hire?"</strong><br>
You tell it how much money you have and what types of projects you run. It tells you how many staff you need across your defined roles.
</p>
</div>
""", unsafe_allow_html=True)

        # ── 2. Key terms (collapsible) ──────────────────────────────────
        with st.expander("New to this? Key terms explained"):
            st.markdown(f"""
<div class="hiw-concept-row">
    <div class="hiw-concept">
        <h6>FTE</h6>
        <p><strong>Full-Time Equivalent.</strong> One FTE = one person working full time for a year. 0.5 FTE = half a person's time. This model calculates how many FTEs your projects need.</p>
    </div>
    <div class="hiw-concept">
        <h6>Pipeline stage</h6>
        <p>The sequence of stages a project goes through from start to finish ({", ".join(cfg.pipeline_stages)}). Think of it like a funnel: many projects start at the early stage, but only some advance to later stages.</p>
    </div>
</div>
<div class="hiw-concept-row">
    <div class="hiw-concept">
        <h6>Archetype</h6>
        <p>A type of R&D project. Different types cost different amounts, take different times, and need different teams. E.g. {", ".join(a.name for a in cfg.archetypes[:3])}.</p>
    </div>
    <div class="hiw-concept">
        <h6>Roles</h6>
        <p>The types of people working on projects. Your current roles: <strong>{_role_list_str}</strong>. Each project type defines how many of each role it needs per stage.</p>
    </div>
</div>
""", unsafe_allow_html=True)

        # ── 3. Step by step ─────────────────────────────────────────────
        st.markdown("#### How the model works — step by step")
        st.markdown(
            '<div class="section-intro">Follow the numbered steps below to see '
            'exactly how your budget turns into a headcount plan.</div>',
            unsafe_allow_html=True,
        )

        # Step 1: Budget
        st.markdown(f"""
<div class="hiw-step">
    <div class="hiw-step-num">1</div>
    <div class="hiw-step-body">
        <h5>Start with the money</h5>
        <p>You have a total R&D budget. First, subtract overhead (admin, facilities, management). What's left is the money available to actually fund projects.</p>
        <div class="hiw-formula">Net project budget = Total budget &times; (1 &minus; Overhead %)</div>
        <p>Example: {cfg.total_budget_m:,.0f}M total &times; (1 &minus; {cfg.overhead_pct*100:.0f}%) = <strong>{_net:,.0f}M</strong> available for projects.</p>
    </div>
</div>
""", unsafe_allow_html=True)

        # Step 2: Budget → Projects (with cost explanation folded in)
        if is_cashflow:
            st.markdown(f"""
<div class="hiw-step">
    <div class="hiw-step-num">2</div>
    <div class="hiw-step-body">
        <h5>Figure out how many projects you can afford</h5>
        <p>Under <strong>cash-flow budgeting</strong>, the {_net:,.0f}M annual budget must cover <em>everything running that year</em> &mdash; ongoing projects from prior years <strong>and</strong> new projects started this year.</p>
        <p>The model works year by year:</p>
        <ol>
            <li><strong>Ongoing cost</strong> &mdash; for each project cohort still running from prior years, compute: <code>count &times; (cost &divide; duration) &times; months active this year</code>. Sum across all cohorts.</li>
            <li><strong>Available budget</strong> = {_net:,.0f}M &minus; ongoing cost. If ongoing costs exceed the budget, no new projects can start that year.</li>
            <li><strong>Partial-year cost per new project</strong> &mdash; a project started mid-year only consumes part of its lifetime cost this year.</li>
            <li><strong>New projects</strong> = available budget &divide; partial-year cost.</li>
        </ol>
        <p>This means project counts <strong>vary year to year</strong>. A big intake year creates a cost wave that squeezes subsequent years. Once those projects complete, budget frees up for another wave.</p>
        <p><strong>Why cost matters here:</strong> cost doesn't change how many people each project needs &mdash; that's set by the FTE-per-role inputs. What cost changes is <strong>how many projects fit in the budget</strong>. Cheaper projects = more projects = higher FTE. This is why a portfolio shift toward lower-cost archetypes can <em>increase</em> headcount.</p>
    </div>
</div>
""", unsafe_allow_html=True)
        else:
            st.markdown(f"""
<div class="hiw-step">
    <div class="hiw-step-num">2</div>
    <div class="hiw-step-body">
        <h5>Figure out how many projects you can afford</h5>
        <p>Under <strong>commitment budgeting</strong>, each year's {_net:,.0f}M budget funds the <em>full lifecycle cost</em> of new projects upfront. Ongoing costs from prior years are not counted &mdash; each year stands alone.</p>
        <div class="hiw-formula">Projects per year = Net budget &divide; Weighted avg lifecycle cost per project</div>
        <p>The project count is the <strong>same every year</strong> within a phase (it only changes if the stage allocation shifts at Phase 2).</p>
        <p><strong>Why cost matters here:</strong> cost doesn't change how many people each project needs &mdash; that's set by the FTE-per-role inputs. What cost changes is <strong>how many projects fit in the budget</strong>. Cheaper projects = more projects = higher FTE. This is why a portfolio shift toward lower-cost archetypes can <em>increase</em> headcount.</p>
    </div>
</div>
""", unsafe_allow_html=True)

        # Step 3: Distribute (with pipeline visualization inside the card)
        pipe_html_parts = []
        for si, sn in enumerate(cfg.pipeline_stages):
            alloc = cfg.stage_mix.get(sn, 0) * 100
            pipe_html_parts.append(f"<strong>{sn}</strong> ({alloc:.0f}% start here)")
            if si < len(cfg.pipeline_stages) - 1:
                conv = cfg.stage_conversion_rates.get(sn, 0) * 100
                pipe_html_parts.append(f" &nbsp;&rarr;&nbsp; <em>{conv:.0f}% advance</em> &nbsp;&rarr;&nbsp; ")
        pipe_html = "".join(pipe_html_parts)

        phase2_html = ""
        if has_phase2:
            p1_summary = " / ".join(f"{cfg.stage_mix.get(sn,0)*100:.0f}%" for sn in cfg.pipeline_stages)
            p2_summary = " / ".join(f"{cfg.stage_mix_phase2.get(sn,0)*100:.0f}%" for sn in cfg.pipeline_stages)
            stage_labels = " / ".join(cfg.pipeline_stages)
            phase2_html = (
                f'<div style="background:#E8F4FD;border-left:4px solid {MCK_BLUE};border-radius:6px;'
                f'padding:0.7rem 1rem;margin-top:0.6rem;font-size:0.85rem;line-height:1.6;">'
                f'<strong>Allocation shifts in {cfg.phase2_start_year}</strong> ({stage_labels})<br>'
                f'Phase 1: {p1_summary} &middot; Phase 2: {p2_summary}<br>'
                f'<em>Conversion rates (% move to next) stay the same.</em></div>'
            )

        st.markdown(f"""
<div class="hiw-step">
    <div class="hiw-step-num">3</div>
    <div class="hiw-step-body">
        <h5>Distribute projects across types and stages</h5>
        <p>The total project count is split across your project types (archetypes) based on portfolio shares. Within each type, projects are assigned to pipeline stages:</p>
        <ul>
            <li><strong>Direct allocation</strong> &mdash; a fixed percentage of new projects start directly at each stage</li>
            <li><strong>Conversion</strong> &mdash; when early-stage projects finish, a percentage of them "graduate" to the next stage, creating additional projects there</li>
        </ul>
        <p>Later stages get projects from two sources: direct allocation + graduates from the previous stage.</p>
        <div class="hiw-formula">{pipe_html}</div>
        {phase2_html}
    </div>
</div>
""", unsafe_allow_html=True)

        # Step 4: Simulate
        st.markdown(f"""
<div class="hiw-step">
    <div class="hiw-step-num">4</div>
    <div class="hiw-step-body">
        <h5>Simulate the pipeline month by month</h5>
        <p>Each year's new projects are spread across the first few months (the "intake window"). Once a project starts, it stays active for its full duration. The model tracks how many projects are running in each stage, every single month.</p>
        <p>Because projects from different years overlap (Year 1 projects may still be running when Year 2 projects start), headcount <strong>builds up</strong> over the first 2&ndash;3 years.</p>
    </div>
</div>
""", unsafe_allow_html=True)

        # Step 5: Active projects → FTE (with contingency folded in)
        st.markdown(f"""
<div class="hiw-step">
    <div class="hiw-step-num">5</div>
    <div class="hiw-step-body">
        <h5>Convert active projects into people needed</h5>
        <p>Every active project needs a team. Multiply the number of active projects by the staff each project requires, broken down by role ({_role_list_str}).</p>
        <div class="hiw-formula">FTE in a month = Active projects &times; Staff per project</div>
        <p>If utilization is less than 100% (people spend time on admin, training, leave), the model inflates the number. If ramp-up is set, new projects start with a partial team that grows to full strength over a few months.</p>
        <p><strong>Contingency buffer:</strong> the base FTE is the <em>minimum</em> to staff all projects. The contingency column adds a buffer (currently <strong>{_cont_pct_display}</strong>) for uncertainty, attrition, leave, and estimation error{f" &mdash; multiplying base FTE by {1+cont:.2f}." if has_contingency else ". Set it above 0% in Advanced Settings to see adjusted numbers."}</p>
    </div>
</div>
""", unsafe_allow_html=True)

        # ── Understand the FTE profile (expander) ─────────────────────────
        _profile_label = "Why does headcount fluctuate?" if is_cashflow else "Why does headcount stabilize?"
        with st.expander(_profile_label):
            if is_cashflow:
                st.markdown(f"""
<p>In Year 1 the pipeline is empty, so the full {_net:,.0f}M goes to new projects &mdash; a large intake. In Year 2, those projects are still running and consuming budget, leaving little room for new starts. As older projects complete, budget frees up again. This creates a <strong>&ldquo;feast and famine&rdquo; cycle</strong>.</p>
<p>Under cash-flow budgeting there is no true steady state &mdash; use the <strong>final-year average</strong> as your planning target.</p>
<ul>
    <li><strong>Peak monthly FTE</strong> &mdash; busiest month across all years</li>
    <li><strong>Avg FTE in {cfg.end_year}</strong> &mdash; final-year average; use as planning baseline</li>
    <li><strong>Min monthly FTE</strong> &mdash; quietest month (e.g. just before a new cohort starts)</li>
</ul>
""", unsafe_allow_html=True)
            else:
                st.markdown(f"""
<p>In Year 1, projects start but none have finished &mdash; the pipeline only fills up. Headcount keeps climbing until starts roughly equal completions. Once that happens, headcount stabilizes &mdash; this is the <strong>steady state</strong>.</p>
<ul>
    <li><strong>Steady-state headcount</strong> &mdash; average monthly FTE in {cfg.end_year}; the long-run level your hiring plan should target</li>
    <li><strong>Min monthly FTE</strong> &mdash; quietest month (e.g. just before a new cohort starts)</li>
    <li><strong>Max monthly FTE</strong> &mdash; busiest month (e.g. when old and new cohorts overlap most)</li>
</ul>
<p>Early years have a wide range (pipeline is still filling). Later years narrow. When min and max converge, you\u2019ve reached steady state.</p>
""", unsafe_allow_html=True)

        # ── Cost range scenarios (expander) ───────────────────────────────
        with st.expander("How cost uncertainty affects the plan"):
            st.markdown("""
<p>When Cost min &ne; Cost max for any stage, the model runs three times:</p>
<ul>
    <li><strong>Base case</strong> &mdash; midpoint cost (main chart bars)</li>
    <li><strong>Low cost</strong> &mdash; all minimums &rarr; more projects fit &rarr; higher FTE (green dashed line)</li>
    <li><strong>High cost</strong> &mdash; all maximums &rarr; fewer projects fit &rarr; lower FTE (red dashed line)</li>
</ul>
<p>The gap between the dashed lines shows how sensitive headcount is to cost uncertainty. A wide gap means small cost changes have a big staffing impact. If all your Cost min and Cost max values are the same, no dashed lines appear.</p>
""", unsafe_allow_html=True)

        # ── What to do next (standalone card) ─────────────────────────────
        if is_cashflow:
            st.markdown(f"""
<div class="hiw-step">
    <div class="hiw-step-body" style="margin-left:0;">
        <h5>What to do next</h5>
        <ul>
            <li><strong>Use the final-year average FTE</strong> as the planning baseline for long-term hiring</li>
            <li><strong>Use peak monthly FTE</strong> to size the maximum staffing capacity you need</li>
            <li><strong>Look at the year-by-year breakdown</strong> to understand when hiring surges and lulls will occur</li>
            <li><strong>Look at the role split</strong> ({_role_list_str}) to decide which roles to prioritise</li>
            <li><strong>Test sensitivity</strong> &mdash; change one assumption and recalculate to see how it moves the needle</li>
        </ul>
    </div>
</div>
""", unsafe_allow_html=True)
        else:
            st.markdown(f"""
<div class="hiw-step">
    <div class="hiw-step-body" style="margin-left:0;">
        <h5>What to do next</h5>
        <ul>
            <li><strong>Use the steady-state headcount</strong> as the basis for long-term hiring</li>
            <li><strong>Use the yearly range</strong> to plan for seasonal variation in staffing needs</li>
            <li><strong>Look at the role split</strong> ({_role_list_str}) to decide which roles to prioritise</li>
            <li><strong>Test sensitivity</strong> &mdash; change one assumption and recalculate to see how it moves the needle</li>
        </ul>
    </div>
</div>
""", unsafe_allow_html=True)

    # ── Monthly Detail ──
    with tab_monthly:
        if monthly.empty:
            st.info("No data.")
        else:
            st.markdown("#### Month-by-month project load and FTE demand")
            if is_cashflow:
                st.markdown(
                    '<div class="section-intro">'
                    'Under <strong>cash-flow budgeting</strong>, the number of new projects varies year to year '
                    'because each year\u2019s budget must first cover ongoing costs from prior cohorts before funding '
                    'new starts. Use the filters below to focus on specific archetypes, stages, or years.'
                    '</div>',
                    unsafe_allow_html=True,
                )
            else:
                st.markdown(
                    '<div class="section-intro">'
                    'Under <strong>commitment budgeting</strong>, the same number of new projects starts every year '
                    '(within each phase). The table below shows how those projects overlap month to month, building '
                    'up FTE demand as the pipeline fills. Use the filters to focus on specific archetypes, stages, or years.'
                    '</div>',
                    unsafe_allow_html=True,
                )
            disp = monthly.copy()
            disp["Month"] = disp["month"].dt.strftime("%Y-%m")
            disp["Archetype"] = disp["archetype"]
            disp["Stage"] = disp["stage"]
            disp["Active Projects"] = disp["effective_projects"].round(1)
            disp["Total FTE"] = disp["fte_total"].round(1)
            nice_cols = ["Month", "Archetype", "Stage", "Active Projects"]
            for role in cfg.all_roles:
                col_src = f"fte_{role}"
                col_nice = f"{role} FTE"
                if col_src in disp.columns:
                    disp[col_nice] = disp[col_src].round(1)
                    nice_cols.append(col_nice)
            nice_cols.append("Total FTE")

            if has_contingency:
                _cont_col = f"Total + {cont*100:.0f}% contingency"
                disp[_cont_col] = (disp["fte_total"] * (1 + cont)).round(1)
                nice_cols.append(_cont_col)

            fc1, fc2, fc3 = st.columns(3)
            with fc1:
                arch_options = sorted(disp["Archetype"].unique())
                sel_arch = st.multiselect("Archetype", arch_options, default=arch_options, key=f"{P}f_arch")
            with fc2:
                stage_options = sorted(disp["Stage"].unique())
                sel_stage = st.multiselect("Stage", stage_options, default=stage_options, key=f"{P}f_stage")
            with fc3:
                year_options = sorted(y for y in disp["month"].dt.year.unique() if cfg.start_year <= y <= cfg.end_year)
                sel_year = st.multiselect("Year", year_options, default=year_options, key=f"{P}f_year")

            filtered = disp[
                disp["Archetype"].isin(sel_arch) &
                disp["Stage"].isin(sel_stage) &
                disp["month"].dt.year.isin(sel_year)
            ]

            st.dataframe(filtered[nice_cols], use_container_width=True, hide_index=True, height=500)
            st.caption(f"Showing {len(filtered):,} rows")
            st.download_button("Download CSV", filtered[nice_cols].to_csv(index=False),
                               f"fte_monthly_{name}.csv", "text/csv", key=f"{P}dl_monthly")

    # ── Annual Summary ──
    with tab_annual:
        if result.annual_summary.empty:
            st.info("No data.")
        else:
            ann_disp = result.annual_summary.copy()
            _yp_for_ann = result.yearly_projects
            ann_disp.insert(1, "New projects", ann_disp["Year"].map(lambda y: round(_yp_for_ann.get(y, 0), 1)))

            _first_avg = ann_disp["Avg monthly FTE"].iloc[0]
            _last_avg_ann = ann_disp["Avg monthly FTE"].iloc[-1]
            _first_yr = int(ann_disp["Year"].iloc[0])
            _last_yr = int(ann_disp["Year"].iloc[-1])
            if is_cashflow:
                _peak_idx = ann_disp["Avg monthly FTE"].idxmax()
                _peak_val = ann_disp["Avg monthly FTE"].iloc[_peak_idx]
                _peak_year = int(ann_disp["Year"].iloc[_peak_idx])
                _min_avg = ann_disp["Avg monthly FTE"].min()
                _max_avg = ann_disp["Avg monthly FTE"].max()
                _ann_title = f"Average FTE fluctuates between {_min_avg:,.0f} and {_max_avg:,.0f} over {_first_yr}–{_last_yr}"
            else:
                _ann_title = f"Steady-state headcount reaches {_last_avg_ann:,.0f} FTE by {_last_yr} (up from {_first_avg:,.0f} in {_first_yr})"
            st.markdown(f"#### {_ann_title}")

            if is_cashflow:
                st.markdown(
                    '<div class="section-intro">'
                    'Each row summarises one year of the simulation. Because cash-flow budgeting ties new starts to '
                    'remaining budget, project counts (and therefore FTE) fluctuate year to year. '
                    'The <strong>Avg monthly FTE</strong> column is the best single number for hiring targets; '
                    'the <strong>Min/Max</strong> columns show the monthly range within that year.'
                    '</div>',
                    unsafe_allow_html=True,
                )
            else:
                st.markdown(
                    '<div class="section-intro">'
                    'Each row summarises one year. Under commitment budgeting the same number of projects starts every '
                    'year, so FTE builds steadily until the pipeline reaches steady state. '
                    'The <strong>Avg monthly FTE</strong> in the final year is the long-run staffing level your '
                    'hiring plan should target.'
                    '</div>',
                    unsafe_allow_html=True,
                )

            # Drop per-role columns for cleaner leadership view
            _role_cols = [c for c in ann_disp.columns if c.startswith("Avg ") and c != "Avg monthly FTE"]
            ann_disp = ann_disp.drop(columns=_role_cols)

            # Round all numeric columns to whole numbers for readability
            _num_cols = ann_disp.select_dtypes(include="number").columns
            ann_disp[_num_cols] = ann_disp[_num_cols].round(0).astype(int)

            if has_contingency:
                _avg_col_idx = ann_disp.columns.get_loc("Avg monthly FTE") + 1
                _cont_label = f"Avg + {cont*100:.0f}% contingency"
                ann_disp.insert(_avg_col_idx, _cont_label, (ann_disp["Avg monthly FTE"] * (1 + cont)).round(0).astype(int))

            # YoY delta column
            ann_disp["YoY change"] = ann_disp["Avg monthly FTE"].diff().fillna(0).astype(int)

            if has_cost_range:
                low_ann = result.cost_low_annual
                high_ann = result.cost_high_annual
                if not low_ann.empty and "Avg monthly FTE" in low_ann.columns:
                    ann_disp["FTE (low cost)"] = low_ann["Avg monthly FTE"].values
                if not high_ann.empty and "Avg monthly FTE" in high_ann.columns:
                    ann_disp["FTE (high cost)"] = high_ann["Avg monthly FTE"].values

            def _highlight_final_year(row):
                if row["Year"] == _last_yr:
                    return ["background-color: #E8F0FE; font-weight: 700"] * len(row)
                return [""] * len(row)
            st.dataframe(
                ann_disp.style.apply(_highlight_final_year, axis=1),
                use_container_width=True, hide_index=True,
            )
            if is_cashflow:
                st.caption(
                    "Under cash-flow budgeting, project counts oscillate because each year's budget "
                    "must first cover ongoing costs from prior cohorts before funding new starts."
                )

            with st.expander("Breakdown by role — average monthly FTE per year"):
                _role_breakdown = result.annual_summary[["Year"]].copy()
                for role in cfg.all_roles:
                    _src_col = f"Avg {role} FTE"
                    if _src_col in result.annual_summary.columns:
                        _role_breakdown[f"Avg monthly {role} FTE"] = result.annual_summary[_src_col]
                _role_breakdown["Avg monthly FTE (all roles)"] = result.annual_summary["Avg monthly FTE"]
                _rb_num = _role_breakdown.select_dtypes(include="number").columns
                _role_breakdown[_rb_num] = _role_breakdown[_rb_num].round(0).astype(int)

                def _highlight_final_yr_role(row):
                    if row["Year"] == _last_yr:
                        return ["background-color: #E8F0FE; font-weight: 700"] * len(row)
                    return [""] * len(row)
                st.dataframe(
                    _role_breakdown.style.apply(_highlight_final_yr_role, axis=1),
                    use_container_width=True, hide_index=True,
                )

    # ── Assumption Register ──
    with tab_assumptions:
        st.markdown("#### All assumptions used in this run")
        st.markdown('<div class="section-intro">Every number in the model traces back to one of these inputs.</div>', unsafe_allow_html=True)

        _budget_mode_label = "Cash-flow" if is_cashflow else "Commitment"
        _budget_mode_meaning = (
            "Annual budget covers all active projects (ongoing + new)"
            if is_cashflow
            else "Annual budget funds full lifecycle cost of new projects upfront"
        )
        st.markdown('<div class="card"><h5>Budget & timeline</h5>', unsafe_allow_html=True)
        st.dataframe(pd.DataFrame([
            {"Assumption": "Budget model", "Value": _budget_mode_label, "Type": "Input", "Meaning": _budget_mode_meaning},
            {"Assumption": "Total R&D budget", "Value": f"{cfg.total_budget_m:,.0f} M", "Type": "Input", "Meaning": "Gross annual R&D spend"},
            {"Assumption": "Overhead", "Value": f"{cfg.overhead_pct*100:.0f}%", "Type": "Input", "Meaning": "Admin, facilities, management"},
            {"Assumption": "Net project budget", "Value": f"{cfg.total_budget_m*(1-cfg.overhead_pct):,.0f} M", "Type": "Derived", "Meaning": "Total x (1 - Overhead)"},
            {"Assumption": "Start year", "Value": str(cfg.start_year), "Type": "Input", "Meaning": "First year of project intake"},
            {"Assumption": "End year", "Value": str(cfg.end_year), "Type": "Input", "Meaning": "Last year of project intake"},
            {"Assumption": "Intake window", "Value": f"{cfg.intake_spread_months} months", "Type": "Input", "Meaning": "New projects spread evenly across this many months"},
        ]), use_container_width=True, hide_index=True)
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="card"><h5>Pipeline & funnel</h5>', unsafe_allow_html=True)
        pipe_rows = []
        for si, sn in enumerate(cfg.pipeline_stages):
            row_data: dict[str, str] = {
                "Stage": sn,
                "Allocation" if not has_phase2 else f"Phase 1 ({cfg.start_year}–{cfg.phase2_start_year - 1})": f"{cfg.stage_mix.get(sn,0)*100:.0f}%",
                "Conversion to next (all phases)": f"{cfg.stage_conversion_rates.get(sn,0)*100:.0f}%" if si < len(cfg.pipeline_stages) - 1 else "Terminal",
            }
            if has_phase2:
                row_data[f"Phase 2 ({cfg.phase2_start_year}–{cfg.end_year})"] = f"{cfg.stage_mix_phase2.get(sn,0)*100:.0f}%"
            pipe_rows.append(row_data)
        st.dataframe(pd.DataFrame(pipe_rows), use_container_width=True, hide_index=True)
        if has_phase2:
            st.caption("Conversion rates (% move to next) stay the same across both phases — only the allocation (% start here) shifts.")
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="card"><h5>Advanced settings</h5>', unsafe_allow_html=True)
        advanced_rows = [
            {"Assumption": "Utilization rate", "Value": f"{cfg.utilization_rate*100:.0f}%", "Type": "Input",
             "Meaning": f"Gross FTE = model FTE ÷ {cfg.utilization_rate:.2f}"},
            {"Assumption": "Ramp-up period", "Value": f"{cfg.ramp_months} months", "Type": "Input",
             "Meaning": "Linear ramp from 0 to full FTE" if cfg.ramp_months > 0 else "Full FTE from day 1"},
        ]
        advanced_rows.append(
            {"Assumption": "Contingency buffer", "Value": f"{cfg.contingency_pct*100:.0f}%", "Type": "Input",
             "Meaning": f"Adjusted FTE = Base × (1 + {cfg.contingency_pct*100:.0f}%)" if has_contingency else "No buffer applied"}
        )
        st.dataframe(pd.DataFrame(advanced_rows), use_container_width=True, hide_index=True)
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="card"><h5>Project type parameters</h5>', unsafe_allow_html=True)
        arch_rows = []
        for arch in cfg.archetypes:
            for sn, sp in arch.stages.items():
                row_data = {
                    "Archetype": arch.name, "Stage": sn,
                    "Share": f"{arch.portfolio_share*100:.0f}%",
                    "Duration (mo)": f"{sp.duration_months}",
                    "Cost Min (M)": f"{sp.cost_min:.1f}",
                    "Cost Max (M)": f"{sp.cost_max:.1f}",
                }
                for role in sp.fte_per_role:
                    row_data[f"{role} FTE"] = f"{sp.fte_per_role[role]:.1f}"
                arch_rows.append(row_data)
        st.dataframe(pd.DataFrame(arch_rows), use_container_width=True, hide_index=True)
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="card"><h5>Derived outputs</h5>', unsafe_allow_html=True)
        try:
            wc = _weighted_cost_per_project(cfg)
        except Exception:
            wc = 0
        derived_rows = [
            {"Metric": "Weighted lifecycle cost per project",
             "Value": f"{wc:,.1f} M",
             "How": "Portfolio-weighted expected cost across all stages"},
            {"Metric": f"Avg FTE in {cfg.end_year}",
             "Value": f"{result.steady_state_avg:,.0f}",
             "How": f"Average monthly total FTE in {cfg.end_year}"},
            {"Metric": f"FTE range in {cfg.end_year}",
             "Value": f"{result.steady_state_min_month:,.0f} – {result.steady_state_max_month:,.0f}",
             "How": f"Min to max monthly FTE in {cfg.end_year}"},
        ]
        if is_cashflow:
            _total_new = sum(result.yearly_projects.get(y, 0) for y in range(cfg.start_year, cfg.end_year + 1))
            peak_fte = monthly.groupby(monthly["month"].dt.to_period("M"))["fte_total"].sum().max() if not monthly.empty else 0
            derived_rows.insert(1, {
                "Metric": "Total new projects (all years)",
                "Value": f"{_total_new:,.0f}",
                "How": "Sum of new project starts across all years",
            })
            derived_rows.append({
                "Metric": "Peak monthly FTE (any month)",
                "Value": f"{peak_fte:,.0f}",
                "How": "Highest single-month total FTE across all years",
            })
        else:
            _ppy = result.projects_per_year
            derived_rows.insert(1, {
                "Metric": "New projects per year",
                "Value": f"{_ppy:,.0f}",
                "How": "Net budget ÷ weighted lifecycle cost",
            })
        st.dataframe(pd.DataFrame(derived_rows), use_container_width=True, hide_index=True)
        st.markdown('</div>', unsafe_allow_html=True)

        _role_list_str = ", ".join(cfg.all_roles)
        with st.expander("Structural assumptions"):
            _structural = [
                {"Assumption": "Constant annual budget", "Meaning": "Same gross budget every year"},
                {"Assumption": "No inflation or cost escalation", "Meaning": "Project costs stay the same regardless of which year they start"},
                {"Assumption": "No economies of scale", "Meaning": "Cost per project is constant regardless of volume"},
            ]
            if is_cashflow:
                _structural += [
                    {"Assumption": "Cash-flow budget constraint",
                     "Meaning": "Each year's budget must cover ongoing costs first; remainder funds new projects"},
                    {"Assumption": "Year-to-year carry-over",
                     "Meaning": "Ongoing project costs from prior years reduce the budget available for new starts"},
                ]
            else:
                _structural.append(
                    {"Assumption": "Independent annual cohorts",
                     "Meaning": "No carry-over of costs between years; each year's budget funds new projects only"}
                )
            _structural += [
                {"Assumption": "Uniform intake spread", "Meaning": "Even distribution across intake window"},
                {"Assumption": "All types share same stages", "Meaning": "You can't have different stage sequences for different archetypes"},
                {"Assumption": "Linear pipeline", "Meaning": "No branching or looping between stages"},
                {"Assumption": "Additive conversion", "Meaning": "Stage survivors add to next-stage direct allocation"},
                {"Assumption": "Same conversion rates for all types", "Meaning": "The % move to next rates are shared across all archetypes"},
                {"Assumption": "No mid-stage failure", "Meaning": "Projects run to completion once started"},
                {"Assumption": "Constant FTE per project", "Meaning": "Flat staffing for full duration (unless ramp set)"},
                {"Assumption": "No resource constraints", "Meaning": "The model doesn't cap FTE supply — assumes you can always hire enough"},
                {"Assumption": "Monthly granularity", "Meaning": "Projects and FTE tracked monthly"},
            ]
            st.dataframe(pd.DataFrame(_structural), use_container_width=True, hide_index=True)

    # ── Single-scenario Excel download ──
    st.divider()
    dc1, dc2, dc3 = st.columns([1, 2, 1])
    with dc2:
        st.markdown(
            '<div class="section-intro" style="text-align:center;">Export this scenario for offline review or sharing.</div>',
            unsafe_allow_html=True,
        )
        excel_bytes = _generate_excel(cfg, result)
        st.download_button(
            "Download Excel — assumptions, annual summary & monthly detail",
            data=excel_bytes,
            file_name=f"FTE_Model_{name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
            key=f"{P}dl_excel",
            type="primary",
        )


# ═══════════════════════════════════════════════════════════════════════════
# COMPARE VIEW
# ═══════════════════════════════════════════════════════════════════════════
def _render_compare_view(results_list):
    # Compute min/max avg FTE across scenarios for insights
    _n = len(results_list)
    _end_yr = results_list[0][1].end_year if results_list else 0
    _avgs = {sname: res.steady_state_avg for sname, cfg, res in results_list}
    _min_name = min(_avgs, key=_avgs.get)
    _max_name = max(_avgs, key=_avgs.get)
    _min_val = _avgs[_min_name]
    _max_val = _avgs[_max_name]
    _spread = _max_val - _min_val

    st.markdown(f"#### Comparing {_n} scenarios — average FTE in {_end_yr} ranges from {_min_val:,.0f} to {_max_val:,.0f}")
    st.markdown(
        '<div class="section-intro">Side-by-side view of all scenarios. '
        'The table summarizes key metrics; the chart shows how headcount evolves over time.</div>',
        unsafe_allow_html=True,
    )

    summary_df = comparison_summary(results_list)
    _preferred_order = [
        "Scenario", "Avg FTE (last yr)", "Peak FTE", "Budget (M)", "Overhead",
        "Net Budget (M)", "Avg Projects/yr", "Portfolio Split", "Success Rate",
        "FTE Range (cost)",
    ]
    _reordered = [c for c in _preferred_order if c in summary_df.columns]
    _remaining = [c for c in summary_df.columns if c not in _reordered]
    summary_df = summary_df[_reordered + _remaining]
    st.dataframe(summary_df, use_container_width=True, hide_index=True)

    # Chart
    st.markdown(f"#### Headcount varies by {_spread:,.0f} FTE across scenarios")

    fig = go.Figure()
    for i, (sname, cfg, res) in enumerate(results_list):
        ann = res.annual_summary
        if ann.empty:
            continue
        color = SCENARIO_COLORS[i % len(SCENARIO_COLORS)]
        fig.add_trace(go.Bar(
            x=ann["Year"], y=ann["Avg monthly FTE"],
            name=sname, marker_color=color, opacity=0.85,
        ))

    fig.update_layout(
        barmode="group", height=420,
        margin=dict(l=20, r=20, t=30, b=20),
        plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Inter, Helvetica Neue, Arial, sans-serif", size=12, color=MCK_DARK),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0),
        xaxis=dict(gridcolor="#E0E0E0", title="", dtick=1),
        yaxis=dict(gridcolor="#E0E0E0", title="Avg Monthly FTE"),
        hovermode="x unified",
    )
    st.plotly_chart(fig, width="stretch", key="compare_chart")

    # Callout
    st.markdown(
        f'<div class="section-intro">'
        f'Leanest: <strong>{_min_name}</strong> ({_min_val:,.0f} FTE). '
        f'Heaviest: <strong>{_max_name}</strong> ({_max_val:,.0f} FTE). '
        f'Spread: <strong>{_spread:,.0f} FTE</strong>.</div>',
        unsafe_allow_html=True,
    )

    # Download comparison Excel
    st.divider()
    dc1, dc2, dc3 = st.columns([1, 2, 1])
    with dc2:
        st.markdown(
            '<div class="section-intro" style="text-align:center;">Export all scenarios for offline comparison.</div>',
            unsafe_allow_html=True,
        )
        excel_bytes = generate_comparison_excel(results_list)
        st.download_button(
            "Download Excel — all scenarios compared",
            data=excel_bytes,
            file_name="FTE_Scenario_Comparison.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
            key="dl_compare_excel",
        )


# ═══════════════════════════════════════════════════════════════════════════
# PAGE 1: CONFIGURE
# ═══════════════════════════════════════════════════════════════════════════
def _page_configure():
    _render_header(
        "Set your R&amp;D budget and project portfolio "
        "&mdash; the model calculates how many staff you need"
    )

    # ── Configuration sub-header ──
    st.markdown("""
    <div class="config-header">
        <h2>Model Configuration</h2>
        <p>Define your R&amp;D budget, project portfolio, and staffing assumptions below.
        The model will calculate the workforce demand across your planning horizon.</p>
    </div>
    """, unsafe_allow_html=True)

    # ── Flow diagram (generic, shared across all scenarios) ──
    st.markdown("""
    <div class="flow-row">
        <div class="flow-box">
            <div class="flow-num">1</div>
            <div class="flow-title">Net R&amp;D Budget</div>
            <div class="flow-formula">Budget &minus; Overhead</div>
            <div class="flow-desc">Funding available for projects</div>
        </div>
        <div class="flow-arrow">&rarr;</div>
        <div class="flow-box">
            <div class="flow-num">2</div>
            <div class="flow-title">New Projects / Year</div>
            <div class="flow-desc">How many projects the budget can fund</div>
        </div>
        <div class="flow-arrow">&rarr;</div>
        <div class="flow-box">
            <div class="flow-num">3</div>
            <div class="flow-title">Pipeline Stages</div>
            <div class="flow-desc">Projects enter at different stages; some advance through gates</div>
        </div>
        <div class="flow-arrow">&rarr;</div>
        <div class="flow-box">
            <div class="flow-num">4</div>
            <div class="flow-title">Active Projects</div>
            <div class="flow-desc">Total projects running at any point, including carry-over</div>
        </div>
        <div class="flow-arrow">&rarr;</div>
        <div class="flow-box">
            <div class="flow-num">5</div>
            <div class="flow-title">FTE Demand</div>
            <div class="flow-formula">Active Projects &times; Team Size</div>
            <div class="flow-desc">Staff needed per workforce role</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Scenario tabs ──
    scenario_names = [s["name"] for s in st.session_state.scenarios]
    tab_labels = scenario_names + ["＋ Add Scenario"]
    tabs = st.tabs(tab_labels)

    for idx in range(len(st.session_state.scenarios)):
        with tabs[idx]:
            _render_scenario_form(idx)

            if len(st.session_state.scenarios) > 1:
                if st.button("Remove this scenario", key=f"remove_scen_{idx}"):
                    st.session_state.scenarios.pop(idx)
                    _clear_scenario_keys(idx)
                    st.rerun()

    with tabs[-1]:
        st.markdown("Click below to add a new scenario.")
        if st.button("Create new scenario", key="add_scenario"):
            n = len(st.session_state.scenarios) + 1
            st.session_state.scenarios.append({
                "name": f"Scenario {n}",
                "cfg": default_baseline(),
            })
            st.rerun()

    # ── Fixed assumptions (non-configurable) ──
    with st.expander("Things the model assumes that cannot be changed"):
        st.markdown("""
| What the model assumes | In plain language |
|---|---|
| Same gross budget every year | The same annual budget envelope applies every year |
| No inflation or cost escalation | Project costs stay the same regardless of which year they start |
| Linear cost burn | Project cost is spread evenly across its duration (cost ÷ months) |
| No bulk discounts | Running 50 projects doesn't make each one cheaper than running 5 |
| Projects start evenly | New projects are spread evenly across the intake months (not all at once) |
| All project types share the same stages | You can't have different stage sequences for different archetypes |
| One path through the pipeline | Projects go forward through stages in order — no skipping ahead or looping back |
| Graduates add to the next stage | Projects that finish an early stage and move on are *added* to whatever's already in the next stage |
| Same conversion rates for all project types | The "% move to next" rates are shared across all archetypes, not set per type |
| No projects get cancelled midway | Once started, every project runs to the end of its stage |
| Same team size throughout a project | Once a project starts, its team stays the same size (unless ramp-up is on) |
| No resource constraints | The model assumes you can always hire enough people — it doesn't cap FTE supply |
| Monthly tracking | The model tracks projects and headcount month by month |
""")

    # ── Generate button ──
    c1, c2, c3 = st.columns([1, 2, 1])
    with c2:
        if st.button("▶  Calculate Headcount", width="stretch", type="primary"):
            configs = [(s["name"], s["cfg"]) for s in st.session_state.scenarios]
            st.session_state.scenario_results = run_all(configs)
            st.session_state.page = "results"
            st.rerun()

    _, rc, _ = st.columns([1, 2, 1])
    with rc:
        if st.button("⟳  Reset all to defaults", width="stretch"):
            st.session_state.scenarios = [{"name": "Scenario 1", "cfg": default_baseline()}]
            st.session_state.scenario_results = []
            _clear_scenario_keys(0)
            st.rerun()


# ═══════════════════════════════════════════════════════════════════════════
# PAGE 2: RESULTS
# ═══════════════════════════════════════════════════════════════════════════
def _page_results():
    results_list = st.session_state.scenario_results
    if not results_list:
        st.warning("No results. Please configure and generate first.")
        if st.button("← Back to Configure"):
            st.session_state.page = "configure"
            st.rerun()
        return

    first_cfg = results_list[0][1]
    _render_header(
        f"Results for {first_cfg.start_year}–{first_cfg.end_year} "
        f"&nbsp;|&nbsp; {len(results_list)} scenario(s)"
    )

    hc1, hc2 = st.columns([3, 1])
    with hc2:
        if st.button("← Modify Assumptions", width="stretch"):
            st.session_state.page = "configure"
            st.rerun()

    # Build tab labels
    has_multiple = len(results_list) > 1
    tab_labels = [name for name, _, _ in results_list]
    if has_multiple:
        tab_labels.append("Compare All")

    tabs = st.tabs(tab_labels)

    for i, (sname, cfg, result) in enumerate(results_list):
        with tabs[i]:
            _render_single_result(sname, cfg, result, key_prefix=f"r{i}")

    if has_multiple:
        with tabs[-1]:
            _render_compare_view(results_list)


# ═══════════════════════════════════════════════════════════════════════════
# Excel generator (single scenario — kept from original)
# ═══════════════════════════════════════════════════════════════════════════
def _generate_excel(cfg: ModelConfig, result) -> bytes:
    has_phase2 = cfg.phase2_start_year > 0 and bool(cfg.stage_mix_phase2)
    yp = result.yearly_projects
    if yp:
        yr_parts = [f"{yp.get(y, 0):,.0f}" for y in range(cfg.start_year, cfg.end_year + 1)]
        _proj_kpi_value = " / ".join(yr_parts) + f" ({cfg.start_year}\u2013{cfg.end_year})"
    else:
        _proj_kpi_value = f"{result.projects_per_year:,.0f}"

    wb = Workbook()

    navy_fill = PatternFill(start_color="051C2C", end_color="051C2C", fill_type="solid")
    light_fill = PatternFill(start_color="F5F6F7", end_color="F5F6F7", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    hdr_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=16, bold=True, color="051C2C")
    sub_font = Font(name="Calibri", size=11, color="7F8C8D")
    sec_font = Font(name="Calibri", size=12, bold=True, color="051C2C")
    body_font = Font(name="Calibri", size=10, color="1A1A2E")
    blue_font = Font(name="Calibri", size=10, color="2251FF", bold=True)
    bdr = Border(
        left=Side(style="thin", color="D0D5DD"), right=Side(style="thin", color="D0D5DD"),
        top=Side(style="thin", color="D0D5DD"), bottom=Side(style="thin", color="D0D5DD"),
    )

    def _hdr_row(ws, row, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = navy_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = bdr

    def _data_row(ws, row, ncol, alt=False):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = light_fill if alt else white_fill
            cell.font = body_font
            cell.alignment = Alignment(vertical="center")
            cell.border = bdr

    ws = wb.active
    ws.title = "Cover"
    ws.sheet_properties.tabColor = "00A9F4"
    ws.merge_cells("B3:F3")
    ws["B3"] = "FTE Baseload Model"
    ws["B3"].font = Font(name="Calibri", size=22, bold=True, color="051C2C")
    ws.merge_cells("B5:F5")
    ws["B5"] = "R&D Workforce Demand Planning"
    ws["B5"].font = sub_font
    _mode_label = "Cash-flow" if cfg.budget_mode == "cashflow" else "Commitment"
    info = [
        ("B8", "Budget", f"{cfg.total_budget_m:,.0f} M total, {cfg.total_budget_m*(1-cfg.overhead_pct):,.0f} M net"),
        ("B9", "Budget model", _mode_label),
        ("B10", "Period", f"{cfg.start_year}\u2013{cfg.end_year}"),
        ("B11", "Archetypes", ", ".join(a.name for a in cfg.archetypes)),
        ("B12", "Avg monthly FTE", f"{result.steady_state_avg:,.0f}"),
        ("B13", "FTE range", f"{result.steady_state_min_month:,.0f} \u2013 {result.steady_state_max_month:,.0f}"),
        ("B14", "Projects/year", _proj_kpi_value),
    ]
    _cover_row = 15
    if cfg.contingency_pct > 0:
        _adj_avg = result.steady_state_avg * (1 + cfg.contingency_pct)
        info.append(("B15", f"Avg + {cfg.contingency_pct*100:.0f}% contingency", f"{_adj_avg:,.0f}"))
        _cover_row = 16
    for ref, lbl, val in info:
        ws[ref] = lbl
        ws[ref].font = Font(name="Calibri", size=10, bold=True, color="051C2C")
        ws[ref.replace("B", "C")].value = val
        ws[ref.replace("B", "C")].font = body_font

    # Key takeaways
    _cover_row += 1
    ws[f"B{_cover_row}"] = "Key takeaways"
    ws[f"B{_cover_row}"].font = sec_font
    _cover_row += 1
    _takeaways = []
    if cfg.archetypes:
        _top_arch = max(cfg.archetypes, key=lambda a: a.portfolio_share)
        _takeaways.append(
            f"{_top_arch.name} projects dominate the portfolio at "
            f"{_top_arch.portfolio_share*100:.0f}% share."
        )
    _all_r = cfg.all_roles
    if _all_r and not result.annual_summary.empty:
        _rt = {}
        for r in _all_r:
            _rc = f"Avg {r} FTE"
            if _rc in result.annual_summary.columns:
                _rt[r] = result.annual_summary[_rc].iloc[-1]
        if _rt:
            _tr = max(_rt, key=_rt.get)
            _tp = (_rt[_tr] / sum(_rt.values()) * 100) if sum(_rt.values()) > 0 else 0
            _takeaways.append(f"The {_tr} role makes up {_tp:.0f}% of total FTE demand in {cfg.end_year}.")
    if not result.annual_summary.empty and len(result.annual_summary) >= 2:
        _y1 = result.annual_summary["Avg monthly FTE"].iloc[0]
        _y3 = result.annual_summary["Avg monthly FTE"].iloc[min(2, len(result.annual_summary) - 1)]
        if _y1 > 0:
            _gr = ((_y3 / _y1) - 1) * 100
            _takeaways.append(
                f"FTE grows {_gr:+.0f}% in the first {min(3, len(result.annual_summary))} years "
                f"as the pipeline fills."
            )
    for t in _takeaways:
        ws.merge_cells(f"B{_cover_row}:F{_cover_row}")
        ws[f"B{_cover_row}"] = f"\u2022  {t}"
        ws[f"B{_cover_row}"].font = body_font
        _cover_row += 1

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 60

    # Assumptions
    ws_a = wb.create_sheet("Assumptions")
    ws_a.sheet_properties.tabColor = "051C2C"
    ws_a["B2"] = "Model Assumptions"
    ws_a["B2"].font = title_font
    ws_a["B3"] = "Blue = user input. Black = derived."
    ws_a["B3"].font = sub_font

    row = 5
    sections = [
        ("BUDGET", [
            ("Total R&D budget", f"{cfg.total_budget_m:,.0f} M", True),
            ("Overhead", f"{cfg.overhead_pct*100:.0f}%", True),
            ("Net project budget", f"{cfg.total_budget_m*(1-cfg.overhead_pct):,.0f} M", False),
            ("Budget model", _mode_label, True),
        ]),
        ("TIMELINE", [
            ("Start year", str(cfg.start_year), True),
            ("End year", str(cfg.end_year), True),
            ("Intake window", f"{cfg.intake_spread_months} months", True),
        ]),
        ("PIPELINE", [(sn, f"{cfg.stage_mix.get(sn,0)*100:.0f}% alloc, "
              f"{cfg.stage_conversion_rates.get(sn,0)*100:.0f}% conv"
              if si < len(cfg.pipeline_stages)-1 else
              f"{cfg.stage_mix.get(sn,0)*100:.0f}% alloc (terminal)", True)
         for si, sn in enumerate(cfg.pipeline_stages)]),
        ("ADVANCED", [
            ("Utilization", f"{cfg.utilization_rate*100:.0f}%", True),
            ("Ramp-up", f"{cfg.ramp_months} months", True),
            ("Contingency buffer", f"{cfg.contingency_pct*100:.0f}%", True),
        ]),
    ]
    for sec_name, items in sections:
        ws_a[f"B{row}"] = sec_name
        ws_a[f"B{row}"].font = sec_font
        row += 1
        for ci, h in enumerate(["Assumption", "Value", "Type"], 2):
            ws_a.cell(row=row, column=ci, value=h)
        _hdr_row(ws_a, row, 4)
        row += 1
        for lbl, val, is_input in items:
            ws_a.cell(row=row, column=2, value=lbl).font = body_font
            ws_a.cell(row=row, column=3, value=val).font = blue_font if is_input else body_font
            ws_a.cell(row=row, column=4, value="Input" if is_input else "Derived").font = body_font
            _data_row(ws_a, row, 4, alt=(row % 2 == 0))
            row += 1
        row += 1

    ws_a[f"B{row}"] = "ARCHETYPE PARAMETERS"
    ws_a[f"B{row}"].font = sec_font
    row += 1
    all_roles = cfg.all_roles
    arch_hdrs = ["Archetype", "Stage", "Share", "Duration", "Cost Min", "Cost Max"]
    arch_hdrs += [f"{r} FTE" for r in all_roles]
    for ci, h in enumerate(arch_hdrs, 2):
        ws_a.cell(row=row, column=ci, value=h)
    ncol_arch = len(arch_hdrs) + 1
    _hdr_row(ws_a, row, ncol_arch)
    row += 1
    for arch in cfg.archetypes:
        for sn, sp in arch.stages.items():
            ws_a.cell(row=row, column=2, value=arch.name).font = body_font
            ws_a.cell(row=row, column=3, value=sn).font = body_font
            ws_a.cell(row=row, column=4, value=f"{arch.portfolio_share*100:.0f}%").font = blue_font
            ws_a.cell(row=row, column=5, value=f"{sp.duration_months} mo").font = blue_font
            ws_a.cell(row=row, column=6, value=f"{sp.cost_min:.1f} M").font = blue_font
            ws_a.cell(row=row, column=7, value=f"{sp.cost_max:.1f} M").font = blue_font
            for ri, role in enumerate(all_roles):
                ws_a.cell(row=row, column=8 + ri,
                          value=f"{sp.fte_per_role.get(role, 0):.1f}").font = blue_font
            _data_row(ws_a, row, ncol_arch, alt=(row % 2 == 0))
            row += 1

    for ci in range(1, ncol_arch + 1):
        ws_a.column_dimensions[get_column_letter(ci)].width = 20
    ws_a.column_dimensions["A"].width = 3

    # Annual Summary (match UI: drop role cols, add YoY/contingency/cost range, whole numbers)
    ws_ann = wb.create_sheet("Annual Summary")
    ws_ann.sheet_properties.tabColor = "051C2C"
    ws_ann["B2"] = "Annual FTE Summary"
    ws_ann["B2"].font = title_font
    ann_df = result.annual_summary.copy()
    if not ann_df.empty:
        _has_cont = cfg.contingency_pct > 0
        _has_cr = (
            isinstance(result.cost_low_annual, pd.DataFrame)
            and not result.cost_low_annual.empty
        )
        _role_drop = [c for c in ann_df.columns if c.startswith("Avg ") and c != "Avg monthly FTE"]
        ann_df = ann_df.drop(columns=_role_drop)
        _num_c = ann_df.select_dtypes(include="number").columns
        ann_df[_num_c] = ann_df[_num_c].round(0).astype(int)
        if _has_cont:
            _avg_idx = list(ann_df.columns).index("Avg monthly FTE") + 1
            _cl = f"Avg + {cfg.contingency_pct*100:.0f}% contingency"
            ann_df.insert(_avg_idx, _cl, (ann_df["Avg monthly FTE"] * (1 + cfg.contingency_pct)).round(0).astype(int))
        ann_df["YoY change"] = ann_df["Avg monthly FTE"].diff().fillna(0).astype(int)
        if _has_cr:
            low_a = result.cost_low_annual
            high_a = result.cost_high_annual
            if not low_a.empty and "Avg monthly FTE" in low_a.columns:
                ann_df["FTE (low cost)"] = low_a["Avg monthly FTE"].round(0).astype(int).values
            if not high_a.empty and "Avg monthly FTE" in high_a.columns:
                ann_df["FTE (high cost)"] = high_a["Avg monthly FTE"].round(0).astype(int).values
        ann_df.insert(1, "New projects", ann_df["Year"].map(lambda y: round(yp.get(y, 0))))
        row = 4
        cols = list(ann_df.columns)
        for ci, h in enumerate(cols, 2):
            ws_ann.cell(row=row, column=ci, value=h)
        _hdr_row(ws_ann, row, len(cols) + 1)
        row += 1
        _last_yr_xl = int(ann_df["Year"].iloc[-1])
        for _, dr in ann_df.iterrows():
            for ci, col in enumerate(cols, 2):
                ws_ann.cell(row=row, column=ci, value=dr[col]).font = body_font
            is_final = dr["Year"] == _last_yr_xl
            if is_final:
                for ci in range(2, len(cols) + 2):
                    cell = ws_ann.cell(row=row, column=ci)
                    cell.fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
                    cell.font = Font(name="Calibri", size=10, bold=True, color="1A1A2E")
                    cell.border = bdr
            else:
                _data_row(ws_ann, row, len(cols) + 1, alt=(row % 2 == 0))
            row += 1
    ws_ann.column_dimensions["A"].width = 3
    for ci in range(2, len(ann_df.columns) + 2 if not ann_df.empty else 12):
        ws_ann.column_dimensions[get_column_letter(ci)].width = 22

    # Monthly Detail
    ws_m = wb.create_sheet("Monthly Detail")
    ws_m.sheet_properties.tabColor = "7F8C8D"
    ws_m["B2"] = "Monthly FTE Detail"
    ws_m["B2"].font = title_font
    mon = result.monthly.copy()
    if not mon.empty:
        mon["month_str"] = mon["month"].dt.strftime("%Y-%m")
        exp_cols = ["month_str", "archetype", "stage", "effective_projects"]
        nice = ["Month", "Archetype", "Stage", "Effective Projects"]
        for role in cfg.all_roles:
            col = f"fte_{role}"
            if col in mon.columns:
                exp_cols.append(col)
                nice.append(f"{role} FTE")
        exp_cols.append("fte_total")
        nice.append("Total FTE")
        for c in exp_cols:
            if c != "month_str" and c != "archetype" and c != "stage" and c in mon.columns:
                mon[c] = mon[c].round(0).astype(int)
        row = 4
        for ci, h in enumerate(nice, 2):
            ws_m.cell(row=row, column=ci, value=h)
        _hdr_row(ws_m, row, len(nice) + 1)
        row += 1
        for _, dr in mon[exp_cols].iterrows():
            for ci, col in enumerate(exp_cols, 2):
                ws_m.cell(row=row, column=ci, value=dr[col]).font = body_font
            _data_row(ws_m, row, len(nice) + 1, alt=(row % 2 == 0))
            row += 1
    ws_m.column_dimensions["A"].width = 3
    for ci in range(2, 13):
        ws_m.column_dimensions[get_column_letter(ci)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════
# Router
# ═══════════════════════════════════════════════════════════════════════════
if st.session_state.page == "configure":
    _page_configure()
else:
    _page_results()

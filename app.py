"""
FTE Baseload Planning Tool — Streamlit UI
Two-page flow: Configure assumptions → View results.
Standard / Custom input modes. No sidebar.
"""

import io
import sys
import copy
from pathlib import Path

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))

from config import Archetype, ModelConfig, StageParams
from defaults import petronas_baseline
from model import run_model, _weighted_cost_per_project

# ---------------------------------------------------------------------------
# Palette
# ---------------------------------------------------------------------------
MCK_NAVY = "#051C2C"
MCK_BLUE = "#2251FF"
MCK_TEAL = "#00A9F4"
MCK_GREEN = "#00B140"
MCK_GREY = "#7F8C8D"
MCK_LIGHT = "#F5F6F7"
MCK_WHITE = "#FFFFFF"
MCK_DARK = "#1A1A2E"

ARCH_COLORS = [
    "#2251FF", "#00A9F4", "#00B140", "#F4A100", "#E74C3C",
    "#8E44AD", "#1ABC9C", "#D35400", "#2C3E50", "#27AE60",
]


def _rgba(hex_color: str, opacity: float) -> str:
    h = hex_color.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return f"rgba({r},{g},{b},{opacity})"


# ---------------------------------------------------------------------------
# Page config & CSS
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="FTE Baseload Planner",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown(f"""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

    :root {{ --primary-color: {MCK_NAVY}; }}

    .stApp {{ font-family: 'Inter', 'Helvetica Neue', Arial, sans-serif; }}
    .main .block-container {{ padding-top: 1.5rem; max-width: 1200px; }}

    /* Hide Streamlit chrome: sidebar, status, deploy, menu */
    [data-testid="collapsedControl"] {{ display: none; }}
    [data-testid="stStatusWidget"] {{ display: none; }}
    [data-testid="stToolbar"] {{ display: none; }}
    header[data-testid="stHeader"] {{ display: none; }}

    /* Override Streamlit accent → navy */
    .stTabs [data-baseweb="tab-highlight"] {{ background-color: {MCK_NAVY} !important; }}
    .stTabs [data-baseweb="tab"] {{ color: {MCK_GREY}; }}
    .stTabs [aria-selected="true"] {{ color: {MCK_NAVY} !important; }}
    button[kind="primary"], .stDownloadButton button {{
        background-color: {MCK_NAVY} !important; border-color: {MCK_NAVY} !important;
        color: {MCK_WHITE} !important;
    }}
    button[kind="primary"] *, .stDownloadButton button * {{
        color: {MCK_WHITE} !important;
    }}
    .stSlider [data-baseweb="slider"] div[role="slider"] {{ background: {MCK_NAVY} !important; }}
    a {{ color: {MCK_NAVY}; }}

    .mck-header {{
        background: {MCK_NAVY}; color: white;
        padding: 1.6rem 2rem; border-radius: 8px; margin-bottom: 1.2rem;
    }}
    .mck-header h1 {{ margin: 0; font-size: 1.5rem; font-weight: 600; letter-spacing: -0.02em; }}
    .mck-header p {{ margin: 0.3rem 0 0 0; font-size: 0.82rem; opacity: 0.7; }}

    .kpi-row {{ display: flex; gap: 1rem; margin-bottom: 1.5rem; }}
    .kpi-card {{
        flex: 1; background: {MCK_WHITE}; border: 1px solid #E0E4E8;
        border-radius: 8px; padding: 1.1rem 1.4rem; box-shadow: 0 1px 3px rgba(0,0,0,0.04);
    }}
    .kpi-card .kpi-label {{
        font-size: 0.7rem; font-weight: 500; color: {MCK_GREY};
        text-transform: uppercase; letter-spacing: 0.05em; margin-bottom: 0.3rem;
    }}
    .kpi-card .kpi-value {{ font-size: 1.5rem; font-weight: 700; color: {MCK_NAVY}; }}
    .kpi-card .kpi-sub {{ font-size: 0.75rem; color: {MCK_GREY}; margin-top: 0.15rem; }}

    .card {{
        background: transparent; border: none;
        border-radius: 0; padding: 0 0 1rem 0; margin-bottom: 1rem;
    }}
    .card h5 {{
        background: #EBF4FA; color: {MCK_NAVY}; font-size: 0.78rem; font-weight: 600;
        text-transform: uppercase; letter-spacing: 0.03em;
        margin: 0 0 1rem 0; padding: 0.65rem 1rem;
        border-radius: 6px; border-bottom: none;
    }}

    .help-text {{
        font-size: 0.76rem; color: #6B7280; line-height: 1.45;
        margin-top: -0.2rem; margin-bottom: 0.7rem;
    }}

    .section-intro {{
        font-size: 0.88rem; color: #4B5563; margin-bottom: 1rem; line-height: 1.6; max-width: 820px;
    }}

    .context-block {{
        font-size: 0.9rem; color: {MCK_DARK}; line-height: 1.7; max-width: 820px;
    }}
    .context-block h4 {{
        color: {MCK_NAVY}; font-size: 1.05rem; font-weight: 600;
        margin-top: 1.5rem; margin-bottom: 0.5rem;
    }}
    .context-block h5 {{
        color: {MCK_BLUE}; font-size: 0.9rem; font-weight: 600;
        margin-top: 1.2rem; margin-bottom: 0.3rem;
    }}
    .context-block ul {{ margin-left: 1.2rem; }}
    .context-block li {{ margin-bottom: 0.3rem; }}
    .context-block strong {{ color: {MCK_NAVY}; }}

    .big-btn {{
        display: flex; justify-content: center; margin-top: 1.5rem; margin-bottom: 1rem;
    }}

    .stDataFrame table {{ font-size: 0.82rem; }}

    /* ── Model flow diagram ── */
    .flow-row {{
        display: flex; align-items: center; justify-content: center;
        gap: 0; margin: 1.2rem 0 1.5rem 0; flex-wrap: nowrap;
    }}
    .flow-box {{
        background: {MCK_WHITE}; border: 2px solid #E0E4E8; border-radius: 10px;
        padding: 0.8rem 1rem; text-align: center; min-width: 120px; max-width: 170px;
        flex-shrink: 0;
    }}
    .flow-box.highlight {{
        border-color: {MCK_NAVY}; background: #EBF4FA;
    }}
    .flow-box .flow-num {{
        display: inline-block; background: {MCK_NAVY}; color: {MCK_WHITE};
        font-size: 0.7rem; font-weight: 700; width: 20px; height: 20px;
        line-height: 20px; border-radius: 50%; text-align: center; margin-bottom: 0.3rem;
    }}
    .flow-box .flow-title {{
        font-size: 0.78rem; font-weight: 600; color: {MCK_NAVY};
        margin-bottom: 0.15rem;
    }}
    .flow-box .flow-desc {{
        font-size: 0.68rem; color: {MCK_GREY}; line-height: 1.3;
    }}
    .flow-arrow {{
        font-size: 1.3rem; color: {MCK_NAVY}; padding: 0 0.3rem; flex-shrink: 0;
    }}

    /* ── How-It-Works visual blocks ── */
    .hiw-step {{
        display: flex; gap: 1rem; align-items: flex-start;
        margin-bottom: 1.5rem; padding: 1.2rem; background: {MCK_WHITE};
        border: 1px solid #E0E4E8; border-radius: 10px;
    }}
    .hiw-step-num {{
        background: {MCK_NAVY}; color: {MCK_WHITE};
        font-size: 1rem; font-weight: 700; min-width: 36px; height: 36px;
        line-height: 36px; border-radius: 50%; text-align: center; flex-shrink: 0;
    }}
    .hiw-step-body {{ flex: 1; }}
    .hiw-step-body h5 {{
        margin: 0 0 0.4rem 0; font-size: 0.95rem; font-weight: 600; color: {MCK_NAVY};
        background: none; padding: 0; border-radius: 0; text-transform: none; letter-spacing: 0;
    }}
    .hiw-step-body p {{ margin: 0 0 0.4rem 0; font-size: 0.85rem; color: #374151; line-height: 1.55; }}
    .hiw-step-body .hiw-formula {{
        display: inline-block; background: #EBF4FA; padding: 0.35rem 0.7rem;
        border-radius: 6px; font-family: 'Courier New', monospace;
        font-size: 0.82rem; color: {MCK_NAVY}; font-weight: 600; margin: 0.3rem 0;
    }}
    .hiw-step-body ul {{ margin: 0.3rem 0 0 1.1rem; padding: 0; }}
    .hiw-step-body li {{ font-size: 0.85rem; color: #374151; margin-bottom: 0.2rem; line-height: 1.5; }}

    .hiw-concept-row {{
        display: flex; gap: 1rem; margin: 1rem 0 1.5rem 0; flex-wrap: wrap;
    }}
    .hiw-concept {{
        flex: 1; min-width: 170px; background: #EBF4FA; border-radius: 10px;
        padding: 1rem 1.1rem; text-align: left;
    }}
    .hiw-concept h6 {{
        margin: 0 0 0.3rem 0; font-size: 0.82rem; font-weight: 700;
        color: {MCK_NAVY}; text-transform: uppercase; letter-spacing: 0.02em;
    }}
    .hiw-concept p {{
        margin: 0; font-size: 0.8rem; color: #374151; line-height: 1.5;
    }}

    .hiw-callout {{
        background: #FFFBEB; border-left: 4px solid #F59E0B;
        padding: 0.9rem 1.1rem; border-radius: 0 8px 8px 0;
        margin: 1rem 0; font-size: 0.85rem; color: #92400E; line-height: 1.55;
    }}

    .hiw-mini-pipe {{
        display: flex; align-items: center; justify-content: flex-start;
        gap: 0; margin: 0.8rem 0;
    }}
    .hiw-pipe-stage {{
        background: {MCK_NAVY}; color: {MCK_WHITE}; padding: 0.5rem 0.9rem;
        border-radius: 8px; font-size: 0.78rem; font-weight: 600; text-align: center;
    }}
    .hiw-pipe-arrow {{
        font-size: 1rem; color: {MCK_NAVY}; padding: 0 0.4rem;
    }}
    .hiw-pipe-label {{
        font-size: 0.68rem; color: {MCK_GREY}; text-align: center; margin-top: 0.2rem;
    }}
</style>
""", unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# Session state
# ---------------------------------------------------------------------------
if "cfg" not in st.session_state or not hasattr(st.session_state.cfg, "pipeline_stages"):
    st.session_state.cfg = petronas_baseline()
if "page" not in st.session_state:
    st.session_state.page = "configure"


def _sync_archetypes(cfg: ModelConfig):
    """Ensure every archetype has entries for every pipeline stage."""
    for arch in cfg.archetypes:
        for sname in cfg.pipeline_stages:
            if sname not in arch.stages:
                arch.stages[sname] = StageParams(9, 8.0, 3.0, 2.0)
        extra = [k for k in arch.stages if k not in cfg.pipeline_stages]
        for k in extra:
            del arch.stages[k]
        ordered = {s: arch.stages[s] for s in cfg.pipeline_stages if s in arch.stages}
        arch.stages = ordered


# ---------------------------------------------------------------------------
# Header (both pages)
# ---------------------------------------------------------------------------
def _render_header(subtitle: str):
    st.markdown(f"""
    <div class="mck-header">
        <h1>FTE Baseload Planning Tool</h1>
        <p>{subtitle}</p>
    </div>
    """, unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════
# PAGE 1: CONFIGURE
# ═══════════════════════════════════════════════════════════════════════════
def _page_configure():
    cfg = st.session_state.cfg

    _render_header("Set your R&amp;D budget and project portfolio &mdash; the model calculates how many staff you need")

    mode = st.radio(
        "Input mode",
        ["Quick estimate", "Full configuration"],
        horizontal=True,
        help="**Quick estimate** — set budget and portfolio mix; all other assumptions use baseline values (shown below). "
             "**Full configuration** — define your own pipeline stages, project parameters, and staffing assumptions.",
        key="input_mode",
    )

    is_custom = mode == "Full configuration"

    # ── Visual: How the model works (compact flow) ──
    st.markdown("""
    <div class="flow-row">
        <div class="flow-box">
            <div class="flow-num">1</div>
            <div class="flow-title">R&D Budget</div>
            <div class="flow-desc">Total spend minus overhead</div>
        </div>
        <div class="flow-arrow">→</div>
        <div class="flow-box">
            <div class="flow-num">2</div>
            <div class="flow-title">÷ Cost per project</div>
            <div class="flow-desc">Weighted by project type mix</div>
        </div>
        <div class="flow-arrow">→</div>
        <div class="flow-box">
            <div class="flow-num">3</div>
            <div class="flow-title"># Projects / year</div>
            <div class="flow-desc">How many fit in the budget</div>
        </div>
        <div class="flow-arrow">→</div>
        <div class="flow-box">
            <div class="flow-num">4</div>
            <div class="flow-title">Pipeline stages</div>
            <div class="flow-desc">Projects flow through gates</div>
        </div>
        <div class="flow-arrow">→</div>
        <div class="flow-box">
            <div class="flow-num">5</div>
            <div class="flow-title">Total FTE needed</div>
            <div class="flow-desc">Active projects × staff each</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Row 1: Budget & Timeline | Pipeline (or baseline summary) ──
    col_left, col_right = st.columns(2)

    with col_left:
        st.markdown('<div class="card"><h5>Budget & timeline</h5>', unsafe_allow_html=True)

        cfg.total_budget_m = st.number_input(
            "Annual R&D budget (USD millions)", value=cfg.total_budget_m, min_value=1.0,
            step=10.0, format="%.0f",
            help="Total yearly R&D spend, before any deductions",
        )

        overhead_pct = st.slider(
            "Overhead deduction (%)", 0, 60,
            int(cfg.overhead_pct * 100), 5, "%d%%",
            help="Admin, facilities, and management costs — subtracted from budget before funding projects",
        )
        cfg.overhead_pct = overhead_pct / 100.0

        avail = cfg.total_budget_m * (1 - cfg.overhead_pct)
        st.markdown(f'<div class="help-text">Net project budget: <strong>{avail:,.0f} M</strong></div>', unsafe_allow_html=True)

        c1, c2 = st.columns(2)
        with c1:
            cfg.start_year = int(st.number_input("First year of new projects", value=cfg.start_year, step=1,
                                                  help="Year new project intake begins"))
        with c2:
            cfg.end_year = int(st.number_input("Last year of new projects", value=cfg.end_year, step=1,
                                                help="Last year new projects are started — projects already in progress continue beyond this"))
            if cfg.end_year <= cfg.start_year:
                cfg.end_year = cfg.start_year + 1

        st.markdown('</div>', unsafe_allow_html=True)

    with col_right:
        if is_custom:
            st.markdown('<div class="card"><h5>Project stages — the journey every project takes</h5>', unsafe_allow_html=True)
            st.markdown("""<div class="help-text">
                Projects go through stages from early research to late development.
                Define your stages below. <strong>"% start here"</strong> = of all new projects each year, what share enters at this stage?
                <strong>"% move to next"</strong> = of projects that finish this stage, what share continues to the next one?
            </div>""", unsafe_allow_html=True)

            stages_to_remove = None

            hc1, hc2, hc3, hc4 = st.columns([3, 2, 2, 1])
            with hc1:
                st.caption("Stage name")
            with hc2:
                st.caption("% start here")
            with hc3:
                st.caption("% move to next")
            with hc4:
                st.caption("")

            for si, sname in enumerate(cfg.pipeline_stages):
                sc1, sc2, sc3, sc4 = st.columns([3, 2, 2, 1])
                with sc1:
                    new_name = st.text_input("Name", value=sname, key=f"sn_{si}",
                                              label_visibility="collapsed")
                with sc2:
                    alloc = st.number_input(
                        "Start", value=int(cfg.stage_mix.get(sname, 0) * 100),
                        min_value=0, max_value=100, step=5, key=f"sa_{si}",
                        label_visibility="collapsed",
                    )
                with sc3:
                    is_terminal = si == len(cfg.pipeline_stages) - 1
                    if not is_terminal:
                        conv = st.number_input(
                            "Move", value=int(cfg.stage_conversion_rates.get(sname, 0) * 100),
                            min_value=0, max_value=100, step=5, key=f"sc_{si}",
                            label_visibility="collapsed",
                        )
                    else:
                        st.markdown("—")
                        conv = None
                with sc4:
                    if len(cfg.pipeline_stages) > 1:
                        if st.button("✕", key=f"sr_{si}"):
                            stages_to_remove = si

                if new_name != sname and new_name.strip():
                    old = sname
                    cfg.pipeline_stages[si] = new_name
                    if old in cfg.stage_mix:
                        cfg.stage_mix[new_name] = cfg.stage_mix.pop(old)
                    if old in cfg.stage_conversion_rates:
                        cfg.stage_conversion_rates[new_name] = cfg.stage_conversion_rates.pop(old)
                    for arch in cfg.archetypes:
                        if old in arch.stages:
                            arch.stages[new_name] = arch.stages.pop(old)
                    sname = new_name

                cfg.stage_mix[sname] = alloc / 100.0
                if conv is not None:
                    cfg.stage_conversion_rates[sname] = conv / 100.0

            if stages_to_remove is not None:
                removed = cfg.pipeline_stages.pop(stages_to_remove)
                cfg.stage_mix.pop(removed, None)
                cfg.stage_conversion_rates.pop(removed, None)
                for arch in cfg.archetypes:
                    arch.stages.pop(removed, None)
                st.rerun()

            if st.button("＋ Add stage"):
                new_s = f"Stage {len(cfg.pipeline_stages) + 1}"
                cfg.pipeline_stages.append(new_s)
                cfg.stage_mix[new_s] = 0.0
                if len(cfg.pipeline_stages) >= 2:
                    prev = cfg.pipeline_stages[-2]
                    cfg.stage_conversion_rates.setdefault(prev, 0.50)
                _sync_archetypes(cfg)
                st.rerun()

            alloc_sum = sum(cfg.stage_mix.get(s, 0) for s in cfg.pipeline_stages)
            if abs(alloc_sum - 1.0) > 0.01:
                st.warning(f"\"% start here\" values add up to {alloc_sum*100:.0f}% — they should total 100%")
            else:
                st.success("Stage percentages add up to 100%")

            st.markdown('</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="card"><h5>Project stages — the journey every project takes</h5>', unsafe_allow_html=True)

            for si, sname in enumerate(cfg.pipeline_stages):
                alloc = cfg.stage_mix.get(sname, 0) * 100
                conv = cfg.stage_conversion_rates.get(sname, 0)

                line = f"**{sname}**  \n"
                line += f"{alloc:.0f}% of new projects start at this stage each year"
                if si < len(cfg.pipeline_stages) - 1 and conv > 0:
                    next_name = cfg.pipeline_stages[si + 1]
                    line += f"  \n{conv*100:.0f}% of projects that finish here move on to {next_name}"
                elif si == len(cfg.pipeline_stages) - 1:
                    line += "  \nThis is the final stage — projects end here"
                st.markdown(line)

            st.divider()

            st.markdown(
                f"**New projects start during the first** {cfg.intake_spread_months} months of each year  \n"
                f"**Staff utilization:** {cfg.utilization_rate*100:.0f}% of time on project work"
                + (f"  \n**Ramp-up:** projects scale to full staffing over {cfg.ramp_months} months" if cfg.ramp_months > 0 else "")
            )
            st.caption("Switch to \"Full configuration\" to change these")
            st.markdown('</div>', unsafe_allow_html=True)

    # ── Row 2: Portfolio mix ──
    st.markdown('<div class="card"><h5>Portfolio mix — what types of R&D projects do you run?</h5>', unsafe_allow_html=True)
    st.markdown('<div class="help-text">Your R&D portfolio is a mix of different project types (e.g. chemistry, hardware, software). Use the sliders to set what percentage of your projects fall into each category. These must add up to 100%.</div>', unsafe_allow_html=True)

    arch_cols = st.columns(max(len(cfg.archetypes), 1))
    for ai, arch in enumerate(cfg.archetypes):
        with arch_cols[ai % len(arch_cols)]:
            share = st.slider(
                arch.name, 0, 100,
                int(arch.portfolio_share * 100), 5, "%d%%",
                key=f"ps_{ai}",
            )
            arch.portfolio_share = share / 100.0

    total_share = sum(a.portfolio_share for a in cfg.archetypes)
    if abs(total_share - 1.0) > 0.01 and cfg.archetypes:
        st.warning(f"Portfolio shares sum to {total_share*100:.0f}% — should be 100%")

    st.markdown('</div>', unsafe_allow_html=True)

    # ── Row 3 (Custom): Archetype details ──
    if is_custom:
        st.markdown('<div class="card"><h5>Project type details — how long, how much, how many people?</h5>', unsafe_allow_html=True)
        st.markdown('<div class="help-text">For each project type and stage, set your best estimate for duration, cost, and team size.</div>', unsafe_allow_html=True)

        _sync_archetypes(cfg)

        arch_tabs = st.tabs([a.name for a in cfg.archetypes] + ["＋ Add"])

        for ai, arch in enumerate(cfg.archetypes):
            with arch_tabs[ai]:
                arch.name = st.text_input("Name", value=arch.name, key=f"an_{ai}")

                for sname in cfg.pipeline_stages:
                    if sname not in arch.stages:
                        continue
                    sp = arch.stages[sname]
                    st.markdown(f"**{sname}**")

                    c1, c2, c3, c4 = st.columns(4)
                    with c1:
                        sp.duration_months = st.number_input(
                            "Duration (months)", value=sp.duration_months,
                            min_value=1, step=1, key=f"dm_{ai}_{sname}",
                        )
                    with c2:
                        sp.cost_millions = st.number_input(
                            "Cost per project (M)", value=sp.cost_millions,
                            min_value=0.1, step=0.5, key=f"cm_{ai}_{sname}",
                        )
                    with c3:
                        sp.fte_research = st.number_input(
                            "Researchers per project", value=sp.fte_research,
                            min_value=0.0, step=0.5, key=f"rm_{ai}_{sname}",
                        )
                    with c4:
                        sp.fte_developer = st.number_input(
                            "Developers per project", value=sp.fte_developer,
                            min_value=0.0, step=0.5, key=f"ddm_{ai}_{sname}",
                        )

                if len(cfg.archetypes) > 1:
                    if st.button(f"Remove {arch.name}", key=f"ra_{ai}"):
                        cfg.archetypes.pop(ai)
                        st.rerun()

        with arch_tabs[-1]:
            st.markdown("Click below to add a new archetype.")
            if st.button("Create archetype"):
                new_name = f"Type {len(cfg.archetypes) + 1}"
                new_stages = {s: StageParams(9, 8.0, 3.0, 2.0)
                              for s in cfg.pipeline_stages}
                cfg.archetypes.append(Archetype(name=new_name, portfolio_share=0.0, stages=new_stages))
                st.rerun()

        st.markdown('</div>', unsafe_allow_html=True)

        # Advanced settings
        st.markdown('<div class="card"><h5>Advanced settings</h5>', unsafe_allow_html=True)

        ac1, ac2, ac3 = st.columns(3)
        with ac1:
            util_pct = st.slider(
                "Utilization rate", 50, 100, int(cfg.utilization_rate * 100), 5, "%d%%",
                help="Fraction of time an FTE spends on project work. If 80%, gross headcount = model FTE ÷ 0.80",
            )
            cfg.utilization_rate = util_pct / 100.0
        with ac2:
            cfg.ramp_months = st.slider(
                "Ramp-up period (months)", 0, 6, cfg.ramp_months, 1,
                help="Projects ramp FTE linearly. Month 1 of a 3-month ramp = 33% staffing, month 2 = 67%, month 3+ = 100%",
            )
        with ac3:
            cfg.intake_spread_months = st.slider(
                "Intake window (months/year)", 1, 12, cfg.intake_spread_months, 1,
                help="New projects start evenly across the first N months of each year",
            )

        st.markdown('</div>', unsafe_allow_html=True)
    else:
        # Standard mode: show baseline archetype detail read-only
        st.markdown('<div class="card"><h5>Project details — how long, how much, how many people? (baseline values)</h5>', unsafe_allow_html=True)

        rows = []
        for arch in cfg.archetypes:
            for sname, sp in arch.stages.items():
                rows.append({
                    "Archetype": arch.name,
                    "Stage": sname,
                    "Duration": f"{sp.duration_months} mo",
                    "Cost / project": f"{sp.cost_millions:.1f} M",
                    "Research FTE": f"{sp.fte_research:.1f}",
                    "Developer FTE": f"{sp.fte_developer:.1f}",
                })
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
        st.caption("Switch to \"Full configuration\" to change these")
        st.markdown('</div>', unsafe_allow_html=True)

    with st.expander("Things the model assumes that cannot be changed"):
        st.markdown("""
| What the model assumes | In plain language |
|------------------------|-------------------|
| Monthly tracking | The model tracks projects and headcount month by month |
| Same team size throughout a project | Once a project starts, its team stays the same size (unless ramp-up is on) |
| Graduates add to the next stage | Projects that finish an early stage and move on are *added* to whatever's already in the next stage |
| Each year stands alone | This year's new projects don't depend on last year's results |
| No projects get cancelled midway | Once started, every project runs to the end of its stage |
| No bulk discounts | Running 50 projects doesn't make each one cheaper than running 5 |
| Same budget every year | The model uses the same annual budget for every year in the range |
| Two roles only | Researchers and Developers — no other role types |
| Projects start evenly | New projects are spread evenly across the intake months (not all at once) |
| One path through the pipeline | Projects go forward through stages in order — no skipping ahead or looping back |
""")
    # ── Generate button ──
    c1, c2, c3 = st.columns([1, 2, 1])
    with c2:
        if st.button("▶  Generate Model", use_container_width=True, type="primary"):
            st.session_state.page = "results"
            st.rerun()

    _, rc, _ = st.columns([1, 2, 1])
    with rc:
        if st.button("⟳  Reset to baseline defaults", use_container_width=True):
            st.session_state.cfg = petronas_baseline()
            st.rerun()


# ═══════════════════════════════════════════════════════════════════════════
# PAGE 2: RESULTS
# ═══════════════════════════════════════════════════════════════════════════
def _page_results():
    cfg = st.session_state.cfg
    result = run_model(cfg)

    # Header with back button
    hc1, hc2 = st.columns([3, 1])
    with hc1:
        _render_header(f"Results for {cfg.start_year}–{cfg.end_year} &nbsp;|&nbsp; Budget {cfg.total_budget_m:,.0f} M")
    with hc2:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("← Modify Assumptions", use_container_width=True):
            st.session_state.page = "configure"
            st.rerun()

    # KPI cards
    st.markdown(f"""
    <div class="kpi-row">
        <div class="kpi-card">
            <div class="kpi-label">Budget available for projects</div>
            <div class="kpi-value">{cfg.total_budget_m*(1-cfg.overhead_pct):,.0f} M</div>
            <div class="kpi-sub">{cfg.total_budget_m:,.0f} M total minus {cfg.overhead_pct*100:.0f}% overhead</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-label">New projects funded per year</div>
            <div class="kpi-value">{result.projects_per_year:,.0f}</div>
            <div class="kpi-sub">How many the budget can support annually</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-label">Yearly FTE range</div>
            <div class="kpi-value">{result.steady_state_min_month:,.0f} – {result.steady_state_max_month:,.0f}</div>
            <div class="kpi-sub">Min to max monthly FTE in {cfg.end_year} — narrows as pipeline stabilizes</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-label">Steady-state headcount</div>
            <div class="kpi-value">{result.steady_state_avg:,.0f}</div>
            <div class="kpi-sub">Avg monthly FTE in {cfg.end_year} — the level the pipeline settles at</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Tabs
    tab_dash, tab_how, tab_monthly, tab_annual, tab_assumptions = st.tabs([
        "Dashboard", "How It Works", "Monthly Detail", "Annual Summary", "Assumption Register",
    ])

    monthly = result.monthly

    # ── Dashboard ──
    with tab_dash:
        if monthly.empty:
            st.info("No data. Check that archetypes and shares are configured.")
        else:
            st.markdown("#### Average headcount needed by year")
            st.markdown('<div class="section-intro">Each bar shows the average monthly FTE for that year, with the range (min–max monthly FTE) shown as markers.</div>', unsafe_allow_html=True)

            ann = result.annual_summary
            if not ann.empty:
                fig_main = go.Figure()
                fig_main.add_trace(go.Bar(
                    x=ann["Year"], y=ann["Avg monthly FTE"], name="Avg monthly FTE",
                    marker_color=MCK_NAVY, opacity=0.85,
                ))
                fig_main.add_trace(go.Scatter(
                    x=ann["Year"], y=ann["Min monthly FTE"], name="Min month",
                    mode="markers", marker=dict(color=MCK_TEAL, size=8, symbol="triangle-down"),
                ))
                fig_main.add_trace(go.Scatter(
                    x=ann["Year"], y=ann["Max monthly FTE"], name="Max month",
                    mode="markers", marker=dict(color=MCK_BLUE, size=8, symbol="triangle-up"),
                ))
                fig_main.update_layout(
                    height=400,
                    margin=dict(l=20, r=20, t=30, b=20),
                    plot_bgcolor=MCK_WHITE, paper_bgcolor=MCK_WHITE,
                    font=dict(family="Inter", size=12, color=MCK_DARK),
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0),
                    xaxis=dict(gridcolor="#E8EAED", title="", dtick=1),
                    yaxis=dict(gridcolor="#E8EAED", title="Monthly FTE"),
                    hovermode="x unified",
                )
                st.plotly_chart(fig_main, use_container_width=True)

            st.markdown("#### Researcher vs Developer split by year")

            if not ann.empty:
                fig2 = go.Figure()
                fig2.add_trace(go.Bar(x=ann["Year"], y=ann["Avg Research FTE"], name="Research", marker_color=MCK_BLUE))
                fig2.add_trace(go.Bar(x=ann["Year"], y=ann["Avg Developer FTE"], name="Developer", marker_color=MCK_TEAL))
                fig2.update_layout(
                    barmode="stack", height=360,
                    margin=dict(l=20, r=20, t=30, b=20),
                    plot_bgcolor=MCK_WHITE, paper_bgcolor=MCK_WHITE,
                    font=dict(family="Inter", size=12, color=MCK_DARK),
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0),
                    xaxis=dict(gridcolor="#E8EAED", title="", dtick=1),
                    yaxis=dict(gridcolor="#E8EAED", title="Avg monthly FTE"),
                    hovermode="x unified",
                )
                st.plotly_chart(fig2, use_container_width=True)

            st.markdown("#### Project type detail")
            bk_rows = []
            for arch in cfg.archetypes:
                for sn, sp in arch.stages.items():
                    bk_rows.append({
                        "Archetype": arch.name, "Stage": sn,
                        "Share": f"{arch.portfolio_share*100:.0f}%",
                        "Duration": f"{sp.duration_months} mo",
                        "Cost / project": f"{sp.cost_millions:.1f} M",
                        "Research FTE": f"{sp.fte_research:.1f}",
                        "Developer FTE": f"{sp.fte_developer:.1f}",
                    })
            st.dataframe(pd.DataFrame(bk_rows), use_container_width=True, hide_index=True)

    # ── How It Works ──
    with tab_how:
        st.markdown("""
<div class="context-block">
<h4>What does this tool do?</h4>
<p style="font-size:1rem; line-height:1.7;">
It answers one question: <strong>"Given our R&D budget, how many people do we need to hire?"</strong><br>
You tell it how much money you have and what types of projects you run. It tells you how many researchers and developers you need on staff.
</p>
</div>
""", unsafe_allow_html=True)

        st.markdown("#### Before we start — a few key terms")
        st.markdown(f"""
<div class="hiw-concept-row">
    <div class="hiw-concept">
        <h6>FTE</h6>
        <p><strong>Full-Time Equivalent.</strong> One FTE = one person working full time for a year. 0.5 FTE = half a person's time. This model calculates how many FTEs your projects need.</p>
    </div>
    <div class="hiw-concept">
        <h6>TRL (Technology Readiness Level)</h6>
        <p>A 1-to-9 scale used worldwide to describe how mature a technology is. Higher TRL = closer to real-world deployment.</p>
    </div>
    <div class="hiw-concept">
        <h6>Pipeline</h6>
        <p>The sequence of stages a project goes through from start to finish. Think of it like a funnel: many projects start at the early stage, but only some advance to later stages.</p>
    </div>
</div>
<div class="hiw-concept-row">
    <div class="hiw-concept">
        <h6>Archetype</h6>
        <p>A type of R&D project. Different types cost different amounts, take different times, and need different teams. E.g. "Chemistry" projects vs "Software" projects.</p>
    </div>
    <div class="hiw-concept">
        <h6>Conversion rate</h6>
        <p>What percentage of projects finishing one stage go on to the next. If conversion from TRL 1–4 to TRL 5–7 is 50%, half the early-stage completers move forward.</p>
    </div>
    <div class="hiw-concept">
        <h6>Steady state</h6>
        <p>In the first few years, headcount grows because new projects pile up faster than old ones finish. Eventually, starts and completions balance out and headcount levels off. That stable level is the <strong>steady state</strong> — the long-run staffing requirement your hiring plan should target.</p>
    </div>
</div>
""", unsafe_allow_html=True)

        st.markdown("#### The pipeline — how projects flow")

        pipe_text_parts = []
        for si, sn in enumerate(cfg.pipeline_stages):
            alloc = cfg.stage_mix.get(sn, 0) * 100
            pipe_text_parts.append(f"**{sn}** ({alloc:.0f}% of new projects start here)")
            if si < len(cfg.pipeline_stages) - 1:
                conv = cfg.stage_conversion_rates.get(sn, 0) * 100
                pipe_text_parts.append(f"  →  *{conv:.0f}% advance*  →  ")
        st.markdown("".join(pipe_text_parts))

        st.markdown("""
<div class="context-block">
<p>Projects enter the pipeline at different stages. Some start early (e.g. TRL 1–4) and must pass a gate to advance. Others skip the early stage and enter directly at a later stage (e.g. TRL 5–7). The model tracks how many projects are active in each stage, every month.</p>
</div>
""", unsafe_allow_html=True)

        st.markdown("#### How the model calculates headcount — step by step")

        st.markdown(f"""
<div class="hiw-step">
    <div class="hiw-step-num">1</div>
    <div class="hiw-step-body">
        <h5>Start with the money</h5>
        <p>You have a total R&D budget. First, subtract overhead (admin, facilities, management). What's left is the money available to actually fund projects.</p>
        <div class="hiw-formula">Net project budget = Total budget × (1 − Overhead %)</div>
        <p>Example: {cfg.total_budget_m:,.0f}M total × (1 − {cfg.overhead_pct*100:.0f}%) = <strong>{cfg.total_budget_m*(1-cfg.overhead_pct):,.0f}M</strong> available for projects.</p>
    </div>
</div>

<div class="hiw-step">
    <div class="hiw-step-num">2</div>
    <div class="hiw-step-body">
        <h5>Figure out how many projects you can afford</h5>
        <p>Each project type has a cost. The model computes a <em>weighted average cost per project</em> based on your portfolio mix (how much of each type you run) and which pipeline stages they go through.</p>
        <div class="hiw-formula">Projects per year = Net budget ÷ Weighted avg cost per project</div>
    </div>
</div>

<div class="hiw-step">
    <div class="hiw-step-num">3</div>
    <div class="hiw-step-body">
        <h5>Distribute projects across types and stages</h5>
        <p>The total project count is split across your project types (archetypes) based on portfolio shares. Within each type, projects are assigned to pipeline stages:</p>
        <ul>
            <li><strong>Direct allocation</strong> — a fixed percentage of new projects start directly at each stage</li>
            <li><strong>Conversion</strong> — when early-stage projects finish, a percentage of them "graduate" to the next stage, creating additional projects there</li>
        </ul>
        <p>This means later stages get projects from two sources: direct allocation + graduates from the previous stage.</p>
    </div>
</div>

<div class="hiw-step">
    <div class="hiw-step-num">4</div>
    <div class="hiw-step-body">
        <h5>Simulate the pipeline month by month</h5>
        <p>Each year's new projects are spread across the first few months (the "intake window"). Once a project starts, it stays active for its full duration — say 12 months. The model tracks how many projects are running in each stage, every single month.</p>
        <p>Because projects from different years overlap (Year 1 projects may still be running when Year 2 projects start), headcount <strong>builds up</strong> over the first 2–3 years before levelling off.</p>
    </div>
</div>

<div class="hiw-step">
    <div class="hiw-step-num">5</div>
    <div class="hiw-step-body">
        <h5>Convert active projects into people needed</h5>
        <p>Every active project needs a team — some researchers and some developers. Multiply the number of active projects by the staff each project requires.</p>
        <div class="hiw-formula">FTE in a month = Active projects × Staff per project</div>
        <p>If utilization is less than 100% (people spend time on admin, training, leave), the model inflates the number to account for that. If ramp-up is set, new projects start with a partial team that grows to full strength over a few months.</p>
    </div>
</div>
""", unsafe_allow_html=True)

        st.markdown("#### Understanding steady state and the yearly range")
        st.markdown(f"""
<div class="context-block">
<p><strong>Why does headcount grow at first?</strong> In Year 1, projects start but none have finished yet — so the pipeline only fills up. In Year 2, new projects start while Year 1 projects are still running. Headcount keeps climbing until the rate of new starts roughly equals the rate of completions. Once that happens, headcount stabilizes — this is the <strong>steady state</strong>.</p>

<p><strong>Why is there a range within each year?</strong> Because new projects start during an intake window (not all at once), FTE demand varies month to month:</p>
<ul>
<li><strong>Steady-state headcount</strong> — the average monthly FTE in the last intake year ({cfg.end_year}). This is the long-run staffing level your hiring plan should target.</li>
<li><strong>Min monthly FTE</strong> — the quietest month (e.g. just before a new annual cohort starts)</li>
<li><strong>Max monthly FTE</strong> — the busiest month (e.g. when old and new cohorts overlap most)</li>
</ul>
<p>Early years have a wide range (pipeline is still filling). Later years have a narrower range (pipeline has stabilized). When min and max converge, you've reached steady state.</p>
</div>
""", unsafe_allow_html=True)

        st.markdown("#### What to do with the results")
        st.markdown(f"""
<div class="context-block">
<ul>
<li><strong>Use the steady-state headcount</strong> (the last KPI card) as the basis for long-term hiring plans — this is where the pipeline settles</li>
<li><strong>Use the yearly range</strong> to plan for seasonal variation in staffing needs</li>
<li><strong>Look at Research vs Developer split</strong> to decide which roles to prioritise</li>
<li><strong>Test sensitivity</strong> — go back, change one assumption (e.g. budget, portfolio mix), and regenerate to see how it moves the needle</li>
<li><strong>Download the Excel</strong> for offline review, presentations, or sharing with leadership</li>
</ul>
</div>
""", unsafe_allow_html=True)

    # ── Monthly Detail ──
    with tab_monthly:
        if monthly.empty:
            st.info("No data.")
        else:
            st.markdown("#### Month-by-month project load and FTE demand")
            st.markdown('<div class="section-intro">Each row shows one project type, one stage, one month. "Effective projects" = number of active projects, adjusted for ramp-up if set.</div>', unsafe_allow_html=True)

            disp = monthly.copy()
            disp["Month"] = disp["month"].dt.strftime("%Y-%m")
            disp["Archetype"] = disp["archetype"]
            disp["Stage"] = disp["stage"]
            disp["Active Projects"] = disp["effective_projects"].round(1)
            disp["Research FTE"] = disp["fte_research"].round(1)
            disp["Developer FTE"] = disp["fte_developer"].round(1)
            disp["Total FTE"] = disp["fte_total"].round(1)

            nice_cols = ["Month", "Archetype", "Stage",
                         "Active Projects", "Research FTE", "Developer FTE", "Total FTE"]

            # Filters
            fc1, fc2, fc3 = st.columns(3)
            with fc1:
                arch_options = sorted(disp["Archetype"].unique())
                sel_arch = st.multiselect("Archetype", arch_options, default=arch_options, key="f_arch")
            with fc2:
                stage_options = sorted(disp["Stage"].unique())
                sel_stage = st.multiselect("Stage", stage_options, default=stage_options, key="f_stage")
            with fc3:
                year_options = sorted(y for y in disp["month"].dt.year.unique() if cfg.start_year <= y <= cfg.end_year)
                sel_year = st.multiselect("Year", year_options, default=year_options, key="f_year")

            filtered = disp[
                disp["Archetype"].isin(sel_arch) &
                disp["Stage"].isin(sel_stage) &
                disp["month"].dt.year.isin(sel_year)
            ]

            st.dataframe(filtered[nice_cols], use_container_width=True, hide_index=True, height=500)
            st.download_button("Download CSV", filtered[nice_cols].to_csv(index=False),
                               "fte_monthly.csv", "text/csv")

    # ── Annual Summary ──
    with tab_annual:
        if result.annual_summary.empty:
            st.info("No data.")
        else:
            st.markdown("#### Average monthly FTE by year")
            st.markdown('<div class="section-intro">Avg = average across all months in the year. Min/Max = the lowest and highest single-month FTE that year (reflects pipeline build-up and seasonal intake variation).</div>', unsafe_allow_html=True)
            st.dataframe(result.annual_summary, use_container_width=True, hide_index=True)
            st.download_button("Download CSV", result.annual_summary.to_csv(index=False),
                               "fte_annual.csv", "text/csv")

    # ── Assumption Register ──
    with tab_assumptions:
        st.markdown("#### All assumptions used in this run")
        st.markdown('<div class="section-intro">Every number in the model traces back to one of these inputs.</div>', unsafe_allow_html=True)

        st.markdown('<div class="card"><h5>1. Budget</h5>', unsafe_allow_html=True)
        st.dataframe(pd.DataFrame([
            {"Assumption": "Total R&D budget", "Value": f"{cfg.total_budget_m:,.0f} M", "Type": "Input", "Meaning": "Gross annual R&D spend"},
            {"Assumption": "Overhead", "Value": f"{cfg.overhead_pct*100:.0f}%", "Type": "Input", "Meaning": "Admin, facilities, management"},
            {"Assumption": "Net project budget", "Value": f"{cfg.total_budget_m*(1-cfg.overhead_pct):,.0f} M", "Type": "Derived", "Meaning": "Total × (1 – Overhead)"},
        ]), use_container_width=True, hide_index=True)
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="card"><h5>2. Timeline</h5>', unsafe_allow_html=True)
        st.dataframe(pd.DataFrame([
            {"Assumption": "Start year", "Value": str(cfg.start_year), "Type": "Input"},
            {"Assumption": "End year", "Value": str(cfg.end_year), "Type": "Input"},
            {"Assumption": "Intake window", "Value": f"{cfg.intake_spread_months} months", "Type": "Input"},
        ]), use_container_width=True, hide_index=True)
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="card"><h5>3. Pipeline & funnel</h5>', unsafe_allow_html=True)
        pipe_rows = []
        for si, sn in enumerate(cfg.pipeline_stages):
            pipe_rows.append({
                "Stage": sn,
                "Direct allocation": f"{cfg.stage_mix.get(sn,0)*100:.0f}%",
                "Conversion to next": f"{cfg.stage_conversion_rates.get(sn,0)*100:.0f}%" if si < len(cfg.pipeline_stages) - 1 else "Terminal",
                "Type": "Input",
            })
        st.dataframe(pd.DataFrame(pipe_rows), use_container_width=True, hide_index=True)
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="card"><h5>4. Advanced</h5>', unsafe_allow_html=True)
        st.dataframe(pd.DataFrame([
            {"Assumption": "Utilization rate", "Value": f"{cfg.utilization_rate*100:.0f}%", "Type": "Input",
             "Meaning": f"Gross FTE = model FTE ÷ {cfg.utilization_rate:.2f}"},
            {"Assumption": "Ramp-up period", "Value": f"{cfg.ramp_months} months", "Type": "Input",
             "Meaning": "Linear ramp from 0 to full FTE" if cfg.ramp_months > 0 else "Full FTE from day 1"},
        ]), use_container_width=True, hide_index=True)
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="card"><h5>5. Project type parameters</h5>', unsafe_allow_html=True)
        arch_rows = []
        for arch in cfg.archetypes:
            for sn, sp in arch.stages.items():
                arch_rows.append({
                    "Archetype": arch.name, "Stage": sn,
                    "Share": f"{arch.portfolio_share*100:.0f}%",
                    "Duration (mo)": f"{sp.duration_months}",
                    "Cost (M)": f"{sp.cost_millions:.1f}",
                    "Research FTE": f"{sp.fte_research:.1f}",
                    "Developer FTE": f"{sp.fte_developer:.1f}",
                })
        st.dataframe(pd.DataFrame(arch_rows), use_container_width=True, hide_index=True)
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="card"><h5>6. Derived outputs</h5>', unsafe_allow_html=True)
        try:
            wc = _weighted_cost_per_project(cfg)
        except Exception:
            wc = 0
        st.dataframe(pd.DataFrame([
            {"Metric": "Weighted cost per project",
             "Value": f"{wc:,.1f} M",
             "How": "Portfolio-weighted expected cost across all stages"},
            {"Metric": "Projects per year",
             "Value": f"{result.projects_per_year:,.1f}",
             "How": f"Net budget ({cfg.total_budget_m*(1-cfg.overhead_pct):,.0f} M) ÷ cost per project ({wc:,.1f} M)"},
            {"Metric": f"Avg FTE in {cfg.end_year}",
             "Value": f"{result.steady_state_avg:,.0f}",
             "How": f"Average monthly total FTE in {cfg.end_year}"},
            {"Metric": f"FTE range in {cfg.end_year}",
             "Value": f"{result.steady_state_min_month:,.0f} – {result.steady_state_max_month:,.0f}",
             "How": f"Min to max monthly FTE in {cfg.end_year}"},
        ]), use_container_width=True, hide_index=True)
        st.markdown('</div>', unsafe_allow_html=True)

        with st.expander("7. Structural assumptions (built into the model)"):
            st.dataframe(pd.DataFrame([
                {"Assumption": "Monthly granularity", "Meaning": "Projects and FTE tracked monthly"},
                {"Assumption": "Constant FTE per project", "Meaning": "Flat staffing for full duration (unless ramp set)"},
                {"Assumption": "Additive conversion", "Meaning": "Stage survivors add to next-stage direct allocation"},
                {"Assumption": "Independent annual cohorts", "Meaning": "No carry-over between years"},
                {"Assumption": "No mid-stage failure", "Meaning": "Projects run to completion once started"},
                {"Assumption": "No economies of scale", "Meaning": "Cost per project is constant"},
                {"Assumption": "Constant annual budget", "Meaning": "Same budget every year"},
                {"Assumption": "Two FTE roles", "Meaning": "Research and Developer only"},
                {"Assumption": "Uniform intake spread", "Meaning": "Even distribution across intake window"},
                {"Assumption": "Linear pipeline", "Meaning": "No branching or looping between stages"},
            ]), use_container_width=True, hide_index=True)

    # ── Excel download ──
    st.divider()
    dc1, dc2, dc3 = st.columns([1, 2, 1])
    with dc2:
        excel_bytes = _generate_excel(cfg, result)
        st.download_button(
            "Download full model as Excel",
            data=excel_bytes,
            file_name="FTE_Baseload_Model.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


# ═══════════════════════════════════════════════════════════════════════════
# Excel generator
# ═══════════════════════════════════════════════════════════════════════════
def _generate_excel(cfg: ModelConfig, result) -> bytes:
    wb = Workbook()

    navy_fill = PatternFill(start_color="051C2C", end_color="051C2C", fill_type="solid")
    light_fill = PatternFill(start_color="F5F6F7", end_color="F5F6F7", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    hdr_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=16, bold=True, color="051C2C")
    sub_font = Font(name="Calibri", size=11, color="7F8C8D")
    sec_font = Font(name="Calibri", size=12, bold=True, color="051C2C")
    body_font = Font(name="Calibri", size=10, color="1A1A2E")
    blue_font = Font(name="Calibri", size=10, color="2251FF", bold=True)
    bdr = Border(
        left=Side(style="thin", color="D0D5DD"), right=Side(style="thin", color="D0D5DD"),
        top=Side(style="thin", color="D0D5DD"), bottom=Side(style="thin", color="D0D5DD"),
    )

    def _hdr_row(ws, row, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = navy_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = bdr

    def _data_row(ws, row, ncol, alt=False):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = light_fill if alt else white_fill
            cell.font = body_font
            cell.alignment = Alignment(vertical="center")
            cell.border = bdr

    # Cover
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_properties.tabColor = "00A9F4"
    ws.merge_cells("B3:F3")
    ws["B3"] = "FTE Baseload Model"
    ws["B3"].font = Font(name="Calibri", size=22, bold=True, color="051C2C")
    ws.merge_cells("B5:F5")
    ws["B5"] = "R&D Workforce Demand Planning"
    ws["B5"].font = sub_font
    info = [
        ("B8", "Budget", f"{cfg.total_budget_m:,.0f} M total, {cfg.total_budget_m*(1-cfg.overhead_pct):,.0f} M net"),
        ("B9", "Period", f"{cfg.start_year}–{cfg.end_year}"),
        ("B10", "Archetypes", ", ".join(a.name for a in cfg.archetypes)),
        ("B11", "Avg monthly FTE", f"{result.steady_state_avg:,.0f}"),
        ("B12", "FTE range", f"{result.steady_state_min_month:,.0f} – {result.steady_state_max_month:,.0f}"),
        ("B13", "Projects/year", f"{result.projects_per_year:,.0f}"),
    ]
    for ref, lbl, val in info:
        ws[ref] = lbl
        ws[ref].font = Font(name="Calibri", size=10, bold=True, color="051C2C")
        ws[ref.replace("B", "C")].value = val
        ws[ref.replace("B", "C")].font = body_font
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 60

    # Assumptions
    ws_a = wb.create_sheet("Assumptions")
    ws_a.sheet_properties.tabColor = "051C2C"
    ws_a["B2"] = "Model Assumptions"
    ws_a["B2"].font = title_font
    ws_a["B3"] = "Blue = user input. Black = derived."
    ws_a["B3"].font = sub_font

    row = 5
    sections = [
        ("BUDGET", [
            ("Total R&D budget", f"{cfg.total_budget_m:,.0f} M", True),
            ("Overhead", f"{cfg.overhead_pct*100:.0f}%", True),
            ("Net project budget", f"{cfg.total_budget_m*(1-cfg.overhead_pct):,.0f} M", False),
        ]),
        ("TIMELINE", [
            ("Start year", str(cfg.start_year), True),
            ("End year", str(cfg.end_year), True),
            ("Intake window", f"{cfg.intake_spread_months} months", True),
        ]),
        ("PIPELINE", [(sn, f"{cfg.stage_mix.get(sn,0)*100:.0f}% alloc, "
                           f"{cfg.stage_conversion_rates.get(sn,0)*100:.0f}% conv"
                           if si < len(cfg.pipeline_stages)-1 else
                           f"{cfg.stage_mix.get(sn,0)*100:.0f}% alloc (terminal)", True)
                       for si, sn in enumerate(cfg.pipeline_stages)]),
        ("ADVANCED", [
            ("Utilization", f"{cfg.utilization_rate*100:.0f}%", True),
            ("Ramp-up", f"{cfg.ramp_months} months", True),
        ]),
    ]
    for sec_name, items in sections:
        ws_a[f"B{row}"] = sec_name
        ws_a[f"B{row}"].font = sec_font
        row += 1
        for ci, h in enumerate(["Assumption", "Value", "Type"], 2):
            ws_a.cell(row=row, column=ci, value=h)
        _hdr_row(ws_a, row, 4)
        row += 1
        for lbl, val, is_input in items:
            ws_a.cell(row=row, column=2, value=lbl).font = body_font
            ws_a.cell(row=row, column=3, value=val).font = blue_font if is_input else body_font
            ws_a.cell(row=row, column=4, value="Input" if is_input else "Derived").font = body_font
            _data_row(ws_a, row, 4, alt=(row % 2 == 0))
            row += 1
        row += 1

    ws_a[f"B{row}"] = "ARCHETYPE PARAMETERS"
    ws_a[f"B{row}"].font = sec_font
    row += 1
    for ci, h in enumerate(["Archetype", "Stage", "Share", "Duration", "Cost", "Research", "Developer"], 2):
        ws_a.cell(row=row, column=ci, value=h)
    _hdr_row(ws_a, row, 8)
    row += 1
    for arch in cfg.archetypes:
        for sn, sp in arch.stages.items():
            ws_a.cell(row=row, column=2, value=arch.name).font = body_font
            ws_a.cell(row=row, column=3, value=sn).font = body_font
            ws_a.cell(row=row, column=4, value=f"{arch.portfolio_share*100:.0f}%").font = blue_font
            ws_a.cell(row=row, column=5, value=f"{sp.duration_months} mo").font = blue_font
            ws_a.cell(row=row, column=6, value=f"{sp.cost_millions:.1f} M").font = blue_font
            ws_a.cell(row=row, column=7, value=f"{sp.fte_research:.1f}").font = blue_font
            ws_a.cell(row=row, column=8, value=f"{sp.fte_developer:.1f}").font = blue_font
            _data_row(ws_a, row, 8, alt=(row % 2 == 0))
            row += 1

    for c_letter in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws_a.column_dimensions[c_letter].width = 20
    ws_a.column_dimensions["A"].width = 3

    # Annual Summary
    ws_ann = wb.create_sheet("Annual Summary")
    ws_ann.sheet_properties.tabColor = "051C2C"
    ws_ann["B2"] = "Annual FTE Summary"
    ws_ann["B2"].font = title_font
    ann_df = result.annual_summary
    if not ann_df.empty:
        row = 4
        cols = list(ann_df.columns)
        for ci, h in enumerate(cols, 2):
            ws_ann.cell(row=row, column=ci, value=h)
        _hdr_row(ws_ann, row, len(cols) + 1)
        row += 1
        for _, dr in ann_df.iterrows():
            for ci, col in enumerate(cols, 2):
                ws_ann.cell(row=row, column=ci, value=dr[col]).font = body_font
            _data_row(ws_ann, row, len(cols) + 1, alt=(row % 2 == 0))
            row += 1
    ws_ann.column_dimensions["A"].width = 3
    for ci in range(2, 9):
        ws_ann.column_dimensions[get_column_letter(ci)].width = 18

    # Monthly Detail
    ws_m = wb.create_sheet("Monthly Detail")
    ws_m.sheet_properties.tabColor = "7F8C8D"
    ws_m["B2"] = "Monthly FTE Detail"
    ws_m["B2"].font = title_font
    mon = result.monthly.copy()
    if not mon.empty:
        mon["month_str"] = mon["month"].dt.strftime("%Y-%m")
        exp_cols = ["month_str", "archetype", "stage",
                    "effective_projects", "fte_research", "fte_developer", "fte_total"]
        nice = ["Month", "Archetype", "Stage", "Effective Projects",
                "Research FTE", "Developer FTE", "Total FTE"]
        for c in ["effective_projects", "fte_research", "fte_developer", "fte_total"]:
            mon[c] = mon[c].round(2)
        row = 4
        for ci, h in enumerate(nice, 2):
            ws_m.cell(row=row, column=ci, value=h)
        _hdr_row(ws_m, row, len(nice) + 1)
        row += 1
        for _, dr in mon[exp_cols].iterrows():
            for ci, col in enumerate(exp_cols, 2):
                ws_m.cell(row=row, column=ci, value=dr[col]).font = body_font
            _data_row(ws_m, row, len(nice) + 1, alt=(row % 2 == 0))
            row += 1
    ws_m.column_dimensions["A"].width = 3
    for ci in range(2, 10):
        ws_m.column_dimensions[get_column_letter(ci)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════
# Router
# ═══════════════════════════════════════════════════════════════════════════
if st.session_state.page == "configure":
    _page_configure()
else:
    _page_results()
